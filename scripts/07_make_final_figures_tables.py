# Optional exact-format manuscript figure section 0
try:
    
    import pandas as pd
    import matplotlib.pyplot as plt
    from matplotlib.ticker import MultipleLocator, FormatStrFormatter
    from pathlib import Path
    
    # -----------------------------
    # Input files
    # -----------------------------
    interaction_file = O4_DIR / "O5_SHAP_interaction_values.csv"
    correlation_file = O4_DIR / "O5_SHAP_complementarity_correlation.csv"
    
    interaction = pd.read_csv(interaction_file)
    correlation = pd.read_csv(correlation_file)
    
    # -----------------------------
    # Build feature-pair labels
    # -----------------------------
    interaction["Feature pair"] = (
        interaction["SE_Feature"].astype(str) + " x " + interaction["Other_Feature"].astype(str)
    )
    
    correlation["Feature pair"] = (
        correlation["SE_Feature"].astype(str) + " x " + correlation["Other_Feature"].astype(str)
    )
    
    # -----------------------------
    # Sort and select top pairs
    # -----------------------------
    top_n = 40
    
    left_df = interaction.sort_values(
        "Mean_abs_SHAP_Interaction",
        ascending=False
    ).head(top_n)
    
    right_df = correlation.sort_values(
        "Abs_SHAP_Correlation",
        ascending=False
    ).head(top_n)
    
    # -----------------------------
    # A4 landscape settings
    # -----------------------------
    A4_LANDSCAPE = (11.69, 8.27)  # inches
    
    y_label_fontsize = 6.5
    axis_label_fontsize = 8
    title_fontsize = 9
    tick_fontsize = 7
    panel_label_fontsize = 11
    
    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 7,
        "axes.titlesize": title_fontsize,
        "axes.labelsize": axis_label_fontsize,
        "xtick.labelsize": tick_fontsize,
        "ytick.labelsize": y_label_fontsize,
        "axes.linewidth": 0.8,
    })
    
    bar_color = "#cd9778"
    
    fig, axes = plt.subplots(
        1, 2,
        figsize=A4_LANDSCAPE,
        dpi=300
    )
    
    # -----------------------------
    # Helper function
    # -----------------------------
    def make_panel(ax, df, value_col, title, xlabel, xlim, tick_step, tick_format):
        y = range(len(df))
    
        ax.barh(
            y,
            df[value_col],
            color=bar_color,
            edgecolor=bar_color,
            height=0.68
        )
    
        ax.set_yticks(y)
        ax.set_yticklabels(
            df["Feature pair"],
            fontsize=y_label_fontsize
        )
    
        ax.invert_yaxis()
    
        ax.set_title(title, fontsize=title_fontsize, pad=8)
        ax.set_xlabel(xlabel, fontsize=axis_label_fontsize)
        ax.set_ylabel("Feature pair", fontsize=axis_label_fontsize, labelpad=8)
    
        ax.set_xlim(xlim)
        ax.xaxis.set_major_locator(MultipleLocator(tick_step))
        ax.xaxis.set_major_formatter(FormatStrFormatter(tick_format))
    
        ax.xaxis.grid(True, color="#e6e6e6", linewidth=0.6)
        ax.set_axisbelow(True)
    
        ax.tick_params(axis="y", length=3, pad=2, labelsize=y_label_fontsize)
        ax.tick_params(axis="x", length=3, labelsize=tick_fontsize)
    
    # -----------------------------
    # Automatic x-axis limits
    # Extra space prevents longest bars from touching the border
    # -----------------------------
    left_xmax = left_df["Mean_abs_SHAP_Interaction"].max() * 1.10
    right_xmax = right_df["Abs_SHAP_Correlation"].max() * 1.08
    
    # -----------------------------
    # Panel A
    # -----------------------------
    make_panel(
        axes[0],
        left_df,
        "Mean_abs_SHAP_Interaction",
        "SHAP interaction strength: SE with CHM/MSI",
        "Mean absolute SHAP interaction value",
        xlim=(0, left_xmax),
        tick_step=0.004,
        tick_format="%.3f"
    )
    
    # -----------------------------
    # Panel B
    # -----------------------------
    make_panel(
        axes[1],
        right_df,
        "Abs_SHAP_Correlation",
        "Complementarity between SE and CHM/MSI predictors",
        "Absolute correlation between SHAP values",
        xlim=(0, right_xmax),
        tick_step=0.1,
        tick_format="%.1f"
    )
    
    # -----------------------------
    # Panel labels
    # -----------------------------
    axes[0].text(
        -0.08, 1.025, "(a)",
        transform=axes[0].transAxes,
        fontsize=panel_label_fontsize,
        fontweight="bold"
    )
    
    axes[1].text(
        -0.08, 1.025, "(b)",
        transform=axes[1].transAxes,
        fontsize=panel_label_fontsize,
        fontweight="bold"
    )
    
    # -----------------------------
    # Layout for A4 landscape
    # Increased left margin helps long y-axis labels fit
    # Reduced right margin gives bars enough visual space
    # -----------------------------
    plt.subplots_adjust(
        left=0.18,
        right=0.97,
        top=0.935,
        bottom=0.085,
        wspace=0.65
    )
    
    # -----------------------------
    # Save figure
    # Do not use bbox_inches="tight" because it changes exact A4 canvas size
    # -----------------------------
    plt.savefig(
        O4_DIR / "SHAP_interaction_complementarity_A4_landscape.png",
        dpi=300
    )
    
    plt.savefig(
        O4_DIR / "SHAP_interaction_complementarity_A4_landscape.pdf"
    )
    
    plt.close("all")
    
except Exception as exc:
    print('Warning: skipped optional exact-format figure section 0:', exc)
    traceback.print_exc(limit=1)


# Optional exact-format manuscript figure section 1
try:

    import pandas as pd
    import matplotlib.pyplot as plt
    from matplotlib.patches import Patch
    from matplotlib.ticker import MultipleLocator, FormatStrFormatter
    from pathlib import Path

    # ============================================================
    # Algorithm-specific SHAP source files
    # ============================================================
    DATA_DIR = OUT_DIR / "O4_SHAP_and_complementarity" / "O4_algorithm_specific_SHAP"

    print("Using algorithm-specific SHAP data folder:")
    print(DATA_DIR)

    all_file = DATA_DIR / "All_algorithm_specific_SHAP_predictor_importance.csv"
    overall_file = DATA_DIR / "Overall_mean_SHAP_importance_across_algorithms.csv"

    if not all_file.exists():
        raise FileNotFoundError(f"File not found:\n{all_file}")
    if not overall_file.exists():
        raise FileNotFoundError(f"File not found:\n{overall_file}")

    all_df = pd.read_csv(all_file)
    overall_df = pd.read_csv(overall_file)

    if "Mean_abs_SHAP" not in all_df.columns:
        raise ValueError(f"Expected column 'Mean_abs_SHAP' in {all_file}")
    if "Mean_abs_SHAP_overall" not in overall_df.columns:
        raise ValueError(f"Expected column 'Mean_abs_SHAP_overall' in {overall_file}")

    # ============================================================
    # Clean group names
    # ============================================================
    def clean_group(group):
        if str(group) in ["RGB-NIR", "RGBNIR", "MSI", "RGB_NIR"]:
            return "MSI"
        return str(group)

    all_df["Plot_Group"] = all_df["Group"].apply(clean_group)
    overall_df["Plot_Group"] = overall_df["Group"].apply(clean_group)

    if "Model_plot" not in all_df.columns:
        all_df["Model_plot"] = all_df["Model"].replace({"GradientBoosting": "GB"})

    # ============================================================
    # Colors same style as manuscript figure
    # ============================================================
    group_colors = {
        "CHM": "#6aaed6",
        "MSI": "#e7ad72",
        "SE": "#7bd88f"
    }

    # ============================================================
    # Panel data
    # ============================================================
    overall_top = (
        overall_df
        .sort_values("Mean_abs_SHAP_overall", ascending=False)
        .head(20)
    )

    rf_top = (
        all_df[all_df["Model_plot"] == "RF"]
        .sort_values("Mean_abs_SHAP", ascending=False)
        .head(10)
    )

    gb_top = (
        all_df[all_df["Model_plot"] == "GB"]
        .sort_values("Mean_abs_SHAP", ascending=False)
        .head(10)
    )

    xgb_top = (
        all_df[all_df["Model_plot"] == "XGBoost"]
        .sort_values("Mean_abs_SHAP", ascending=False)
        .head(10)
    )

    # ============================================================
    # Plot settings
    # ============================================================
    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 8,
        "axes.titlesize": 10,
        "axes.labelsize": 8,
        "xtick.labelsize": 8,
        "ytick.labelsize": 8,
        "axes.linewidth": 0.9,
    })

    fig, axes = plt.subplots(
        2, 2,
        figsize=(11.69, 8.27),
        dpi=300
    )

    axes = axes.flatten()

    # ============================================================
    # Helper function
    # ============================================================
    def plot_barh(ax, df, value_col, title, panel_label, tick_step=0.02):
        df_plot = df.sort_values(value_col, ascending=True).copy()
        colors = df_plot["Plot_Group"].map(group_colors).fillna("#bdbdbd")
        xmax = float(df[value_col].max()) * 1.10 if len(df) else 0.1
        xmax = max(xmax, 0.01)

        ax.barh(
            df_plot["Predictor"],
            df_plot[value_col],
            color=colors,
            edgecolor=colors,
            height=0.70
        )

        ax.set_title(title, fontweight="bold", pad=6)
        ax.set_xlabel("Mean |SHAP value|")
        ax.set_xlim(0, xmax)
        ax.xaxis.set_major_locator(MultipleLocator(tick_step))
        ax.xaxis.set_major_formatter(FormatStrFormatter("%.2f"))

        ax.grid(
            axis="x",
            linestyle="--",
            linewidth=0.6,
            alpha=0.35
        )
        ax.set_axisbelow(True)
        ax.tick_params(axis="y", labelsize=8)
        ax.tick_params(axis="x", labelsize=8)

        ax.text(
            0.0, 1.04,
            panel_label,
            transform=ax.transAxes,
            fontsize=10,
            fontweight="bold",
            ha="left",
            va="bottom"
        )

    # ============================================================
    # Draw panels
    # ============================================================
    plot_barh(axes[0], overall_top, "Mean_abs_SHAP_overall", "Overall Top 20 Predictors", "(a)")
    plot_barh(axes[1], rf_top, "Mean_abs_SHAP", "RF | Top 10 Predictors", "(b)")
    plot_barh(axes[2], gb_top, "Mean_abs_SHAP", "GB | Top 10 Predictors", "(c)")
    plot_barh(axes[3], xgb_top, "Mean_abs_SHAP", "XGBoost | Top 10 Predictors", "(d)")

    # ============================================================
    # Legend
    # ============================================================
    legend_handles = [
        Patch(facecolor=group_colors["CHM"], edgecolor="black", label="CHM"),
        Patch(facecolor=group_colors["MSI"], edgecolor="black", label="MSI"),
        Patch(facecolor=group_colors["SE"], edgecolor="black", label="SE"),
    ]

    fig.legend(
        handles=legend_handles,
        loc="lower center",
        ncol=3,
        frameon=False,
        bbox_to_anchor=(0.5, 0.035),
        fontsize=8
    )

    # ============================================================
    # Layout
    # ============================================================
    plt.subplots_adjust(
        left=0.13,
        right=0.985,
        top=0.94,
        bottom=0.12,
        wspace=0.30,
        hspace=0.26
    )

    # ============================================================
    # Save outputs in the same folder
    # ============================================================
    out_png = DATA_DIR / "Top_predictors_A4_exact_style.png"
    out_pdf = DATA_DIR / "Top_predictors_A4_exact_style.pdf"

    plt.savefig(out_png, dpi=300)
    plt.savefig(out_pdf)

    plt.close("all")

    print("Saved PNG to:", out_png)
    print("Saved PDF to:", out_pdf)

except Exception as exc:
    print('Warning: skipped optional exact-format figure section 1:', exc)
    traceback.print_exc(limit=1)


# Optional exact-format manuscript figure section 2
try:
    
    import pandas as pd
    import matplotlib.pyplot as plt
    from pathlib import Path
    
    # ============================================================
    # Your data folder
    # ============================================================
    DATA_DIR = OUT_DIR / "O3_label_efficiency"
    
    summary_file = DATA_DIR / "O3_label_efficiency_summary.csv"
    
    if not summary_file.exists():
        raise FileNotFoundError(
            f"File not found:\n{summary_file}\n\n"
            "Please check that O3_label_efficiency_summary.csv is inside this folder."
        )
    
    df = pd.read_csv(summary_file)
    
    # ============================================================
    # Rename feature-set labels for the figure
    # ============================================================
    feature_label_map = {
        "CHM+RGBNIR": "CHM+MSI",
        "CHM+RGBNIR+SE": "CHM+MSI+SE"
    }
    
    df["Feature_Set_Plot"] = df["Feature_Set"].replace(feature_label_map)
    
    # ============================================================
    # Model order and display names
    # ============================================================
    model_order = ["RF", "GradientBoosting", "XGBoost"]
    
    model_title_map = {
        "RF": "RF",
        "GradientBoosting": "GB",
        "XGBoost": "XGBoost"
    }
    
    feature_order = ["CHM+MSI", "CHM+MSI+SE"]
    
    # ============================================================
    # Colors matching the shown figure
    # ============================================================
    colors = {
        "CHM+MSI": "#1f77b4",
        "CHM+MSI+SE": "#ff7f0e"
    }
    
    # ============================================================
    # Metrics and axis limits
    # ============================================================
    metrics = [
        {
            "mean_col": "R2_mean",
            "sd_col": "R2_sd",
            "ylabel": "R²",
            "ylim": (-0.55, 0.70),
            "yticks": [-0.4, -0.2, 0.0, 0.2, 0.4, 0.6]
        },
        {
            "mean_col": "RMSE_mean",
            "sd_col": "RMSE_sd",
            "ylabel": "RMSE",
            "ylim": (0.11, 0.24),
            "yticks": [0.12, 0.14, 0.16, 0.18, 0.20, 0.22, 0.24]
        },
        {
            "mean_col": "MAE_mean",
            "sd_col": "MAE_sd",
            "ylabel": "MAE",
            "ylim": (0.08, 0.16),
            "yticks": [0.08, 0.09, 0.10, 0.11, 0.12, 0.13, 0.14, 0.15, 0.16]
        }
    ]
    
    # ============================================================
    # Figure style
    # ============================================================
    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 8,
        "axes.titlesize": 9,
        "axes.labelsize": 8,
        "xtick.labelsize": 8,
        "ytick.labelsize": 8,
        "axes.linewidth": 0.9,
    })
    
    # A4 landscape
    fig, axes = plt.subplots(
        3, 3,
        figsize=(11.69, 8.27),
        dpi=300
    )
    
    panel_letters = [
        "(a)", "(b)", "(c)",
        "(d)", "(e)", "(f)",
        "(g)", "(h)", "(i)"
    ]
    
    # ============================================================
    # Plot panels
    # ============================================================
    panel_i = 0
    
    for row_i, metric in enumerate(metrics):
        for col_i, model in enumerate(model_order):
    
            ax = axes[row_i, col_i]
    
            model_df = df[df["Model"] == model].copy()
    
            for feature_set in feature_order:
                plot_df = (
                    model_df[model_df["Feature_Set_Plot"] == feature_set]
                    .sort_values("Train_Proportion")
                    .copy()
                )
    
                x = plot_df["Train_Proportion"].to_numpy() * 100
                y = plot_df[metric["mean_col"]].to_numpy()
                sd = plot_df[metric["sd_col"]].fillna(0).to_numpy()
    
                ax.plot(
                    x,
                    y,
                    marker="o",
                    markersize=4.5,
                    linewidth=1.8,
                    color=colors[feature_set],
                    label=feature_set
                )
    
                ax.fill_between(
                    x,
                    y - sd,
                    y + sd,
                    color=colors[feature_set],
                    alpha=0.16,
                    linewidth=0
                )
    
            # Column titles only on first row
            if row_i == 0:
                ax.set_title(
                    model_title_map[model],
                    fontweight="bold",
                    pad=8
                )
    
            # Y-axis labels only on first column
            if col_i == 0:
                ax.set_ylabel(metric["ylabel"])
            else:
                ax.set_ylabel("")
    
            # X-axis labels only on bottom row
            if row_i == 2:
                ax.set_xlabel("Training data used (%)")
            else:
                ax.set_xticklabels([])
    
            ax.set_xlim(17, 103)
            ax.set_xticks([20, 40, 60, 80, 100])
    
            ax.set_ylim(metric["ylim"])
            ax.set_yticks(metric["yticks"])
    
            ax.grid(
                True,
                axis="both",
                color="#d9d9d9",
                linewidth=0.6,
                alpha=0.55
            )
    
            ax.set_axisbelow(True)
    
            # Panel label
            ax.text(
                0.02,
                1.04,
                panel_letters[panel_i],
                transform=ax.transAxes,
                fontsize=10,
                fontweight="bold",
                ha="left",
                va="bottom"
            )
    
            panel_i += 1
    
    # ============================================================
    # Legend
    # ============================================================
    handles, labels = axes[0, 0].get_legend_handles_labels()
    
    fig.legend(
        handles,
        labels,
        loc="upper center",
        bbox_to_anchor=(0.5, 0.985),
        ncol=2,
        frameon=False,
        fontsize=8
    )
    
    # ============================================================
    # Bottom note
    # ============================================================
    fig.text(
        0.5,
        0.035,
        "Lines show mean performance across training proportions; shaded bands indicate ±1 SD.",
        ha="center",
        va="center",
        fontsize=8
    )
    
    # ============================================================
    # A4 layout
    # Do not use bbox_inches='tight' because it changes A4 size
    # ============================================================
    plt.subplots_adjust(
        left=0.075,
        right=0.985,
        top=0.92,
        bottom=0.10,
        wspace=0.18,
        hspace=0.22
    )
    
    # ============================================================
    # Save outputs
    # ============================================================
    out_png = DATA_DIR / "O3_label_efficiency_3x3_A4.png"
    out_pdf = DATA_DIR / "O3_label_efficiency_3x3_A4.pdf"
    
    plt.savefig(out_png, dpi=300)
    plt.savefig(out_pdf)
    
    plt.close("all")
    
    print("Saved PNG to:", out_png)
    print("Saved PDF to:", out_pdf)
    
except Exception as exc:
    print('Warning: skipped optional exact-format figure section 2:', exc)
    traceback.print_exc(limit=1)


# Optional exact-format manuscript figure section 3
try:
    
    import pandas as pd
    import numpy as np
    import matplotlib.pyplot as plt
    from pathlib import Path
    
    # ============================================================
    # Data folder
    # Change this to your O2 folder if needed
    # ============================================================
    DATA_DIR = OUT_DIR / "O2_SE_added_value"
    
    # Main input file
    summary_file = DATA_DIR / "O2_SE_added_value_summary.csv"
    
    # If CSV is not found, try Excel file
    excel_file = DATA_DIR / "O2_SE_added_value_results.xlsx"
    
    if summary_file.exists():
        df = pd.read_csv(summary_file)
    elif excel_file.exists():
        df = pd.read_excel(excel_file, sheet_name="Summary")
    else:
        raise FileNotFoundError(
            f"Could not find:\n{summary_file}\n\nor:\n{excel_file}"
        )
    
    # ============================================================
    # Label cleaning
    # ============================================================
    def clean_feature_name(name):
        name = str(name)
        name = name.replace("RGBNIR", "MSI")
        name = name.replace("+", " + ")
        return name
    
    def clean_model_name(model):
        if model == "GradientBoosting":
            return "GB"
        return model
    
    df["With_SE_clean"] = df["With_SE"].apply(clean_feature_name)
    df["Model_clean"] = df["Model"].apply(clean_model_name)
    
    # ============================================================
    # Plot order
    # ============================================================
    row_order = [
        "CHM + SE",
        "MSI + SE",
        "CHM + MSI + SE"
    ]
    
    col_order = [
        "RF",
        "GB",
        "XGBoost"
    ]
    
    # ============================================================
    # Improvement calculations
    # ============================================================
    # R² improvement is calculated as relative improvement:
    # 100 * (R2_with_SE - R2_without_SE) / R2_without_SE
    df["R2_improvement_percent"] = (
        100 * (df["R2_with_SE_mean"] - df["R2_without_SE_mean"]) / df["R2_without_SE_mean"]
    )
    
    # RMSE and MAE improvement columns already exist in the summary file
    df["RMSE_improvement_percent"] = df["RMSE_percent_improvement_mean"]
    df["MAE_improvement_percent"] = df["MAE_percent_improvement_mean"]
    
    # ============================================================
    # Significance labels
    # ============================================================
    def p_to_stars(p):
        if p < 0.001:
            return "***"
        elif p < 0.01:
            return "**"
        elif p < 0.05:
            return "*"
        else:
            return "ns"
    
    # ============================================================
    # Build matrix for each metric
    # ============================================================
    def build_matrix(value_col, p_col):
        values = pd.DataFrame(index=row_order, columns=col_order, dtype=float)
        stars = pd.DataFrame(index=row_order, columns=col_order, dtype=object)
    
        for _, row in df.iterrows():
            r = row["With_SE_clean"]
            c = row["Model_clean"]
    
            if r in row_order and c in col_order:
                values.loc[r, c] = row[value_col]
                stars.loc[r, c] = p_to_stars(row[p_col])
    
        return values, stars
    
    r2_values, r2_stars = build_matrix(
        "R2_improvement_percent",
        "Wilcoxon_p_R2"
    )
    
    rmse_values, rmse_stars = build_matrix(
        "RMSE_improvement_percent",
        "Wilcoxon_p_RMSE"
    )
    
    mae_values, mae_stars = build_matrix(
        "MAE_improvement_percent",
        "Wilcoxon_p_MAE"
    )
    
    # ============================================================
    # Figure style
    # ============================================================
    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 8,
        "axes.titlesize": 10,
        "axes.labelsize": 9,
        "xtick.labelsize": 8,
        "ytick.labelsize": 8,
        "axes.linewidth": 0.8,
    })
    
    fig, axes = plt.subplots(
        1, 3,
        figsize=(11.69, 3.25),
        dpi=300
    )
    
    # Same green/yellow/red style
    cmap = "RdYlGn"
    
    # Color limits matching your example
    r2_vmin, r2_vmax = -5, 85
    err_vmin, err_vmax = -5, 30
    
    panels = [
        {
            "ax": axes[0],
            "values": r2_values,
            "stars": r2_stars,
            "title": "R² improvement (%)",
            "label": "(a)",
            "vmin": r2_vmin,
            "vmax": r2_vmax,
            "cbar_ticks": np.arange(0, 81, 10)
        },
        {
            "ax": axes[1],
            "values": rmse_values,
            "stars": rmse_stars,
            "title": "RMSE improvement (%)",
            "label": "(b)",
            "vmin": err_vmin,
            "vmax": err_vmax,
            "cbar_ticks": np.arange(-5, 31, 5)
        },
        {
            "ax": axes[2],
            "values": mae_values,
            "stars": mae_stars,
            "title": "MAE improvement (%)",
            "label": "(c)",
            "vmin": err_vmin,
            "vmax": err_vmax,
            "cbar_ticks": np.arange(-5, 31, 5)
        }
    ]
    
    # ============================================================
    # Draw heatmaps
    # ============================================================
    for panel in panels:
        ax = panel["ax"]
        values = panel["values"]
        stars = panel["stars"]
    
        im = ax.imshow(
            values.to_numpy(dtype=float),
            cmap=cmap,
            vmin=panel["vmin"],
            vmax=panel["vmax"],
            aspect="auto"
        )
    
        # Axis ticks
        ax.set_xticks(np.arange(len(col_order)))
        ax.set_xticklabels(col_order)
    
        ax.set_yticks(np.arange(len(row_order)))
        ax.set_yticklabels(row_order)
    
        # Only first panel has y-axis label and y tick labels
        if ax == axes[0]:
            ax.set_ylabel("Feature set after adding SE")
        else:
            ax.set_yticklabels([])
    
        # White cell borders
        ax.set_xticks(np.arange(-0.5, len(col_order), 1), minor=True)
        ax.set_yticks(np.arange(-0.5, len(row_order), 1), minor=True)
    
        ax.grid(
            which="minor",
            color="white",
            linestyle="-",
            linewidth=2
        )
    
        ax.tick_params(which="minor", bottom=False, left=False)
    
        # Cell text
        for i in range(len(row_order)):
            for j in range(len(col_order)):
                value = values.iloc[i, j]
                sig = stars.iloc[i, j]
    
                ax.text(
                    j,
                    i,
                    f"{value:.1f}%\n{sig}",
                    ha="center",
                    va="center",
                    color="black",
                    fontsize=8,
                    fontweight="bold" if sig != "ns" else "normal"
                )
    
        # Title and panel label
        ax.set_title(panel["title"], pad=8)
    
        ax.text(
            -0.23,
            1.04,
            panel["label"],
            transform=ax.transAxes,
            fontsize=10,
            fontweight="bold",
            ha="left",
            va="bottom"
        )
    
        # Colorbar
        cbar = fig.colorbar(
            im,
            ax=ax,
            fraction=0.046,
            pad=0.035,
            ticks=panel["cbar_ticks"]
        )
    
        cbar.ax.tick_params(labelsize=7)
    
    # ============================================================
    # Footnote
    # ============================================================
    fig.text(
        0.5,
        0.04,
        "Positive values indicate improved performance after adding SE predictors. "
        "* p < 0.05, ** p < 0.01, *** p < 0.001; ns = not significant.",
        ha="center",
        va="center",
        fontsize=8
    )
    
    # ============================================================
    # Layout
    # ============================================================
    plt.subplots_adjust(
        left=0.08,
        right=0.965,
        top=0.84,
        bottom=0.22,
        wspace=0.14
    )
    
    # ============================================================
    # Save
    # ============================================================
    out_png = DATA_DIR / "O2_SE_added_value_heatmap.png"
    out_pdf = DATA_DIR / "O2_SE_added_value_heatmap.pdf"
    
    plt.savefig(out_png, dpi=300)
    plt.savefig(out_pdf)
    
    plt.close("all")
    
    print("Saved PNG to:", out_png)
    print("Saved PDF to:", out_pdf)
    
except Exception as exc:
    print('Warning: skipped optional exact-format figure section 3:', exc)
    traceback.print_exc(limit=1)


# Optional exact-format manuscript figure section 4
try:
    
    import pandas as pd
    import matplotlib.pyplot as plt
    from pathlib import Path
    
    # ============================================================
    # Data folder
    # ============================================================
    DATA_DIR = OUT_DIR / "O3_label_efficiency"
    
    # Main input file
    summary_file = DATA_DIR / "O3_label_efficiency_SE_added_value_by_training_size.csv"
    
    if not summary_file.exists():
        raise FileNotFoundError(
            f"File not found:\n{summary_file}\n\n"
            "Check that O3_label_efficiency_SE_added_value_by_training_size.csv is inside this folder."
        )
    
    df = pd.read_csv(summary_file)
    
    # ============================================================
    # Clean model names
    # ============================================================
    model_name_map = {
        "RF": "RF",
        "GradientBoosting": "GB",
        "XGBoost": "XGBoost"
    }
    
    df["Model_plot"] = df["Model"].replace(model_name_map)
    
    # ============================================================
    # Plot order and colors
    # ============================================================
    model_order = ["RF", "GB", "XGBoost"]
    
    colors = {
        "RF": "#1f77b4",
        "GB": "#ff7f0e",
        "XGBoost": "#2ca02c"
    }
    
    # ============================================================
    # Figure style
    # ============================================================
    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 12,
        "axes.labelsize": 13,
        "xtick.labelsize": 12,
        "ytick.labelsize": 12,
        "legend.fontsize": 12,
        "legend.title_fontsize": 13,
        "axes.linewidth": 1.0,
    })
    
    fig, ax = plt.subplots(
        figsize=(8.6, 6.6),
        dpi=300
    )
    
    # ============================================================
    # Plot lines with ±1 SD error bars
    # ============================================================
    for model in model_order:
        plot_df = (
            df[df["Model_plot"] == model]
            .sort_values("Train_Proportion")
            .copy()
        )
    
        x = plot_df["Train_Proportion"] * 100
        y = plot_df["RMSE_percent_improvement_mean"]
        yerr = plot_df["RMSE_percent_improvement_sd"]
    
        ax.errorbar(
            x,
            y,
            yerr=yerr,
            marker="o",
            markersize=7,
            linewidth=2.2,
            elinewidth=1.2,
            capsize=4,
            capthick=1.0,
            color=colors[model],
            alpha=0.95,
            label=model
        )
    
    # ============================================================
    # Zero reference line
    # ============================================================
    ax.axhline(
        0,
        color="black",
        linestyle="--",
        linewidth=1.2,
        alpha=0.65
    )
    
    # ============================================================
    # Axes, labels, ticks
    # ============================================================
    ax.set_xlabel("Training data used (%)")
    ax.set_ylabel("RMSE improvement from SE (%)")
    
    ax.set_xlim(16, 104)
    ax.set_xticks([20, 30, 40, 50, 60, 70, 80, 90, 100])
    
    ax.set_ylim(-32, 18)
    ax.set_yticks([-30, -20, -10, 0, 10])
    
    # Grid
    ax.grid(
        True,
        axis="both",
        color="#d9d9d9",
        linewidth=0.7,
        alpha=0.55
    )
    
    ax.set_axisbelow(True)
    
    # Legend
    ax.legend(
        title="Model",
        loc="upper right",
        frameon=True
    )
    
    # ============================================================
    # Layout and save
    # ============================================================
    plt.tight_layout()
    
    out_png = DATA_DIR / "O3_RMSE_improvement_from_SE_by_training_size.png"
    out_pdf = DATA_DIR / "O3_RMSE_improvement_from_SE_by_training_size.pdf"
    
    plt.savefig(out_png, dpi=300)
    plt.savefig(out_pdf)
    
    plt.close("all")
    
    print("Saved PNG to:", out_png)
    print("Saved PDF to:", out_pdf)
except Exception as exc:
    print('Warning: skipped optional exact-format figure section 4:', exc)
    traceback.print_exc(limit=1)




# ============================================================
# FINAL MANUSCRIPT FIGURES WITH FIXED MARGINS
# This block regenerates only the final manuscript figures from
# the CSV outputs created above. It does NOT rerun models.
# Figures are saved before final collection so they are included
# in ALL_FIGURES_TOGETHER and the final zip package.
# ============================================================
from typing import Optional, Iterable
from matplotlib.patches import Patch
from matplotlib.lines import Line2D
from matplotlib.ticker import MultipleLocator, FormatStrFormatter
from matplotlib.gridspec import GridSpec

def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def save_fig(fig: plt.Figure, out_base: Path, dpi: int = 300, tight: bool = False) -> None:
    out_base.parent.mkdir(parents=True, exist_ok=True)
    if tight:
        fig.savefig(out_base.with_suffix(".png"), dpi=dpi, bbox_inches="tight", pad_inches=0.08)
        fig.savefig(out_base.with_suffix(".pdf"), bbox_inches="tight", pad_inches=0.08)
    else:
        fig.savefig(out_base.with_suffix(".png"), dpi=dpi)
        fig.savefig(out_base.with_suffix(".pdf"))
    plt.close(fig)
    print("Saved:", out_base.with_suffix(".png"))
    print("Saved:", out_base.with_suffix(".pdf"))


def first_existing(paths: Iterable[Path]) -> Optional[Path]:
    for p in paths:
        if p.exists():
            return p
    return None


def find_file(root: Path, filename: str) -> Optional[Path]:
    direct = root / filename
    if direct.exists():
        return direct
    matches = list(root.rglob(filename))
    if matches:
        # Prefer shortest path / likely original, not collected copy.
        matches = sorted(matches, key=lambda p: ("FINAL_NECESSARY" in str(p), len(str(p))))
        return matches[0]
    return None


def clean_feature_set_name(name: str) -> str:
    return str(name).replace("RGBNIR", "MSI")


def clean_model_name(name: str) -> str:
    return {"GradientBoosting": "GB", "Gradient_Boosting": "GB"}.get(str(name), str(name))


def safe_label(text: str) -> str:
    return str(text).replace("RGBNIR", "MSI")


# ============================================================
# Figure 1: A4 top predictors
# ============================================================

def generate_top_predictors(o1_dir: Path, out_dir: Path, dpi: int, xlabel: str) -> None:
    data_dir = o1_dir / "O4_SHAP_and_complementarity" / "O4_algorithm_specific_SHAP"

    all_file = data_dir / "All_algorithm_specific_SHAP_predictor_importance.csv"
    overall_file = data_dir / "Overall_mean_SHAP_importance_across_algorithms.csv"

    if not all_file.exists():
        all_file = find_file(o1_dir, "All_algorithm_specific_SHAP_predictor_importance.csv")
    if not overall_file.exists():
        overall_file = find_file(o1_dir, "Overall_mean_SHAP_importance_across_algorithms.csv")

    if all_file is None or not Path(all_file).exists():
        raise FileNotFoundError("Could not find All_algorithm_specific_SHAP_predictor_importance.csv")
    if overall_file is None or not Path(overall_file).exists():
        raise FileNotFoundError("Could not find Overall_mean_SHAP_importance_across_algorithms.csv")

    all_df = pd.read_csv(all_file)
    overall_df = pd.read_csv(overall_file)

    if "Mean_abs_SHAP" not in all_df.columns:
        raise ValueError(f"Expected column 'Mean_abs_SHAP' in {all_file}")
    if "Mean_abs_SHAP_overall" not in overall_df.columns:
        raise ValueError(f"Expected column 'Mean_abs_SHAP_overall' in {overall_file}")

    # Clean group names.
    def clean_group(group: str) -> str:
        group = str(group)
        if group in ["RGB-NIR", "RGBNIR", "MSI", "RGB_NIR"]:
            return "MSI"
        return group

    all_df["Plot_Group"] = all_df["Group"].apply(clean_group)
    overall_df["Plot_Group"] = overall_df["Group"].apply(clean_group)

    if "Model_plot" not in all_df.columns:
        all_df["Model_plot"] = all_df["Model"].apply(clean_model_name)

    group_colors = {
        "CHM": "#6aaed6",
        "MSI": "#e7ad72",
        "SE": "#7bd88f",
    }

    overall_df = (
        overall_df
        .sort_values("Mean_abs_SHAP_overall", ascending=False)
        .head(20)
    )

    per_model = {}
    for m in ["RF", "GB", "XGBoost"]:
        mdf = all_df[all_df["Model_plot"] == m].copy()
        if mdf.empty:
            continue
        per_model[m] = mdf.sort_values("Mean_abs_SHAP", ascending=False).head(10)

    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 8,
        "axes.titlesize": 10,
        "axes.labelsize": 8,
        "xtick.labelsize": 8,
        "ytick.labelsize": 8,
        "axes.linewidth": 0.9,
    })

    fig, axes = plt.subplots(2, 2, figsize=(11.69, 8.27), dpi=dpi)
    axes = axes.flatten()

    def plot_barh(ax, df, title, panel_label, tick_step=0.02):
        value_col = "Mean_abs_SHAP_overall" if "Mean_abs_SHAP_overall" in df.columns else "Mean_abs_SHAP"
        df_plot = df.sort_values(value_col, ascending=True).copy()
        colors = df_plot["Plot_Group"].map(group_colors).fillna("#bdbdbd")
        xmax = float(df[value_col].max()) * 1.10 if len(df) else 0.1
        xmax = max(xmax, 0.01)

        ax.barh(
            df_plot["Predictor"],
            df_plot[value_col],
            color=colors,
            edgecolor=colors,
            height=0.70,
        )
        ax.set_title(title, fontweight="bold", pad=7)
        ax.set_xlabel("Mean |SHAP value|")
        ax.set_xlim(0, xmax)
        ax.xaxis.set_major_locator(MultipleLocator(tick_step))
        ax.xaxis.set_major_formatter(FormatStrFormatter("%.2f"))
        ax.grid(axis="x", linestyle="--", linewidth=0.6, alpha=0.35)
        ax.set_axisbelow(True)
        ax.tick_params(axis="y", labelsize=8, pad=2)
        ax.tick_params(axis="x", labelsize=8)
        ax.text(
            0.0, 1.045, panel_label,
            transform=ax.transAxes,
            fontsize=10,
            fontweight="bold",
            ha="left",
            va="bottom",
            clip_on=False,
        )

    plot_barh(axes[0], overall_df, "Overall Top 20 Predictors", "(a)")
    plot_barh(axes[1], per_model.get("RF", overall_df.head(10)), "RF | Top 10 Predictors", "(b)")
    plot_barh(axes[2], per_model.get("GB", overall_df.head(10)), "GB | Top 10 Predictors", "(c)")
    plot_barh(axes[3], per_model.get("XGBoost", overall_df.head(10)), "XGBoost | Top 10 Predictors", "(d)")

    legend_handles = [
        Patch(facecolor=group_colors["CHM"], edgecolor="black", label="CHM"),
        Patch(facecolor=group_colors["MSI"], edgecolor="black", label="MSI"),
        Patch(facecolor=group_colors["SE"], edgecolor="black", label="SE"),
    ]
    fig.legend(
        handles=legend_handles,
        loc="lower center",
        ncol=3,
        frameon=False,
        bbox_to_anchor=(0.5, 0.035),
        fontsize=8,
    )

    # Larger left margin prevents long labels being clipped.
    fig.subplots_adjust(
        left=0.185,
        right=0.985,
        top=0.94,
        bottom=0.12,
        wspace=0.32,
        hspace=0.28,
    )

    save_fig(fig, out_dir / "Top_predictors_A4_exact_style_FIXED", dpi=dpi, tight=False)


# ============================================================
# Figure 2: SHAP interaction + complementarity A4 landscape
# ============================================================

def generate_shap_interaction(o1_dir: Path, out_dir: Path, dpi: int) -> None:
    o4_dir = o1_dir / "O4_SHAP_and_complementarity"
    interaction_file = first_existing([
        o4_dir / "O5_SHAP_interaction_values.csv",
        o1_dir / "O5_SHAP_interaction_values.csv",
    ]) or find_file(o1_dir, "O5_SHAP_interaction_values.csv")
    correlation_file = first_existing([
        o4_dir / "O5_SHAP_complementarity_correlation.csv",
        o1_dir / "O5_SHAP_complementarity_correlation.csv",
    ]) or find_file(o1_dir, "O5_SHAP_complementarity_correlation.csv")

    if interaction_file is None or correlation_file is None:
        raise FileNotFoundError(
            "Could not find O5_SHAP_interaction_values.csv and/or "
            "O5_SHAP_complementarity_correlation.csv"
        )

    interaction = pd.read_csv(interaction_file)
    correlation = pd.read_csv(correlation_file)

    interaction["Feature pair"] = interaction["SE_Feature"].astype(str) + " x " + interaction["Other_Feature"].astype(str)
    correlation["Feature pair"] = correlation["SE_Feature"].astype(str) + " x " + correlation["Other_Feature"].astype(str)

    top_n = min(35, len(interaction), len(correlation))
    left_df = interaction.sort_values("Mean_abs_SHAP_Interaction", ascending=False).head(top_n).copy()
    right_df = correlation.sort_values("Abs_SHAP_Correlation", ascending=False).head(top_n).copy()

    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 7,
        "axes.titlesize": 9,
        "axes.labelsize": 8,
        "xtick.labelsize": 7,
        "ytick.labelsize": 5.9,
        "axes.linewidth": 0.8,
    })

    fig, axes = plt.subplots(1, 2, figsize=(11.69, 8.27), dpi=dpi)
    bar_color = "#cd9778"

    def make_panel(ax, df, value_col, title, xlabel, tick_step, tick_format):
        df_plot = df.sort_values(value_col, ascending=True).copy()
        y = np.arange(len(df_plot))
        xmax = float(df[value_col].max()) * 1.10 if len(df) else 1.0
        ax.barh(y, df_plot[value_col], color=bar_color, edgecolor=bar_color, height=0.68)
        ax.set_yticks(y)
        ax.set_yticklabels(df_plot["Feature pair"], fontsize=5.9)
        ax.set_title(title, fontsize=9, pad=8)
        ax.set_xlabel(xlabel, fontsize=8)
        ax.set_ylabel("Feature pair", fontsize=8, labelpad=8)
        ax.set_xlim(0, xmax)
        ax.xaxis.set_major_locator(MultipleLocator(tick_step))
        ax.xaxis.set_major_formatter(FormatStrFormatter(tick_format))
        ax.xaxis.grid(True, color="#e6e6e6", linewidth=0.6)
        ax.set_axisbelow(True)
        ax.tick_params(axis="y", length=3, pad=2, labelsize=5.9)
        ax.tick_params(axis="x", length=3, labelsize=7)

    make_panel(
        axes[0],
        left_df,
        "Mean_abs_SHAP_Interaction",
        "SHAP interaction strength: SE with CHM/MSI",
        "Mean absolute SHAP interaction value",
        tick_step=0.004,
        tick_format="%.3f",
    )
    make_panel(
        axes[1],
        right_df,
        "Abs_SHAP_Correlation",
        "Complementarity between SE and CHM/MSI predictors",
        "Absolute correlation between SHAP values",
        tick_step=0.1,
        tick_format="%.1f",
    )

    axes[0].text(-0.08, 1.025, "(a)", transform=axes[0].transAxes, fontsize=11, fontweight="bold", clip_on=False)
    axes[1].text(-0.08, 1.025, "(b)", transform=axes[1].transAxes, fontsize=11, fontweight="bold", clip_on=False)

    fig.subplots_adjust(left=0.245, right=0.985, top=0.935, bottom=0.085, wspace=0.72)
    save_fig(fig, out_dir / "SHAP_interaction_complementarity_A4_landscape_FIXED", dpi=dpi, tight=False)


# ============================================================
# Figure 3: O3 label efficiency 3 x 3 A4
# ============================================================

def generate_o3_label_efficiency(o1_dir: Path, out_dir: Path, dpi: int) -> None:
    data_dir = o1_dir / "O3_label_efficiency"
    summary_file = data_dir / "O3_label_efficiency_summary.csv"
    if not summary_file.exists():
        summary_file = find_file(o1_dir, "O3_label_efficiency_summary.csv")
    if summary_file is None:
        raise FileNotFoundError("Could not find O3_label_efficiency_summary.csv")

    df = pd.read_csv(summary_file)
    df["Feature_Set_Plot"] = df["Feature_Set"].replace({
        "CHM+RGBNIR": "CHM+MSI",
        "CHM+RGBNIR+SE": "CHM+MSI+SE",
    })

    model_order = ["RF", "GradientBoosting", "XGBoost"]
    model_title_map = {"RF": "RF", "GradientBoosting": "GB", "XGBoost": "XGBoost"}
    feature_order = ["CHM+MSI", "CHM+MSI+SE"]
    colors = {"CHM+MSI": "#1f77b4", "CHM+MSI+SE": "#ff7f0e"}

    metrics = [
        {"mean_col": "R2_mean", "sd_col": "R2_sd", "ylabel": "R²", "ylim": (-0.55, 0.72), "yticks": [-0.4, -0.2, 0.0, 0.2, 0.4, 0.6]},
        {"mean_col": "RMSE_mean", "sd_col": "RMSE_sd", "ylabel": "RMSE", "ylim": (0.105, 0.245), "yticks": [0.12, 0.14, 0.16, 0.18, 0.20, 0.22, 0.24]},
        {"mean_col": "MAE_mean", "sd_col": "MAE_sd", "ylabel": "MAE", "ylim": (0.08, 0.16), "yticks": [0.08, 0.09, 0.10, 0.11, 0.12, 0.13, 0.14, 0.15, 0.16]},
    ]

    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 8,
        "axes.titlesize": 9,
        "axes.labelsize": 8,
        "xtick.labelsize": 8,
        "ytick.labelsize": 8,
        "axes.linewidth": 0.9,
    })

    fig, axes = plt.subplots(3, 3, figsize=(11.69, 8.27), dpi=dpi)
    panel_letters = ["(a)", "(b)", "(c)", "(d)", "(e)", "(f)", "(g)", "(h)", "(i)"]
    panel_i = 0

    for row_i, metric in enumerate(metrics):
        for col_i, model in enumerate(model_order):
            ax = axes[row_i, col_i]
            model_df = df[df["Model"] == model].copy()

            for feature_set in feature_order:
                plot_df = model_df[model_df["Feature_Set_Plot"] == feature_set].sort_values("Train_Proportion").copy()
                if plot_df.empty:
                    continue
                x = plot_df["Train_Proportion"].to_numpy() * 100
                y = plot_df[metric["mean_col"]].to_numpy(dtype=float)
                sd = plot_df[metric["sd_col"]].fillna(0).to_numpy(dtype=float)

                ax.plot(x, y, marker="o", markersize=4.5, linewidth=1.8, color=colors[feature_set], label=feature_set)
                ax.fill_between(x, y - sd, y + sd, color=colors[feature_set], alpha=0.16, linewidth=0)

            if row_i == 0:
                ax.set_title(model_title_map[model], fontweight="bold", pad=8)
            if col_i == 0:
                ax.set_ylabel(metric["ylabel"])
            if row_i == 2:
                ax.set_xlabel("Training data used (%)")
            else:
                ax.set_xticklabels([])

            ax.set_xlim(17, 103)
            ax.set_xticks([20, 40, 60, 80, 100])
            ax.set_ylim(metric["ylim"])
            ax.set_yticks(metric["yticks"])
            ax.grid(True, axis="both", color="#d9d9d9", linewidth=0.6, alpha=0.55)
            ax.set_axisbelow(True)
            ax.text(0.02, 1.04, panel_letters[panel_i], transform=ax.transAxes, fontsize=10, fontweight="bold", ha="left", va="bottom", clip_on=False)
            panel_i += 1

    handles, labels = axes[0, 0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="upper center", bbox_to_anchor=(0.5, 0.985), ncol=2, frameon=False, fontsize=8)
    fig.text(0.5, 0.035, "Lines show mean performance across training proportions; shaded bands indicate ±1 SD.", ha="center", va="center", fontsize=8)
    fig.subplots_adjust(left=0.075, right=0.985, top=0.92, bottom=0.10, wspace=0.18, hspace=0.22)

    save_fig(fig, out_dir / "O3_label_efficiency_3x3_A4_FIXED", dpi=dpi, tight=False)


# ============================================================
# Figure 4: O3 RMSE improvement from SE by training size
# ============================================================

def generate_o3_rmse_improvement(o1_dir: Path, out_dir: Path, dpi: int) -> None:
    data_dir = o1_dir / "O3_label_efficiency"
    summary_file = data_dir / "O3_label_efficiency_SE_added_value_by_training_size.csv"
    if not summary_file.exists():
        summary_file = find_file(o1_dir, "O3_label_efficiency_SE_added_value_by_training_size.csv")
    if summary_file is None:
        raise FileNotFoundError("Could not find O3_label_efficiency_SE_added_value_by_training_size.csv")

    df = pd.read_csv(summary_file)
    df["Model_plot"] = df["Model"].replace({"GradientBoosting": "GB", "Gradient_Boosting": "GB"})

    model_order = ["RF", "GB", "XGBoost"]
    colors = {"RF": "#1f77b4", "GB": "#ff7f0e", "XGBoost": "#2ca02c"}

    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 12,
        "axes.labelsize": 13,
        "xtick.labelsize": 12,
        "ytick.labelsize": 12,
        "legend.fontsize": 12,
        "legend.title_fontsize": 13,
        "axes.linewidth": 1.0,
    })

    fig, ax = plt.subplots(figsize=(8.6, 6.6), dpi=dpi)

    for model in model_order:
        plot_df = df[df["Model_plot"] == model].sort_values("Train_Proportion").copy()
        if plot_df.empty:
            continue
        x = plot_df["Train_Proportion"] * 100
        y = plot_df["RMSE_percent_improvement_mean"]
        yerr = plot_df.get("RMSE_percent_improvement_sd", pd.Series(np.zeros(len(plot_df)), index=plot_df.index)).fillna(0)
        ax.errorbar(x, y, yerr=yerr, marker="o", markersize=7, linewidth=2.2, elinewidth=1.2, capsize=4, capthick=1.0, color=colors[model], alpha=0.95, label=model)

    ax.axhline(0, color="black", linestyle="--", linewidth=1.2, alpha=0.65)
    ax.set_xlabel("Training data used (%)")
    ax.set_ylabel("RMSE improvement from SE (%)")
    ax.set_xlim(16, 104)
    ax.set_xticks([20, 30, 40, 50, 60, 70, 80, 90, 100])

    yvals = df["RMSE_percent_improvement_mean"].replace([np.inf, -np.inf], np.nan).dropna()
    if len(yvals):
        ymin = min(-32, float((df["RMSE_percent_improvement_mean"] - df.get("RMSE_percent_improvement_sd", 0)).min()) - 2)
        ymax = max(18, float((df["RMSE_percent_improvement_mean"] + df.get("RMSE_percent_improvement_sd", 0)).max()) + 2)
        ax.set_ylim(ymin, ymax)
    else:
        ax.set_ylim(-32, 18)

    ax.grid(True, axis="both", color="#d9d9d9", linewidth=0.7, alpha=0.55)
    ax.set_axisbelow(True)
    ax.legend(title="Model", loc="upper right", frameon=True)
    fig.tight_layout(pad=1.2)

    save_fig(fig, out_dir / "O3_RMSE_improvement_from_SE_by_training_size_FIXED", dpi=dpi, tight=True)


# ============================================================
# Figure 5: O7 condition-specific SE benefit split dotplot
# ============================================================

def generate_o7_condition_dotplot(o1_dir: Path, out_dir: Path, dpi: int) -> None:
    data_dir = o1_dir / "O7_condition_specific_SE_benefit"
    corr_file = data_dir / "O7_condition_correlations_with_SE_improvement.csv"
    if not corr_file.exists():
        corr_file = find_file(o1_dir, "O7_condition_correlations_with_SE_improvement.csv")
    if corr_file is None:
        raise FileNotFoundError("Could not find O7_condition_correlations_with_SE_improvement.csv")

    corr_df = pd.read_csv(corr_file)
    corr_df.columns = corr_df.columns.str.strip()
    corr_df["Model"] = corr_df["Model"].replace({"GradientBoosting": "GB", "Gradient_Boosting": "GB"})

    model_order = ["RF", "GB", "XGBoost"]
    model_colors = {"RF": "#1f77b4", "GB": "#ff7f0e", "XGBoost": "#2ca02c"}
    model_markers = {"RF": "o", "GB": "s", "XGBoost": "^"}
    corr_col = "Spearman_Correlation_with_SE_Improvement"

    if corr_col not in corr_df.columns:
        raise ValueError(f"Missing column: {corr_col}")

    corr_df["Condition_Label"] = corr_df["Condition_Variable"].astype(str).str.replace("_", " ", regex=False)
    corr_df["Full_Label"] = corr_df["Model"] + " | " + corr_df["Condition_Label"]

    n_show = 10
    neg_df = corr_df[corr_df[corr_col] < 0].sort_values(corr_col, ascending=True).head(n_show).copy()
    pos_df = corr_df[corr_df[corr_col] > 0].sort_values(corr_col, ascending=False).head(n_show).copy()
    neg_df = neg_df.sort_values(corr_col, ascending=False).reset_index(drop=True)
    pos_df = pos_df.sort_values(corr_col, ascending=True).reset_index(drop=True)

    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 9,
        "axes.labelsize": 9,
        "xtick.labelsize": 9,
        "ytick.labelsize": 8.5,
        "axes.linewidth": 0.9,
    })

    fig, axes = plt.subplots(nrows=1, ncols=2, figsize=(13.4, 5.9), dpi=dpi, sharex=False)

    def panel(ax, df, positive: bool):
        y = np.arange(len(df))
        if df.empty:
            ax.axvline(0, linestyle="--", linewidth=1, color="black")
            ax.set_yticks([])
            ax.set_xlabel("Spearman correlation", fontsize=9)
            ax.grid(axis="x", alpha=0.25)
            ax.set_xlim(-0.1, 0.1)
            ax.text(0.5, 0.5, "No associations", transform=ax.transAxes,
                    ha="center", va="center", fontsize=9)
            return
        for i, row in df.iterrows():
            model = row["Model"]
            ax.scatter(row[corr_col], i, color=model_colors.get(model, "gray"), marker=model_markers.get(model, "o"), s=55, zorder=3)
            if positive:
                ax.text(row[corr_col] + 0.006, i, f"{row[corr_col]:.2f}", ha="left", va="center", fontsize=8, clip_on=False)
            else:
                ax.text(row[corr_col] - 0.006, i, f"{row[corr_col]:.2f}", ha="right", va="center", fontsize=8, clip_on=False)
        ax.axvline(0, linestyle="--", linewidth=1, color="black")
        ax.set_yticks(y)
        ax.set_yticklabels(df["Full_Label"], fontsize=8.5)
        ax.set_xlabel("Spearman correlation", fontsize=9)
        ax.grid(axis="x", alpha=0.25)
        ax.set_axisbelow(True)

    panel(axes[0], neg_df, positive=False)
    if not neg_df.empty:
        axes[0].set_xlim(min(neg_df[corr_col].min() - 0.04, -0.22), 0.02)
    axes[0].text(0.01, 1.04, "(a)", transform=axes[0].transAxes, fontsize=11, fontweight="bold", ha="left", va="bottom", clip_on=False)
    axes[0].text(1.12, 0.5, "Conditions associated with weaker SE benefit", transform=axes[0].transAxes, rotation=-90, fontsize=9, ha="center", va="center", clip_on=False)

    panel(axes[1], pos_df, positive=True)
    if not pos_df.empty:
        axes[1].set_xlim(-0.02, max(pos_df[corr_col].max() + 0.04, 0.14))
    axes[1].text(0.01, 1.04, "(b)", transform=axes[1].transAxes, fontsize=11, fontweight="bold", ha="left", va="bottom", clip_on=False)
    axes[1].text(1.12, 0.5, "Conditions associated with stronger SE benefit", transform=axes[1].transAxes, rotation=-90, fontsize=9, ha="center", va="center", clip_on=False)

    legend_handles = [
        Line2D([0], [0], marker=model_markers[m], color="w", markerfacecolor=model_colors[m], markeredgecolor=model_colors[m], markersize=7, label=m)
        for m in model_order
    ]
    fig.legend(handles=legend_handles, loc="upper center", ncol=3, frameon=False, bbox_to_anchor=(0.5, 1.02), fontsize=9)
    fig.text(0.5, 0.015, "Positive correlations indicate greater SE benefit at higher condition-variable values; negative correlations indicate weaker SE benefit.", ha="center", fontsize=8)
    fig.subplots_adjust(left=0.20, right=0.91, top=0.88, bottom=0.14, wspace=0.78)

    save_fig(fig, out_dir / "Figure_O7_split_dotplot_condition_associations_FIXED", dpi=dpi, tight=True)


# ============================================================
# Figure 6: Observed vs predicted overview grid
# ============================================================

def generate_observed_predicted_grid(o1_dir: Path, out_dir: Path, dpi: int) -> None:
    pred_file = o1_dir / "O1_all_models_mean_repeated_outer_test_predictions.csv"
    if not pred_file.exists():
        pred_file = find_file(o1_dir, "O1_all_models_mean_repeated_outer_test_predictions.csv")
    if pred_file is None:
        raise FileNotFoundError("Could not find O1_all_models_mean_repeated_outer_test_predictions.csv")

    df = pd.read_csv(pred_file)
    required = {"Feature_Set", "Model", "Observed", "Predicted_Mean"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Prediction file is missing columns: {missing}")

    feature_sets_left = ["CHM", "RGBNIR", "SE", "CHM+RGBNIR", "CHM+SE", "RGBNIR+SE"]
    full_feature_set = "CHM+RGBNIR+SE"
    model_order = ["RF", "GradientBoosting", "XGBoost"]
    model_title = {"RF": "RF", "GradientBoosting": "GB", "XGBoost": "XGBoost"}

    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 7,
        "axes.titlesize": 7,
        "axes.labelsize": 8,
        "xtick.labelsize": 6,
        "ytick.labelsize": 6,
        "axes.linewidth": 0.8,
    })

    fig = plt.figure(figsize=(8.27, 11.69), dpi=dpi)  # A4 portrait
    gs = GridSpec(6, 5, figure=fig, width_ratios=[1, 1, 1, 0.18, 1.65], wspace=0.42, hspace=0.52)

    all_obs = df["Observed"].to_numpy(dtype=float)
    all_pred = df["Predicted_Mean"].to_numpy(dtype=float)
    lim_min = min(0.0, np.nanmin([all_obs.min(), all_pred.min()]))
    lim_max = max(0.95, np.nanmax([all_obs.max(), all_pred.max()]))
    lims = (lim_min, lim_max)

    def scatter_panel(ax, sub, title, big=False):
        ax.scatter(sub["Observed"], sub["Predicted_Mean"], s=10 if not big else 14, color="#1f77b4", alpha=0.70, edgecolor="none")
        ax.plot(lims, lims, "--", color="red", linewidth=0.9)
        ax.set_xlim(lims)
        ax.set_ylim(lims)
        ax.set_title(title, fontweight="bold", pad=4, fontsize=7 if not big else 9)
        ax.grid(True, color="#d9d9d9", linewidth=0.45, alpha=0.55)
        ax.set_axisbelow(True)
        ax.tick_params(labelsize=6 if not big else 7)

    # Left matrix: six feature sets x three models.
    for row_i, fs in enumerate(feature_sets_left):
        for col_i, model in enumerate(model_order):
            ax = fig.add_subplot(gs[row_i, col_i])
            sub = df[(df["Feature_Set"] == fs) & (df["Model"] == model)]
            if sub.empty:
                ax.axis("off")
                continue
            title = f"{clean_feature_set_name(fs)} | {model_title[model]}"
            scatter_panel(ax, sub, title, big=False)
            if row_i == len(feature_sets_left) - 1:
                ax.set_xlabel("Observed", fontsize=7)
            if col_i == 0:
                ax.set_ylabel("Predicted", fontsize=7)

    # Right larger panels: full predictor set for each model.
    spans = [(0, 2), (2, 4), (4, 6)]
    for model, (r0, r1) in zip(model_order, spans):
        ax = fig.add_subplot(gs[r0:r1, 4])
        sub = df[(df["Feature_Set"] == full_feature_set) & (df["Model"] == model)]
        if sub.empty:
            ax.axis("off")
            continue
        title = f"CHM+MSI+SE | {model_title[model]}"
        scatter_panel(ax, sub, title, big=True)
        ax.set_xlabel("Observed", fontsize=8)
        ax.set_ylabel("Predicted", fontsize=8)

    fig.subplots_adjust(left=0.07, right=0.97, top=0.98, bottom=0.05)
    save_fig(fig, out_dir / "Observed_vs_predicted_grid_A4_FIXED", dpi=dpi, tight=False)




try:
    FINAL_FIXED_FIG_DIR = RUN_ROOT / "FINAL_NECESSARY_FIGURES_FIXED"
    ensure_dir(FINAL_FIXED_FIG_DIR)

    print("\n============================================================")
    print("Generating final manuscript figures with fixed margins")
    print("============================================================")
    print("Output folder:", FINAL_FIXED_FIG_DIR)

    _final_figure_tasks = [
        ("Top predictors A4", lambda: generate_top_predictors(OUT_DIR, FINAL_FIXED_FIG_DIR, 300, "Mean |SHAP value|")),
        ("SHAP interaction/complementarity A4", lambda: generate_shap_interaction(OUT_DIR, FINAL_FIXED_FIG_DIR, 300)),
        ("O3 label-efficiency 3x3 A4", lambda: generate_o3_label_efficiency(OUT_DIR, FINAL_FIXED_FIG_DIR, 300)),
        ("O3 RMSE improvement from SE", lambda: generate_o3_rmse_improvement(OUT_DIR, FINAL_FIXED_FIG_DIR, 300)),
        ("O7 condition-specific SE benefit dotplot", lambda: generate_o7_condition_dotplot(OUT_DIR, FINAL_FIXED_FIG_DIR, 300)),
        ("Observed-vs-predicted grid", lambda: generate_observed_predicted_grid(OUT_DIR, FINAL_FIXED_FIG_DIR, 300)),
    ]

    _generated_final_figures = []
    _skipped_final_figures = []

    for _label, _func in _final_figure_tasks:
        print("\nGenerating:", _label)
        try:
            _func()
            _generated_final_figures.append(_label)
        except Exception as _exc:
            _skipped_final_figures.append((_label, str(_exc)))
            print("WARNING: skipped", _label)
            print(_exc)
            traceback.print_exc(limit=1)

    print("\nFinal fixed figures generated:")
    for _label in _generated_final_figures:
        print(" -", _label)

    if _skipped_final_figures:
        print("\nFinal fixed figures skipped:")
        for _label, _reason in _skipped_final_figures:
            print(f" - {_label}: {_reason}")
        print("Skipped figures usually mean the required CSV output was not produced in this run.")

except Exception as exc:
    print("Warning: skipped final fixed manuscript figure block:", exc)
    traceback.print_exc(limit=1)



# ============================================================
# FINAL MANUSCRIPT TABLES: generate tables as before
# ============================================================

print("\nGenerating final manuscript-style tables...")

FINAL_TABLE_DIR = RUN_ROOT / "FINAL_TABLES_AS_BEFORE"
FINAL_TABLE_DIR.mkdir(parents=True, exist_ok=True)


def _display_feature_set(value):
    """Use manuscript naming while keeping original analysis names in the raw columns."""
    return str(value).replace("RGBNIR", "MSI")


def _display_model(value):
    value = str(value)
    if value == "GradientBoosting":
        return "GB"
    return value


def _display_comparison(value):
    return str(value).replace("RGBNIR", "MSI")


def _fmt_mean_sd(mean_val, sd_val, decimals=3):
    try:
        if pd.isna(mean_val):
            return ""
        if pd.isna(sd_val):
            return f"{mean_val:.{decimals}f}"
        return f"{mean_val:.{decimals}f} ± {sd_val:.{decimals}f}"
    except Exception:
        return ""


def _round_numeric(df, digits=4):
    out = df.copy()
    for col in out.columns:
        if pd.api.types.is_numeric_dtype(out[col]):
            out[col] = out[col].round(digits)
    return out


def _safe_table(df, name):
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        print(f"Skipping empty table: {name}")
        return None
    out = df.copy()
    out.to_csv(FINAL_TABLE_DIR / f"{name}.csv", index=False)
    return out


final_tables = {}

# ------------------------------------------------------------
# Table 1. Main nested-CV model comparison
# ------------------------------------------------------------
try:
    t1 = summary_df.copy()
    t1["Predictor_Set"] = t1["Feature_Set"].apply(_display_feature_set)
    t1["Model_Display"] = t1["Model"].apply(_display_model)
    t1["R2_mean_sd"] = [
        _fmt_mean_sd(m, s, 3)
        for m, s in zip(t1["R2_mean"], t1["R2_sd"])
    ]
    t1["RMSE_mean_sd"] = [
        _fmt_mean_sd(m, s, 3)
        for m, s in zip(t1["RMSE_mean"], t1["RMSE_sd"])
    ]
    t1["MAE_mean_sd"] = [
        _fmt_mean_sd(m, s, 3)
        for m, s in zip(t1["MAE_mean"], t1["MAE_sd"])
    ]
    t1["N_Selected_Features_mean_sd"] = [
        _fmt_mean_sd(m, s, 1)
        for m, s in zip(t1["N_Selected_Features_mean"], t1["N_Selected_Features_sd"])
    ]
    t1 = t1[[
        "Predictor_Set", "Model_Display", "R2_mean_sd", "RMSE_mean_sd", "MAE_mean_sd",
        "N_Selected_Features_mean_sd", "Feature_Set", "Model", "R2_mean", "R2_sd",
        "RMSE_mean", "RMSE_sd", "MAE_mean", "MAE_sd",
        "N_Selected_Features_mean", "N_Selected_Features_sd"
    ]].rename(columns={
        "Predictor_Set": "Predictor set",
        "Model_Display": "Model",
        "R2_mean_sd": "R² mean ± SD",
        "RMSE_mean_sd": "RMSE mean ± SD",
        "MAE_mean_sd": "MAE mean ± SD",
        "N_Selected_Features_mean_sd": "Selected predictors mean ± SD",
        "Feature_Set": "Original feature-set name",
        "Model": "Original model name"
    })
    t1 = _round_numeric(t1, 4)
    final_tables["Table_01_nested_CV_model_comparison"] = _safe_table(t1, "Table_01_nested_CV_model_comparison")
except Exception as exc:
    print("Could not create Table 01:", exc)

# ------------------------------------------------------------
# Table 2. Best tuned model selected by nested-CV R²
# ------------------------------------------------------------
try:
    t2 = best_info.copy()
    t2["Best_Feature_Set_Display"] = t2["Best_Feature_Set"].apply(_display_feature_set)
    t2["Best_Model_Display"] = t2["Best_Model"].apply(_display_model)
    front_cols = ["Best_Feature_Set_Display", "Best_Model_Display"]
    remaining_cols = [c for c in t2.columns if c not in front_cols]
    t2 = t2[front_cols + remaining_cols].rename(columns={
        "Best_Feature_Set_Display": "Best predictor set",
        "Best_Model_Display": "Best model"
    })
    t2 = _round_numeric(t2, 4)
    final_tables["Table_02_best_model_selected_by_R2"] = _safe_table(t2, "Table_02_best_model_selected_by_R2")
except Exception as exc:
    print("Could not create Table 02:", exc)

# ------------------------------------------------------------
# Table 3. Best feature set by algorithm
# ------------------------------------------------------------
try:
    if "all_algorithm_best_info_df" in globals():
        t3 = all_algorithm_best_info_df.copy()
    elif "all_algorithm_best_info" in globals() and len(all_algorithm_best_info) > 0:
        t3 = pd.DataFrame(all_algorithm_best_info)
    else:
        t3 = pd.DataFrame()
    if not t3.empty:
        t3["Model_Display"] = t3["Model"].apply(_display_model)
        t3["Best_Feature_Set_Display"] = t3["Best_Feature_Set"].apply(_display_feature_set)
        t3["R2_mean_sd"] = [_fmt_mean_sd(m, s, 3) for m, s in zip(t3["Nested_CV_R2_mean"], t3["Nested_CV_R2_sd"])]
        t3["RMSE_mean_sd"] = [_fmt_mean_sd(m, s, 3) for m, s in zip(t3["Nested_CV_RMSE_mean"], t3["Nested_CV_RMSE_sd"])]
        t3["MAE_mean_sd"] = [_fmt_mean_sd(m, s, 3) for m, s in zip(t3["Nested_CV_MAE_mean"], t3["Nested_CV_MAE_sd"])]
        t3 = t3[[
            "Model_Display", "Best_Feature_Set_Display", "R2_mean_sd", "RMSE_mean_sd", "MAE_mean_sd",
            "N_Selected_Features", "Model", "Best_Feature_Set", "Nested_CV_R2_mean", "Nested_CV_R2_sd",
            "Nested_CV_RMSE_mean", "Nested_CV_RMSE_sd", "Nested_CV_MAE_mean", "Nested_CV_MAE_sd",
            "Final_Best_Params_JSON", "Selected_Features", "Saved_Pipeline"
        ]].rename(columns={
            "Model_Display": "Model",
            "Best_Feature_Set_Display": "Best predictor set",
            "R2_mean_sd": "R² mean ± SD",
            "RMSE_mean_sd": "RMSE mean ± SD",
            "MAE_mean_sd": "MAE mean ± SD",
            "N_Selected_Features": "Number of selected predictors",
            "Best_Feature_Set": "Original feature-set name",
            "Model": "Original model name"
        })
        t3 = _round_numeric(t3, 4)
    final_tables["Table_03_best_model_by_algorithm"] = _safe_table(t3, "Table_03_best_model_by_algorithm")
except Exception as exc:
    print("Could not create Table 03:", exc)

# ------------------------------------------------------------
# Table 4. SE added-value paired comparison
# ------------------------------------------------------------
try:
    t4 = added_value_df.copy()
    t4["Model_Display"] = t4["Model"].apply(_display_model)
    t4["Comparison_Display"] = t4["Comparison"].apply(_display_comparison)
    t4["Without_SE_Display"] = t4["Without_SE"].apply(_display_feature_set)
    t4["With_SE_Display"] = t4["With_SE"].apply(_display_feature_set)
    keep = [
        "Model_Display", "Comparison_Display", "Without_SE_Display", "With_SE_Display",
        "N_Paired_Outer_Splits", "R2_without_SE_mean", "R2_with_SE_mean", "Delta_R2_mean", "Delta_R2_sd", "Wilcoxon_p_R2",
        "RMSE_without_SE_mean", "RMSE_with_SE_mean", "Delta_RMSE_mean", "Delta_RMSE_sd",
        "RMSE_percent_improvement_mean", "Wilcoxon_p_RMSE",
        "MAE_without_SE_mean", "MAE_with_SE_mean", "Delta_MAE_mean", "Delta_MAE_sd",
        "MAE_percent_improvement_mean", "Wilcoxon_p_MAE"
    ]
    keep = [c for c in keep if c in t4.columns]
    t4 = t4[keep].rename(columns={
        "Model_Display": "Model",
        "Comparison_Display": "Comparison",
        "Without_SE_Display": "Without SE",
        "With_SE_Display": "With SE",
        "N_Paired_Outer_Splits": "Paired outer-CV splits",
        "RMSE_percent_improvement_mean": "RMSE improvement (%) mean",
        "MAE_percent_improvement_mean": "MAE improvement (%) mean"
    })
    t4 = _round_numeric(t4, 4)
    final_tables["Table_04_SE_added_value_paired_tests"] = _safe_table(t4, "Table_04_SE_added_value_paired_tests")
except Exception as exc:
    print("Could not create Table 04:", exc)

# ------------------------------------------------------------
# Table 5. Label-efficiency summary by training size
# ------------------------------------------------------------
try:
    t5 = label_eff_summary_df.copy()
    t5["Training data used (%)"] = (t5["Train_Proportion"] * 100).round(0).astype(int)
    t5["Predictor set"] = t5["Feature_Set"].apply(_display_feature_set)
    t5["Model display"] = t5["Model"].apply(_display_model)
    t5["R2_mean_sd"] = [_fmt_mean_sd(m, s, 3) for m, s in zip(t5["R2_mean"], t5["R2_sd"])]
    t5["RMSE_mean_sd"] = [_fmt_mean_sd(m, s, 3) for m, s in zip(t5["RMSE_mean"], t5["RMSE_sd"])]
    t5["MAE_mean_sd"] = [_fmt_mean_sd(m, s, 3) for m, s in zip(t5["MAE_mean"], t5["MAE_sd"])]
    t5 = t5[[
        "Training data used (%)", "Predictor set", "Model display", "Train_N_mean",
        "R2_mean_sd", "RMSE_mean_sd", "MAE_mean_sd", "Train_Proportion", "Feature_Set", "Model",
        "R2_mean", "R2_sd", "RMSE_mean", "RMSE_sd", "MAE_mean", "MAE_sd"
    ]].rename(columns={
        "Model display": "Model",
        "Train_N_mean": "Mean training N",
        "R2_mean_sd": "R² mean ± SD",
        "RMSE_mean_sd": "RMSE mean ± SD",
        "MAE_mean_sd": "MAE mean ± SD",
        "Feature_Set": "Original feature-set name",
    })
    t5 = _round_numeric(t5, 4)
    final_tables["Table_05_label_efficiency_summary"] = _safe_table(t5, "Table_05_label_efficiency_summary")
except Exception as exc:
    print("Could not create Table 05:", exc)

# ------------------------------------------------------------
# Table 6. SE gain by training size
# ------------------------------------------------------------
try:
    if "label_eff_added_df" in globals() and isinstance(label_eff_added_df, pd.DataFrame) and not label_eff_added_df.empty:
        t6 = label_eff_added_df.copy()
        t6["Training data used (%)"] = (t6["Train_Proportion"] * 100).round(0).astype(int)
        t6["Model display"] = t6["Model"].apply(_display_model)
        front = ["Training data used (%)", "Model display"]
        remaining = [c for c in t6.columns if c not in front]
        t6 = t6[front + remaining].rename(columns={"Model display": "Model"})
    elif "label_eff_gain_df" in globals() and isinstance(label_eff_gain_df, pd.DataFrame):
        t6 = label_eff_gain_df.copy()
        if "Train_Proportion" in t6.columns:
            t6["Training data used (%)"] = (t6["Train_Proportion"] * 100).round(0).astype(int)
        t6["Model display"] = t6["Model"].apply(_display_model)
    else:
        t6 = pd.DataFrame()
    t6 = _round_numeric(t6, 4)
    final_tables["Table_06_SE_gain_by_training_size"] = _safe_table(t6, "Table_06_SE_gain_by_training_size")
except Exception as exc:
    print("Could not create Table 06:", exc)

# ------------------------------------------------------------
# Table 7. SHAP predictor importance
# ------------------------------------------------------------
try:
    algo_shap_dir = OUT_DIR / "O4_SHAP_and_complementarity" / "O4_algorithm_specific_SHAP"

    if "all_algorithm_shap_df" in globals() and isinstance(all_algorithm_shap_df, pd.DataFrame):
        t7 = all_algorithm_shap_df.copy()
    elif (algo_shap_dir / "All_algorithm_specific_SHAP_predictor_importance.csv").exists():
        t7 = pd.read_csv(algo_shap_dir / "All_algorithm_specific_SHAP_predictor_importance.csv")
    else:
        t7 = shap_importance_df.copy()

    t7 = _round_numeric(t7, 6)
    final_tables["Table_07_SHAP_predictor_importance"] = _safe_table(
        t7,
        "Table_07_SHAP_predictor_importance"
    )

    if "overall_shap_df" in globals() and isinstance(overall_shap_df, pd.DataFrame):
        t7a = overall_shap_df.head(20).copy()
    elif (algo_shap_dir / "Table_07a_SHAP_top20_predictor_importance_across_algorithms.csv").exists():
        t7a = pd.read_csv(algo_shap_dir / "Table_07a_SHAP_top20_predictor_importance_across_algorithms.csv")
    elif (algo_shap_dir / "Overall_mean_SHAP_importance_across_algorithms.csv").exists():
        t7a = pd.read_csv(algo_shap_dir / "Overall_mean_SHAP_importance_across_algorithms.csv").head(20)
    else:
        t7a = t7.head(20).copy()

    t7a = _round_numeric(t7a, 6)
    final_tables["Table_07a_SHAP_top20_predictor_importance"] = _safe_table(
        t7a,
        "Table_07a_SHAP_top20_predictor_importance"
    )

except Exception as exc:
    print("Could not create Table 07:", exc)

# ------------------------------------------------------------
# Table 8. SHAP group importance
# ------------------------------------------------------------
try:
    algo_shap_dir = OUT_DIR / "O4_SHAP_and_complementarity" / "O4_algorithm_specific_SHAP"

    if "all_algorithm_group_shap_df" in globals() and isinstance(all_algorithm_group_shap_df, pd.DataFrame):
        t8 = all_algorithm_group_shap_df.copy()
    elif (algo_shap_dir / "All_algorithm_specific_SHAP_group_importance.csv").exists():
        t8 = pd.read_csv(algo_shap_dir / "All_algorithm_specific_SHAP_group_importance.csv")
    else:
        t8 = group_shap_df.copy()

    t8 = _round_numeric(t8, 6)
    final_tables["Table_08_SHAP_group_importance"] = _safe_table(
        t8,
        "Table_08_SHAP_group_importance"
    )

    if "overall_group_shap_df" in globals() and isinstance(overall_group_shap_df, pd.DataFrame):
        t8_overall = overall_group_shap_df.copy()
    elif (algo_shap_dir / "Overall_SHAP_group_importance_across_algorithms.csv").exists():
        t8_overall = pd.read_csv(algo_shap_dir / "Overall_SHAP_group_importance_across_algorithms.csv")
    else:
        t8_overall = pd.DataFrame()

    if not t8_overall.empty:
        t8_overall = _round_numeric(t8_overall, 6)
        final_tables["Table_08a_overall_SHAP_group_importance"] = _safe_table(
            t8_overall,
            "Table_08a_overall_SHAP_group_importance"
        )

except Exception as exc:
    print("Could not create Table 08:", exc)

# ------------------------------------------------------------
# Table 9. SHAP complementarity correlations
# ------------------------------------------------------------
try:
    t9 = complementarity_df.copy()
    t9 = _round_numeric(t9, 6)
    final_tables["Table_09_SHAP_complementarity_correlations"] = _safe_table(t9, "Table_09_SHAP_complementarity_correlations")
except Exception as exc:
    print("Could not create Table 09:", exc)

# ------------------------------------------------------------
# Table 10. SHAP interactions
# ------------------------------------------------------------
try:
    if "interaction_df" in globals() and isinstance(interaction_df, pd.DataFrame) and not interaction_df.empty:
        t10 = _round_numeric(interaction_df.copy(), 6)
    else:
        t10 = pd.DataFrame()
    final_tables["Table_10_SHAP_interactions"] = _safe_table(t10, "Table_10_SHAP_interactions")
except Exception as exc:
    print("Could not create Table 10:", exc)

# ------------------------------------------------------------
# Table 11. Condition-specific SE-benefit correlations
# ------------------------------------------------------------
try:
    t11 = condition_corr_df.copy()
    if "Model" in t11.columns:
        t11["Model display"] = t11["Model"].apply(_display_model)
        cols = ["Model display"] + [c for c in t11.columns if c != "Model display"]
        t11 = t11[cols].rename(columns={"Model display": "Model"})
    t11 = _round_numeric(t11, 6)
    final_tables["Table_11_condition_specific_SE_benefit_correlations"] = _safe_table(t11, "Table_11_condition_specific_SE_benefit_correlations")
except Exception as exc:
    print("Could not create Table 11:", exc)

# ------------------------------------------------------------
# Table 12. Condition-bin summary and best conditions
# ------------------------------------------------------------
try:
    if "bin_summary_df" in globals() and isinstance(bin_summary_df, pd.DataFrame) and not bin_summary_df.empty:
        t12 = _round_numeric(bin_summary_df.copy(), 5)
    else:
        t12 = pd.DataFrame()
    final_tables["Table_12_condition_bin_summary"] = _safe_table(t12, "Table_12_condition_bin_summary")
except Exception as exc:
    print("Could not create Table 12:", exc)

try:
    if "best_condition_table" in globals() and isinstance(best_condition_table, pd.DataFrame) and not best_condition_table.empty:
        t13 = _round_numeric(best_condition_table.copy(), 5)
    else:
        t13 = pd.DataFrame()
    final_tables["Table_13_best_condition_specific_SE_benefit"] = _safe_table(t13, "Table_13_best_condition_specific_SE_benefit")
except Exception as exc:
    print("Could not create Table 13:", exc)

# ------------------------------------------------------------
# Table 14. Mean repeated outer-test prediction metrics
# ------------------------------------------------------------
try:
    t14 = outer_prediction_metrics_df.copy()
    t14["Predictor set"] = t14["Feature_Set"].apply(_display_feature_set)
    t14["Model display"] = t14["Model"].apply(_display_model)
    front = ["Predictor set", "Model display"]
    t14 = t14[front + [c for c in t14.columns if c not in front]].rename(columns={"Model display": "Model"})
    t14 = _round_numeric(t14, 4)
    final_tables["Table_14_mean_repeated_outer_test_prediction_metrics"] = _safe_table(t14, "Table_14_mean_repeated_outer_test_prediction_metrics")
except Exception as exc:
    print("Could not create Table 14:", exc)

# ------------------------------------------------------------
# Combined final manuscript table workbook
# ------------------------------------------------------------
try:
    final_workbook = FINAL_TABLE_DIR / "FINAL_TABLES_AS_BEFORE.xlsx"
    with pd.ExcelWriter(final_workbook, engine="openpyxl") as writer:
        for sheet_name, table in final_tables.items():
            if table is None or not isinstance(table, pd.DataFrame) or table.empty:
                continue
            safe_sheet = _safe_filename(sheet_name, 31)[:31]
            table.to_excel(writer, sheet_name=safe_sheet, index=False)

        # Add a small table index sheet last/first by writing then moving it.
        index_table = pd.DataFrame([
            {"Sheet": _safe_filename(name, 31)[:31], "CSV file": f"{name}.csv", "Rows": 0 if table is None else len(table)}
            for name, table in final_tables.items()
            if table is not None and isinstance(table, pd.DataFrame) and not table.empty
        ])
        index_table.to_excel(writer, sheet_name="Table_index", index=False)

        # Basic formatting for readability.
        wb = writer.book
        if "Table_index" in wb.sheetnames:
            ws_index = wb["Table_index"]
            wb._sheets.remove(ws_index)
            wb._sheets.insert(0, ws_index)

        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)
        thin = Side(style="thin", color="DDDDDD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for col_cells in ws.columns:
                col_letter = col_cells[0].column_letter
                max_len = 0
                for cell in col_cells[:80]:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 45))

    # Also keep a copy at the run root for easy access.
    shutil.copy2(final_workbook, RUN_ROOT / "FINAL_TABLES_AS_BEFORE.xlsx")
    print("Final manuscript-style table workbook saved:", final_workbook)
except Exception as exc:
    print("Could not create FINAL_TABLES_AS_BEFORE.xlsx:", exc)
    traceback.print_exc(limit=1)

# Write a text summary of table outputs.
try:
    (FINAL_TABLE_DIR / "README_FINAL_TABLES.txt").write_text(
        "Final manuscript-style tables generated from the full reanalysis.\n"
        "Display names use MSI instead of RGBNIR for manuscript consistency, while original feature-set names are retained in raw columns where useful.\n\n"
        "Main file: FINAL_TABLES_AS_BEFORE.xlsx\n"
        "Each sheet is also exported as a separate CSV in this folder.\n",
        encoding="utf-8"
    )
except Exception:
    pass


try:
    _write_figure_selection_report()
except Exception as _fig_report_exc:
    print("Warning: could not write figure selection report:", _fig_report_exc)

# ============================================================
# FINAL COLLECTION: all tables, figures, logs, and a zip package
# ============================================================

def _safe_filename(text: str, max_len: int = 150) -> str:
    bad = '<>:"/\\|?*\n\r\t'
    out = ''.join('_' if ch in bad else ch for ch in str(text))
    out = out.replace(' ', '_')
    while '__' in out:
        out = out.replace('__', '_')
    return out[:max_len].strip('._') or 'file'


def _unique_path(folder: Path, name: str) -> Path:
    base = _safe_filename(Path(name).stem)
    suffix = Path(name).suffix
    candidate = folder / f"{base}{suffix}"
    i = 2
    while candidate.exists():
        candidate = folder / f"{base}_{i}{suffix}"
        i += 1
    return candidate


def _collect_outputs(run_root: Path):
    print("\nCollecting all generated tables and figures...")
    figures_dir = run_root / "FIGURES"
    tables_dir = run_root / "TABLES"
    files_dir = run_root / "FILES"
    figures_dir.mkdir(parents=True, exist_ok=True)
    tables_dir.mkdir(parents=True, exist_ok=True)
    files_dir.mkdir(parents=True, exist_ok=True)

    fig_exts = {'.png', '.jpg', '.jpeg', '.tif', '.tiff', '.svg', '.pdf'}
    table_exts = {'.csv', '.xlsx', '.xls'}
    other_exts = {'.json', '.txt', '.joblib', '.pkl'}
    index_rows = []

    skip_dirs = {figures_dir.resolve(), tables_dir.resolve(), files_dir.resolve()}

    for src in sorted(run_root.rglob('*')):
        if not src.is_file():
            continue
        if any(str(src.resolve()).startswith(str(d)) for d in skip_dirs):
            continue
        if src.name.startswith('ALL_RESULTS_PACKAGE'):
            continue
        ext = src.suffix.lower()
        if ext in fig_exts:
            dst = _unique_path(figures_dir, src.name)
            category = 'figure'
        elif ext in table_exts:
            dst = _unique_path(tables_dir, src.name)
            category = 'table'
        elif ext in other_exts:
            dst = _unique_path(files_dir, src.name)
            category = 'other_result_file'
        else:
            continue
        try:
            shutil.copy2(src, dst)
            index_rows.append({
                'category': category,
                'original_path': str(src),
                'collected_path': str(dst),
                'filename': dst.name,
                'extension': ext,
            })
        except Exception as exc:
            index_rows.append({
                'category': f'{category}_COPY_FAILED',
                'original_path': str(src),
                'collected_path': '',
                'filename': src.name,
                'extension': ext,
                'error': str(exc),
            })

    index_df = pd.DataFrame(index_rows)
    index_path = run_root / 'ALL_RESULTS_INDEX.csv'
    index_df.to_csv(index_path, index=False)

    # Combined workbook: all CSV files, plus sheets from XLSX files if readable.
    workbook_path = run_root / 'ALL_RESULTS_TABLES.xlsx'
    try:
        with pd.ExcelWriter(workbook_path, engine='openpyxl') as writer:
            used_sheets = set()
            for table_file in sorted(tables_dir.glob('*')):
                ext = table_file.suffix.lower()
                try:
                    if ext == '.csv':
                        df_tmp = pd.read_csv(table_file)
                        sheet = _safe_filename(table_file.stem, 25)
                        original_sheet = sheet
                        j = 2
                        while sheet in used_sheets:
                            sheet = _safe_filename(f'{original_sheet}_{j}', 31)
                            j += 1
                        used_sheets.add(sheet)
                        df_tmp.to_excel(writer, sheet_name=sheet[:31], index=False)
                    elif ext in {'.xlsx', '.xls'}:
                        xls = pd.ExcelFile(table_file)
                        for sheet_name in xls.sheet_names:
                            df_tmp = pd.read_excel(table_file, sheet_name=sheet_name)
                            sheet = _safe_filename(f'{table_file.stem}_{sheet_name}', 25)
                            original_sheet = sheet
                            j = 2
                            while sheet in used_sheets:
                                sheet = _safe_filename(f'{original_sheet}_{j}', 31)
                                j += 1
                            used_sheets.add(sheet)
                            df_tmp.to_excel(writer, sheet_name=sheet[:31], index=False)
                except Exception as exc:
                    pd.DataFrame([{'file': str(table_file), 'error': str(exc)}]).to_excel(
                        writer,
                        sheet_name=_safe_filename(f'ERROR_{table_file.stem}', 31)[:31],
                        index=False,
                    )
    except Exception as exc:
        (run_root / 'ALL_RESULTS_TABLES_WORKBOOK_ERROR.txt').write_text(
            f'Could not create ALL_RESULTS_TABLES.xlsx. Reason: {exc}\n',
            encoding='utf-8',
        )
        workbook_path = None

    # Simple HTML gallery for image formats that browsers can display.
    gallery_path = run_root / 'ALL_FIGURES_GALLERY.html'
    html_parts = [
        '<html><head><meta charset="utf-8"><title>All generated figures</title>',
        '<style>body{font-family:Arial,sans-serif;margin:24px;} .fig{margin:24px 0;padding-bottom:24px;border-bottom:1px solid #ddd;} img{max-width:1100px;width:100%;height:auto;border:1px solid #ddd;} code{background:#f6f6f6;padding:2px 4px;}</style>',
        '</head><body>',
        '<h1>All generated figures</h1>',
        f'<p>Run folder: <code>{_html.escape(str(run_root))}</code></p>',
    ]
    browser_exts = {'.png', '.jpg', '.jpeg', '.svg'}
    for fig in sorted(figures_dir.glob('*')):
        if fig.suffix.lower() in browser_exts:
            rel = fig.relative_to(run_root).as_posix()
            html_parts.append(
                f'<div class="fig"><h3>{_html.escape(fig.name)}</h3><img src="{_html.escape(rel)}"></div>'
            )
        else:
            rel = fig.relative_to(run_root).as_posix()
            html_parts.append(f'<div class="fig"><h3>{_html.escape(fig.name)}</h3><p><a href="{_html.escape(rel)}">Open file</a></p></div>')
    html_parts.append('</body></html>')
    gallery_path.write_text('\n'.join(html_parts), encoding='utf-8')

    # README / manifest
    readme = run_root / 'README_RUN_OUTPUTS.txt'
    readme.write_text(
        'Full analysis rerun without predictor extraction\n'
        f'Input predictor CSV: {CSV_PATH_FROM_ARGS}\n'
        f'Run folder: {run_root}\n\n'
        'Main folders:\n'
        f'- {OUT_DIR}: actual analysis outputs by objective\n'
        f'- {figures_dir}: all copied figure files together\n'
        f'- {tables_dir}: all copied table files together\n'
        f'- {files_dir}: models, JSON, text logs, and other result files\n\n'
        'Main combined files:\n'
        f'- {index_path}\n'
        f'- {workbook_path if workbook_path else "ALL_RESULTS_TABLES.xlsx was not created"}\n'
        f'- {gallery_path}\n',
        encoding='utf-8',
    )

    zip_path = run_root / f'ALL_RESULTS_PACKAGE_{run_root.name}.zip'
    try:
        with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
            for f in run_root.rglob('*'):
                if f == zip_path or not f.is_file():
                    continue
                zf.write(f, f.relative_to(run_root))
    except Exception as exc:
        (run_root / 'ZIP_PACKAGE_ERROR.txt').write_text(str(exc), encoding='utf-8')
        zip_path = None

    print('\nDONE. New full-analysis outputs are here:')
    print(run_root)
    print('\nMain collected folders:')
    print(' -', figures_dir)
    print(' -', tables_dir)
    print(' -', files_dir)
    print('\nMain files:')
    print(' -', index_path)
    if workbook_path:
        print(' -', workbook_path)
    print(' -', gallery_path)
    if zip_path:
        print(' -', zip_path)


try:
    _collect_outputs(RUN_ROOT)
except Exception as _collect_exc:
    print('Warning: final collection step failed:', _collect_exc)
    traceback.print_exc()
