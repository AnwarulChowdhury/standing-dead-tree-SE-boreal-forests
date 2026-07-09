# ============================================================
# This script starts from an already-extracted predictor CSV:
#   Task 1: nested CV model comparison
#   Task 2: SE added-value tests
#   Task 3: label-efficiency / training-size analysis
#   Task 4-5: SHAP interpretation and complementarity
#   Task 7: plot-specific SE-benefit analysis
# ============================================================

from __future__ import annotations

import argparse
import datetime as _datetime
import html as _html
import json as _json
import os
import shutil
import sys
import traceback
import zipfile
from pathlib import Path

import matplotlib
matplotlib.use("Agg")


def _parse_args():
    parser = argparse.ArgumentParser(
        description="Run"
    )
    parser.add_argument(
        "--data",
        required=True,
        help="Path to predictor CSV, e.g. data.csv",
    )
    parser.add_argument(
        "--target-col",
        default="Dead_F",
        help="Target column name. Default: Dead_F",
    )
    parser.add_argument(
        "--out-root",
        default=None,
        help="Output root folder. Default: ANALYSIS_<timestamp> beside the CSV.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Allow writing into an existing --out-root folder.",
    )
    parser.add_argument(
        "--figures",
        "--selected-figures",
        default="final",
        help=(
            "Comma-separated figure selection. Tables are always generated. "
            "Use final, all, none, model, ovp, top, shap, o3, o3se, o7, "
            "or *_all groups such as shap_all, o3_all, o7_all, importance_all. "
            "Default: final (only final manuscript figures with fixed margins)."
        ),
    )
    return parser.parse_args()


ARGS = _parse_args()
CSV_PATH_FROM_ARGS = Path(ARGS.data).expanduser().resolve()
if not CSV_PATH_FROM_ARGS.exists():
    raise FileNotFoundError(f"Predictor CSV not found: {CSV_PATH_FROM_ARGS}")

_timestamp = _datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
if ARGS.out_root:
    RUN_ROOT = Path(ARGS.out_root).expanduser().resolve()
else:
    RUN_ROOT = CSV_PATH_FROM_ARGS.parent / f"REANALYSIS_R2_{_timestamp}"

if RUN_ROOT.exists() and not ARGS.overwrite:
    raise FileExistsError(
        f"Output folder already exists: {RUN_ROOT}\n"
        "Use --overwrite or choose a new --out-root."
    )
RUN_ROOT.mkdir(parents=True, exist_ok=True)

# Keep a copy of the exact predictor CSV used for the rerun.
try:
    shutil.copy2(CSV_PATH_FROM_ARGS, RUN_ROOT / "input_predictor_matrix_used.csv")
except Exception as exc:
    print(f"Warning: could not copy input CSV into run folder: {exc}")

# Run all relative outputs inside the new analysis folder.
os.chdir(RUN_ROOT)
print("Input predictor CSV:", CSV_PATH_FROM_ARGS)
print("All new outputs will be saved in:", RUN_ROOT)

# ============================================================
# Figure selection
# ============================================================
# Tables and model outputs are always generated. This selection only controls
# which figure files are written to disk. The default is "final", meaning only
# the final manuscript-style figures with fixed margins are saved.

_FIGURE_SELECTION_RAW = str(getattr(ARGS, "figures", "final") or "final")

def _parse_figure_selection(raw: str) -> set[str]:
    parts = []
    for chunk in str(raw).replace(";", ",").split(","):
        item = chunk.strip().lower().replace("-", "_").replace(" ", "_")
        if item:
            parts.append(item)
    if not parts:
        parts = ["final"]
    aliases = {
        "selected": "final",
        "manuscript": "final",
        "manuscript_final": "final",
        "fixed": "final",
        "no": "none",
        "false": "none",
        "0": "none",
        "yes": "all",
        "true": "all",
        "1": "all",
        "condition": "o7",
        "conditions": "o7",
        "plot": "o7",
        "plots": "o7",
        "plot_variable": "o7",
        "plot_variables": "o7",
        "label": "o3",
        "label_efficiency": "o3",
        "se_gain": "o3se",
        "se_training": "o3se",
        "top_predictors": "top",
        "interaction": "shap",
        "complementarity": "shap",
        "observed_predicted": "ovp",
        "obs_pred": "ovp",
    }
    return {aliases.get(p, p) for p in parts}

_SELECTED_FIGURE_GROUPS = _parse_figure_selection(_FIGURE_SELECTION_RAW)
_SKIPPED_FIGURES_BY_SELECTION: list[str] = []
_SAVED_FIGURES_BY_SELECTION: list[str] = []

# Keywords are checked against the lower-case output filename/path.
# The short group names below are intentionally conservative: e.g. "top" saves
# only the final fixed top-predictor figure, whereas "importance_all" saves all
# diagnostic importance figures as well.
_FIGURE_GROUP_KEYWORDS = {
    "final": [
        "top_predictors_a4_exact_style_fixed",
        "shap_interaction_complementarity_a4_landscape_fixed",
        "o3_label_efficiency_3x3_a4_fixed",
        "o3_rmse_improvement_from_se_by_training_size_fixed",
        "figure_o7_split_dotplot_plot_variable_associations_fixed",
        "observed_vs_predicted_grid_a4_fixed",
    ],
    "top": ["top_predictors_a4_exact_style_fixed"],
    "shap": ["shap_interaction_complementarity_a4_landscape_fixed"],
    "o3": ["o3_label_efficiency_3x3_a4_fixed"],
    "o3se": ["o3_rmse_improvement_from_se_by_training_size_fixed"],
    "o7": ["figure_o7_split_dotplot_plot_variable_associations_fixed"],
    "ovp": ["observed_vs_predicted_grid_a4_fixed"],
    "model": [
        "figure_01_nested_tuned_model_comparison",
        "figure_02_nested_tuned_model_comparison",
        "figure_03_nested_tuned_model_comparison",
    ],
    "ovp_all": ["fig_ovp_"],
    "importance_all": [
        "figure_05_best_tuned_model_top_10_predictors",
        "figure_06_best_tuned_model_group_level_importance",
        "figure_top_",
        "figure_group_importance",
        "top_predictors_a4_exact_style",
    ],
    "o2_all": ["figure_o2_", "o2_se_added_value_heatmap"],
    "o3_all": [
        "figure_o3_label_efficiency",
        "o3_label_efficiency_rmse",
        "o3_label_efficiency_3x3_a4",
    ],
    "o3se_all": [
        "figure_o3_se_added_value_by_training_size",
        "o3_se_improvement_under_reduced_labels",
        "o3_rmse_improvement_from_se_by_training_size",
    ],
    "shap_all": [
        "figure_o4_",
        "figure_o5_",
        "algorithm_specific_shap",
        "top_predictors_a4_exact_style",
        "shap_interaction_complementarity_a4_landscape",
    ],
    "o7_all": ["figure_o7_"],
}

def _should_save_figure(fname) -> bool:
    groups = _SELECTED_FIGURE_GROUPS
    if "all" in groups:
        return True
    if "none" in groups:
        return False
    filename = Path(str(fname)).name.lower()
    full = str(fname).lower()
    for group in groups:
        for keyword in _FIGURE_GROUP_KEYWORDS.get(group, [group]):
            key = str(keyword).lower()
            if key in filename or key in full:
                return True
    return False

def _record_figure_decision(fname, saved: bool) -> None:
    try:
        s = str(fname)
        if saved:
            _SAVED_FIGURES_BY_SELECTION.append(s)
        else:
            _SKIPPED_FIGURES_BY_SELECTION.append(s)
    except Exception:
        pass

def _write_figure_selection_report() -> None:
    report = RUN_ROOT / "FIGURE_SELECTION_REPORT.txt"
    lines = [
        f"Requested --figures: {_FIGURE_SELECTION_RAW}",
        f"Normalized groups: {', '.join(sorted(_SELECTED_FIGURE_GROUPS))}",
        "",
        f"Saved figures: {len(_SAVED_FIGURES_BY_SELECTION)}",
        *[f"  SAVED: {p}" for p in _SAVED_FIGURES_BY_SELECTION],
        "",
        f"Skipped figures: {len(_SKIPPED_FIGURES_BY_SELECTION)}",
        *[f"  SKIPPED: {p}" for p in _SKIPPED_FIGURES_BY_SELECTION],
    ]
    report.write_text("\n".join(lines), encoding="utf-8")
    print("Figure selection report:", report)

print("Figure selection:", ", ".join(sorted(_SELECTED_FIGURE_GROUPS)))

# ============================================================
# Objective 1: Nested CV model development with group-wise
# feature selection and observed-vs-predicted figures for
# every model using repeated outer-CV test predictions
# ============================================================

import warnings
warnings.filterwarnings("ignore")

from pathlib import Path
import json
import joblib

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

# Robust figure saving for Windows/OneDrive paths.
# Ensures parent folders exist before every savefig call and applies --figures selection.
_ORIGINAL_PLT_SAVEFIG = plt.savefig
import matplotlib.figure as _mpl_figure
_ORIGINAL_FIGURE_SAVEFIG = _mpl_figure.Figure.savefig

def safe_savefig(fname, *args, **kwargs):
    path = Path(fname)
    if not _should_save_figure(path):
        _record_figure_decision(path, saved=False)
        print(f"Skipping figure by --figures: {path.name}")
        return None
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    _record_figure_decision(path, saved=True)
    return _ORIGINAL_PLT_SAVEFIG(str(path), *args, **kwargs)

def safe_figure_savefig(self, fname, *args, **kwargs):
    path = Path(fname)
    if not _should_save_figure(path):
        _record_figure_decision(path, saved=False)
        print(f"Skipping figure by --figures: {path.name}")
        return None
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    _record_figure_decision(path, saved=True)
    return _ORIGINAL_FIGURE_SAVEFIG(self, str(path), *args, **kwargs)

plt.savefig = safe_savefig
_mpl_figure.Figure.savefig = safe_figure_savefig

from sklearn.base import BaseEstimator, TransformerMixin, clone
from sklearn.feature_selection import f_regression
from sklearn.model_selection import (
    RepeatedStratifiedKFold,
    RepeatedKFold,
    RandomizedSearchCV
)
from sklearn.pipeline import Pipeline
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score

from xgboost import XGBRegressor


# ============================================================
# 1. Load data
# ============================================================

CSV_PATH = CSV_PATH_FROM_ARGS

OUT_DIR = RUN_ROOT / "O1"
OUT_DIR.mkdir(parents=True, exist_ok=True)

PRED_FIG_DIR = OUT_DIR / "OvP"
PRED_FIG_DIR.mkdir(parents=True, exist_ok=True)

target_col = ARGS.target_col

df = pd.read_csv(CSV_PATH)

if target_col not in df.columns:
    raise ValueError(f"Target column '{target_col}' was not found.")

# Drop target and plot identifier.
drop_cols = [target_col]

if "plot_id" in df.columns:
    drop_cols.append("plot_id")
    sample_names = df["plot_id"].astype(str).copy()
elif "Name" in df.columns:
    drop_cols.append("Name")
    sample_names = df["Name"].astype(str).copy()
else:
    sample_names = pd.Series(df.index.astype(str), index=df.index, name="Sample_Name")

X = df.drop(columns=drop_cols)
y = df[target_col]

# Keep only numeric predictors
non_numeric_cols = X.select_dtypes(exclude=[np.number]).columns.tolist()

if len(non_numeric_cols) > 0:
    print("Dropping non-numeric predictor columns:")
    print(non_numeric_cols)
    X = X.drop(columns=non_numeric_cols)

# Basic checks
if y.isna().any():
    raise ValueError("Target column contains missing values.")

if X.isna().all(axis=0).any():
    bad_cols = X.columns[X.isna().all(axis=0)].tolist()
    raise ValueError(f"These predictors are entirely missing: {bad_cols}")

if X.shape[1] == 0:
    raise ValueError("No numeric predictors available.")

print("Dataset shape:", df.shape)
print("Predictor matrix shape:", X.shape)
print("Target:", target_col)


# ============================================================
# 2. Predictor groups
# ============================================================

SE_cols = [
    c for c in X.columns
    if c.startswith("SE2025") or c.startswith("dSE")
]

CHM_cols = [
    c for c in X.columns
    if c.startswith("CHM")
]

RGBNIR_cols = [
    c for c in X.columns
    if c not in SE_cols + CHM_cols
]

print("\nPredictor groups:")
print("CHM predictors:", len(CHM_cols))
print("RGB-NIR predictors:", len(RGBNIR_cols))
print("SE predictors:", len(SE_cols))


# ============================================================
# 3. Robust repeated stratified CV for regression
# ============================================================

class RepeatedStratifiedKFoldReg:
    """
    Repeated stratified K-fold for regression.

    The continuous target is binned using quantiles.
    If valid bins cannot be created, the class falls back to RepeatedKFold.
    """

    def __init__(self, n_splits=5, n_repeats=1, n_bins=5, random_state=42):
        self.n_splits = n_splits
        self.n_repeats = n_repeats
        self.n_bins = n_bins
        self.random_state = random_state

    def _make_bins(self, y):
        y_series = pd.Series(y).reset_index(drop=True)

        max_bins = min(self.n_bins, y_series.nunique())

        for q in range(max_bins, 1, -1):
            try:
                bins = pd.qcut(
                    y_series,
                    q=q,
                    labels=False,
                    duplicates="drop"
                )

                if pd.Series(bins).isna().any():
                    continue

                bin_counts = pd.Series(bins).value_counts()

                if bin_counts.min() >= self.n_splits:
                    return bins

            except Exception:
                continue

        return None

    def split(self, X, y, groups=None):
        y_bins = self._make_bins(y)

        if y_bins is not None:
            cv = RepeatedStratifiedKFold(
                n_splits=self.n_splits,
                n_repeats=self.n_repeats,
                random_state=self.random_state
            )
            return cv.split(X, y_bins)

        print(
            "Warning: valid regression stratification was not possible. "
            "Using RepeatedKFold instead."
        )

        cv = RepeatedKFold(
            n_splits=self.n_splits,
            n_repeats=self.n_repeats,
            random_state=self.random_state
        )

        return cv.split(X, y)

    def get_n_splits(self, X=None, y=None, groups=None):
        return self.n_splits * self.n_repeats


# ============================================================
# 4. Leakage-safe group-wise selector
# ============================================================

class GroupWiseSelectKBest(BaseEstimator, TransformerMixin):
    """
    Selects top k predictors per group using f_regression.

    Leakage-safe because:
    - median imputation is fitted only on training data,
    - f_regression scores are fitted only on training data,
    - selector is inside the sklearn Pipeline.
    """

    def __init__(self, groups, k_per_group):
        self.groups = groups
        self.k_per_group = k_per_group

    def fit(self, X, y):
        if not isinstance(X, pd.DataFrame):
            raise TypeError("X must be a pandas DataFrame with column names.")

        X = X.copy()

        self.feature_names_in_ = X.columns.to_list()
        self.selected_features_ = []
        self.group_selected_features_ = {}

        self.medians_ = X.median(numeric_only=True)
        X_filled = X.fillna(self.medians_)

        for group_name, cols in self.groups.items():

            cols = [c for c in cols if c in X_filled.columns]

            if len(cols) == 0:
                self.group_selected_features_[group_name] = []
                continue

            requested_k = self.k_per_group.get(group_name, len(cols))
            k = min(requested_k, len(cols))

            scores, _ = f_regression(X_filled[cols], y)
            scores = np.nan_to_num(
                scores,
                nan=0.0,
                posinf=0.0,
                neginf=0.0
            )

            ranked_cols = (
                pd.DataFrame({
                    "feature": cols,
                    "score": scores
                })
                .sort_values("score", ascending=False)
                .head(k)["feature"]
                .tolist()
            )

            self.group_selected_features_[group_name] = ranked_cols
            self.selected_features_.extend(ranked_cols)

        if len(self.selected_features_) == 0:
            raise ValueError("No predictors were selected.")

        return self

    def transform(self, X):
        if not isinstance(X, pd.DataFrame):
            raise TypeError("X must be a pandas DataFrame with column names.")

        X = X.copy()
        X = X.fillna(self.medians_)

        return X[self.selected_features_]

    def get_feature_names_out(self, input_features=None):
        return np.array(self.selected_features_)


# ============================================================
# 5. k-per-group tuning grid
# ============================================================

k_grid = [
    {"CHM": 1, "RGBNIR": 5,  "SE": 5},
    {"CHM": 2, "RGBNIR": 10, "SE": 10},
    {"CHM": 3, "RGBNIR": 15, "SE": 15},
    {"CHM": 4, "RGBNIR": 20, "SE": 20},
    {"CHM": 5, "RGBNIR": 25, "SE": 25},
]


# ============================================================
# 6. Feature-set combinations
# ============================================================

feature_sets = {
    "CHM": {
        "CHM": CHM_cols
    },
    "RGBNIR": {
        "RGBNIR": RGBNIR_cols
    },
    "SE": {
        "SE": SE_cols
    },
    "CHM+RGBNIR": {
        "CHM": CHM_cols,
        "RGBNIR": RGBNIR_cols
    },
    "CHM+SE": {
        "CHM": CHM_cols,
        "SE": SE_cols
    },
    "RGBNIR+SE": {
        "RGBNIR": RGBNIR_cols,
        "SE": SE_cols
    },
    "CHM+RGBNIR+SE": {
        "CHM": CHM_cols,
        "RGBNIR": RGBNIR_cols,
        "SE": SE_cols
    }
}

# Remove feature sets with zero predictors
feature_sets = {
    feature_name: groups
    for feature_name, groups in feature_sets.items()
    if sum(len(cols) for cols in groups.values()) > 0
}


# ============================================================
# 7. Models and hyperparameter grids
# ============================================================

models_and_params = {
    "RF": {
        "model": RandomForestRegressor(
            random_state=42,
            n_jobs=1
        ),
        "params": {
            "selector__k_per_group": k_grid,
            "model__n_estimators": [300, 500, 800],
            "model__max_depth": [None, 5, 10, 20],
            "model__min_samples_leaf": [1, 2, 4, 6],
            "model__max_features": ["sqrt", 0.5, 0.7, 1.0],
            "model__bootstrap": [True]
        }
    },

    "GradientBoosting": {
        "model": GradientBoostingRegressor(
            random_state=42
        ),
        "params": {
            "selector__k_per_group": k_grid,
            "model__n_estimators": [200, 300, 500, 800],
            "model__learning_rate": [0.01, 0.03, 0.05, 0.1],
            "model__max_depth": [2, 3, 4],
            "model__subsample": [0.6, 0.8, 1.0],
            "model__min_samples_leaf": [1, 2, 3],
            "model__max_features": ["sqrt", 0.5, 0.7, 1.0]
        }
    },

    "XGBoost": {
        "model": XGBRegressor(
            objective="reg:squarederror",
            random_state=42,
            n_jobs=1,
            tree_method="hist"
        ),
        "params": {
            "selector__k_per_group": k_grid,
            "model__n_estimators": [200, 300, 500, 800],
            "model__learning_rate": [0.01, 0.03, 0.05, 0.1],
            "model__max_depth": [2, 3, 4, 5],
            "model__subsample": [0.6, 0.8, 1.0],
            "model__colsample_bytree": [0.6, 0.8, 1.0],
            "model__reg_alpha": [0, 0.01, 0.1, 1],
            "model__reg_lambda": [0.1, 1, 5, 10]
        }
    }
}


# ============================================================
# 8. Nested CV settings
# ============================================================

outer_cv = RepeatedStratifiedKFoldReg(
    n_splits=5,
    n_repeats=10,
    n_bins=5,
    random_state=42
)

inner_cv = RepeatedStratifiedKFoldReg(
    n_splits=5,
    n_repeats=3,
    n_bins=5,
    random_state=123
)

N_ITER = 10

all_results = []
all_predictions = []
all_best_params = []
all_selected_features = []


# ============================================================
# 9. Nested CV hyperparameter tuning
# ============================================================

for feature_name, groups in feature_sets.items():

    selected_input_cols = []

    for cols in groups.values():
        selected_input_cols.extend(cols)

    selected_input_cols = list(dict.fromkeys(selected_input_cols))

    X_sub = X[selected_input_cols].copy()

    print("\n=====================================")
    print("Feature set:", feature_name)
    print("Number of predictors:", X_sub.shape[1])
    print("=====================================")

    for model_name, model_info in models_and_params.items():

        print("\nTuning model:", model_name)

        outer_split_id = 0

        for train_idx, test_idx in outer_cv.split(X_sub, y):

            outer_split_id += 1

            X_train = X_sub.iloc[train_idx].copy()
            X_test = X_sub.iloc[test_idx].copy()

            y_train = y.iloc[train_idx].copy()
            y_test = y.iloc[test_idx].copy()

            selector = GroupWiseSelectKBest(
                groups=groups,
                k_per_group=k_grid[0]
            )

            pipe = Pipeline([
                ("selector", selector),
                ("model", clone(model_info["model"]))
            ])

            search = RandomizedSearchCV(
                estimator=pipe,
                param_distributions=model_info["params"],
                n_iter=N_ITER,
                scoring="r2",
                cv=inner_cv,
                n_jobs=-1,
                random_state=42 + outer_split_id,
                refit=True,
                error_score="raise"
            )

            # Inner CV tuning happens only inside outer training data
            search.fit(X_train, y_train)

            best_pipe = search.best_estimator_

            # Prediction is made only on outer test data
            y_pred = best_pipe.predict(X_test)

            r2 = r2_score(y_test, y_pred)
            rmse = np.sqrt(mean_squared_error(y_test, y_pred))
            mae = mean_absolute_error(y_test, y_pred)

            selected_features = (
                best_pipe
                .named_steps["selector"]
                .get_feature_names_out()
                .tolist()
            )

            all_results.append({
                "Objective": "O1_Model_development_nested_tuning",
                "Feature_Set": feature_name,
                "Model": model_name,
                "Outer_Split": outer_split_id,
                "R2": r2,
                "RMSE": rmse,
                "MAE": mae,
                "Best_Inner_R2": search.best_score_,
                "N_Selected_Features": len(selected_features)
            })

            all_best_params.append({
                "Feature_Set": feature_name,
                "Model": model_name,
                "Outer_Split": outer_split_id,
                "Best_Params_JSON": json.dumps(search.best_params_, default=str)
            })

            all_selected_features.append({
                "Feature_Set": feature_name,
                "Model": model_name,
                "Outer_Split": outer_split_id,
                "Selected_Features": ", ".join(selected_features)
            })

            # Store outer-CV test predictions
            for row_position, row_index, obs, pred in zip(
                test_idx,
                X_sub.index[test_idx],
                y_test,
                y_pred
            ):
                all_predictions.append({
                    "Feature_Set": feature_name,
                    "Model": model_name,
                    "Outer_Split": outer_split_id,
                    "Row_Position": row_position,
                    "Row_Index": row_index,
                    "Sample_Name": sample_names.iloc[row_position],
                    "Observed": obs,
                    "Predicted": pred
                })

            print(
                f"{feature_name} | {model_name} | Outer split {outer_split_id} | "
                f"RMSE={rmse:.4f}, R2={r2:.4f}, MAE={mae:.4f}"
            )


# ============================================================
# 10. Save nested CV results
# ============================================================

results_df = pd.DataFrame(all_results)
params_df = pd.DataFrame(all_best_params)
selected_features_df = pd.DataFrame(all_selected_features)
pred_all_df = pd.DataFrame(all_predictions)

results_df.to_csv(
    OUT_DIR / "O1_nested_hyperparameter_tuning_CV_results.csv",
    index=False
)

params_df.to_csv(
    OUT_DIR / "O1_nested_hyperparameter_tuning_best_params_by_outer_split.csv",
    index=False
)

selected_features_df.to_csv(
    OUT_DIR / "O1_nested_hyperparameter_tuning_selected_features_by_outer_split.csv",
    index=False
)

pred_all_df.to_csv(
    OUT_DIR / "O1_nested_hyperparameter_tuning_all_outer_test_predictions.csv",
    index=False
)


# ============================================================
# 11. Nested CV summary table
# ============================================================

summary_df = (
    results_df
    .groupby(["Feature_Set", "Model"])
    .agg(
        R2_mean=("R2", "mean"),
        R2_sd=("R2", "std"),
        RMSE_mean=("RMSE", "mean"),
        RMSE_sd=("RMSE", "std"),
        MAE_mean=("MAE", "mean"),
        MAE_sd=("MAE", "std"),
        N_Selected_Features_mean=("N_Selected_Features", "mean"),
        N_Selected_Features_sd=("N_Selected_Features", "std")
    )
    .reset_index()
    .sort_values("R2_mean", ascending=False)
)

summary_df.to_csv(
    OUT_DIR / "O1_nested_hyperparameter_tuning_summary.csv",
    index=False
)

print("\nNested CV tuning summary:")
print(summary_df)


# ============================================================
# 12. Best model from nested CV
# ============================================================

best_row = summary_df.iloc[0]

best_feature_set = best_row["Feature_Set"]
best_model_name = best_row["Model"]

print("\nBest tuned model based on nested CV R2:")
print("Feature set:", best_feature_set)
print("Model:", best_model_name)
print("Nested CV RMSE:", best_row["RMSE_mean"])
print("Nested CV R2:", best_row["R2_mean"])
print("Nested CV MAE:", best_row["MAE_mean"])


# ============================================================
# 13. Model comparison figures
# ============================================================

feature_order = [
    "CHM",
    "RGBNIR",
    "SE",
    "CHM+RGBNIR",
    "CHM+SE",
    "RGBNIR+SE",
    "CHM+RGBNIR+SE"
]

feature_order = [
    f for f in feature_order
    if f in summary_df["Feature_Set"].unique()
]

model_order = ["RF", "GradientBoosting", "XGBoost"]


def plot_metric(summary, metric_mean, metric_sd, ylabel, title, filename):

    x = np.arange(len(feature_order))
    width = 0.25

    plt.figure(figsize=(13, 6))

    for i, model in enumerate(model_order):

        data = (
            summary[summary["Model"] == model]
            .set_index("Feature_Set")
            .reindex(feature_order)
        )

        plt.bar(
            x + i * width,
            data[metric_mean],
            width,
            yerr=data[metric_sd],
            capsize=4,
            label=model
        )

    plt.xticks(x + width, feature_order, rotation=35, ha="right")
    plt.ylabel(ylabel)
    plt.xlabel("Predictor set")
    plt.title(title)
    plt.legend(title="Model")
    plt.tight_layout()
    plt.savefig(OUT_DIR / filename, dpi=300, bbox_inches="tight")
    plt.close("all")


plot_metric(
    summary_df,
    "R2_mean",
    "R2_sd",
    "Nested CV R²",
    "Nested CV tuned model comparison: R²",
    "Figure_01_nested_tuned_model_comparison_R2.png"
)

plot_metric(
    summary_df,
    "RMSE_mean",
    "RMSE_sd",
    "Nested CV RMSE",
    "Nested CV tuned model comparison: RMSE",
    "Figure_02_nested_tuned_model_comparison_RMSE.png"
)

plot_metric(
    summary_df,
    "MAE_mean",
    "MAE_sd",
    "Nested CV MAE",
    "Nested CV tuned model comparison: MAE",
    "Figure_03_nested_tuned_model_comparison_MAE.png"
)


# ============================================================
# 14. Observed vs predicted figures for every model
#     Based only on repeated outer-CV test predictions
# ============================================================

all_mean_prediction_summaries = []
all_oop_metrics = []

for feature_name in feature_order:

    for model_name in model_order:

        this_pred_repeated_df = pred_all_df[
            (pred_all_df["Feature_Set"] == feature_name) &
            (pred_all_df["Model"] == model_name)
        ].copy()

        if this_pred_repeated_df.empty:
            continue

        # Each sample has repeated outer-test predictions.
        # Average them to get one plotted prediction per sample.
        this_pred_df = (
            this_pred_repeated_df
            .groupby(["Row_Position", "Row_Index", "Sample_Name"], as_index=False)
            .agg(
                Observed=("Observed", "first"),
                Predicted_Mean=("Predicted", "mean"),
                Predicted_SD=("Predicted", "std"),
                N_Outer_Test_Predictions=("Predicted", "count")
            )
        )

        this_pred_df["Feature_Set"] = feature_name
        this_pred_df["Model"] = model_name

        all_mean_prediction_summaries.append(this_pred_df)

        outer_test_r2 = r2_score(
            this_pred_df["Observed"],
            this_pred_df["Predicted_Mean"]
        )

        outer_test_rmse = np.sqrt(
            mean_squared_error(
                this_pred_df["Observed"],
                this_pred_df["Predicted_Mean"]
            )
        )

        outer_test_mae = mean_absolute_error(
            this_pred_df["Observed"],
            this_pred_df["Predicted_Mean"]
        )

        all_oop_metrics.append({
            "Feature_Set": feature_name,
            "Model": model_name,
            "Mean_Outer_Test_R2": outer_test_r2,
            "Mean_Outer_Test_RMSE": outer_test_rmse,
            "Mean_Outer_Test_MAE": outer_test_mae,
            "N_Samples": this_pred_df.shape[0],
            "Mean_N_Outer_Test_Predictions": this_pred_df["N_Outer_Test_Predictions"].mean()
        })

        safe_feature_name = (
            feature_name
            .replace("+", "_plus_")
            .replace("/", "_")
            .replace(" ", "_")
        )

        safe_model_name = (
            model_name
            .replace("+", "_plus_")
            .replace("/", "_")
            .replace(" ", "_")
        )

        PRED_FIG_DIR.mkdir(parents=True, exist_ok=True)

        this_pred_df.to_csv(
            PRED_FIG_DIR / f"Pred_{safe_feature_name}_{safe_model_name}.csv",
            index=False
        )

        plt.figure(figsize=(6.5, 6.5))

        plt.scatter(
            this_pred_df["Observed"],
            this_pred_df["Predicted_Mean"],
            alpha=0.75,
            edgecolor="k",
            linewidth=0.3
        )

        min_val = min(
            this_pred_df["Observed"].min(),
            this_pred_df["Predicted_Mean"].min()
        )

        max_val = max(
            this_pred_df["Observed"].max(),
            this_pred_df["Predicted_Mean"].max()
        )

        plt.plot(
            [min_val, max_val],
            [min_val, max_val],
            linestyle="--",
            linewidth=1.5
        )

        plt.xlabel("Observed dead-tree fraction")
        plt.ylabel("Mean repeated outer-CV test prediction")
        plt.title(f"Observed vs predicted\n{model_name} | {feature_name}")

        plt.text(
            0.05,
            0.95,
            (
                f"R² = {outer_test_r2:.3f}\n"
                f"RMSE = {outer_test_rmse:.3f}\n"
                f"MAE = {outer_test_mae:.3f}"
            ),
            transform=plt.gca().transAxes,
            va="top",
            bbox=dict(
                boxstyle="round",
                facecolor="white",
                alpha=0.8
            )
        )

        plt.tight_layout()

        plt.savefig(
            PRED_FIG_DIR / f"Fig_OvP_{safe_feature_name}_{safe_model_name}.png",
            dpi=300,
            bbox_inches="tight"
        )

        plt.close("all")


all_mean_predictions_df = pd.concat(
    all_mean_prediction_summaries,
    ignore_index=True
)

all_mean_predictions_df.to_csv(
    OUT_DIR / "O1_all_models_mean_repeated_outer_test_predictions.csv",
    index=False
)

outer_prediction_metrics_df = pd.DataFrame(all_oop_metrics)

outer_prediction_metrics_df.to_csv(
    OUT_DIR / "O1_all_models_mean_repeated_outer_test_prediction_metrics.csv",
    index=False
)

print("\nObserved vs predicted figures saved to:")
print(PRED_FIG_DIR)


# ============================================================
# 15. Best-model observed vs predicted table
# ============================================================

best_pred_df = all_mean_predictions_df[
    (all_mean_predictions_df["Feature_Set"] == best_feature_set) &
    (all_mean_predictions_df["Model"] == best_model_name)
].copy()

best_pred_df.to_csv(
    OUT_DIR / "O1_best_model_mean_repeated_outer_test_predictions.csv",
    index=False
)

best_pred_metrics = outer_prediction_metrics_df[
    (outer_prediction_metrics_df["Feature_Set"] == best_feature_set) &
    (outer_prediction_metrics_df["Model"] == best_model_name)
].iloc[0]

mean_outer_test_r2 = best_pred_metrics["Mean_Outer_Test_R2"]
mean_outer_test_rmse = best_pred_metrics["Mean_Outer_Test_RMSE"]
mean_outer_test_mae = best_pred_metrics["Mean_Outer_Test_MAE"]


# ============================================================
# 16. Final tuning on full data for final deployable model
# ============================================================

best_groups = feature_sets[best_feature_set]

best_input_cols = []

for cols in best_groups.values():
    best_input_cols.extend(cols)

best_input_cols = list(dict.fromkeys(best_input_cols))

X_best = X[best_input_cols].copy()

final_selector = GroupWiseSelectKBest(
    groups=best_groups,
    k_per_group=k_grid[0]
)

final_pipe = Pipeline([
    ("selector", final_selector),
    ("model", clone(models_and_params[best_model_name]["model"]))
])

final_search = RandomizedSearchCV(
    estimator=final_pipe,
    param_distributions=models_and_params[best_model_name]["params"],
    n_iter=N_ITER,
    scoring="r2",
    cv=inner_cv,
    n_jobs=-1,
    random_state=999,
    refit=True,
    error_score="raise"
)

final_search.fit(X_best, y)

best_pipe = final_search.best_estimator_
best_final_params = final_search.best_params_

print("\nFinal best parameters fitted on full data:")
print(best_final_params)

selected_features = (
    best_pipe
    .named_steps["selector"]
    .get_feature_names_out()
    .tolist()
)

final_model = best_pipe.named_steps["model"]

joblib.dump(
    best_pipe,
    OUT_DIR / "O1_best_tuned_final_pipeline.joblib"
)


# ============================================================
# 17. Predictor importance for final fitted model
# ============================================================

importance_df = None
group_importance_df = None

if hasattr(final_model, "feature_importances_"):

    importance_df = pd.DataFrame({
        "Predictor": selected_features,
        "Importance": final_model.feature_importances_
    })

    def get_group_name(feature):
        if feature.startswith("CHM"):
            return "CHM"
        elif feature.startswith("SE2025") or feature.startswith("dSE"):
            return "SE"
        else:
            return "RGB-NIR"

    importance_df["Group"] = importance_df["Predictor"].apply(get_group_name)

    importance_df = importance_df.sort_values(
        "Importance",
        ascending=False
    )

    importance_df.to_csv(
        OUT_DIR / "O1_best_tuned_model_predictor_importance.csv",
        index=False
    )

    print("\nTop important predictors:")
    print(importance_df.head(20))

    top_n = min(10, len(importance_df))

    plot_df = (
        importance_df
        .head(top_n)
        .sort_values("Importance", ascending=True)
    )

    # Manuscript colours: CHM = blue, MSI/RGB-NIR = orange, SE = green
    group_colors = {
        "CHM": "#6aaed6",
        "RGB-NIR": "#e7ad72",
        "SE": "#7bd88f"
    }

    plt.figure(figsize=(8, 6))

    plt.barh(
        plot_df["Predictor"],
        plot_df["Importance"],
        color=plot_df["Group"].map(group_colors)
    )

    plt.xlabel("Built-in tree feature importance")
    plt.ylabel("Predictor")
    plt.title(
        f"Top {top_n} important predictors\n"
        f"{best_model_name} | {best_feature_set}"
    )

    from matplotlib.patches import Patch

    present_groups = plot_df["Group"].unique()

    legend_elements = [
        Patch(facecolor=group_colors[g], label=g)
        for g in present_groups
        if g in group_colors
    ]

    plt.legend(handles=legend_elements, title="Predictor group")

    plt.tight_layout()

    plt.savefig(
        OUT_DIR / "Figure_05_best_tuned_model_top_10_predictors.png",
        dpi=300,
        bbox_inches="tight"
    )

    plt.close("all")

    group_importance_df = (
        importance_df
        .groupby("Group", as_index=False)
        .agg(Total_Importance=("Importance", "sum"))
        .sort_values("Total_Importance", ascending=False)
    )

    group_importance_df.to_csv(
        OUT_DIR / "O1_best_tuned_model_group_level_importance.csv",
        index=False
    )

    plt.figure(figsize=(6, 5))

    plt.bar(
        group_importance_df["Group"],
        group_importance_df["Total_Importance"]
    )

    plt.xlabel("Predictor group")
    plt.ylabel("Total built-in feature importance")
    plt.title(
        f"Group-level predictor importance\n"
        f"{best_model_name} | {best_feature_set}"
    )

    plt.tight_layout()

    plt.savefig(
        OUT_DIR / "Figure_06_best_tuned_model_group_level_importance.png",
        dpi=300,
        bbox_inches="tight"
    )

    plt.close("all")

else:
    print("The selected final model does not support built-in feature importance.")


# ============================================================
# 18. Save best model information
# ============================================================

best_info = pd.DataFrame([{
    "Best_Feature_Set": best_feature_set,
    "Best_Model": best_model_name,

    "Nested_CV_R2_mean": best_row["R2_mean"],
    "Nested_CV_R2_sd": best_row["R2_sd"],

    "Nested_CV_RMSE_mean": best_row["RMSE_mean"],
    "Nested_CV_RMSE_sd": best_row["RMSE_sd"],

    "Nested_CV_MAE_mean": best_row["MAE_mean"],
    "Nested_CV_MAE_sd": best_row["MAE_sd"],

    "Mean_Repeated_Outer_Test_R2": mean_outer_test_r2,
    "Mean_Repeated_Outer_Test_RMSE": mean_outer_test_rmse,
    "Mean_Repeated_Outer_Test_MAE": mean_outer_test_mae,

    "Final_Best_Params_JSON": json.dumps(best_final_params, default=str),

    "Selected_Features": ", ".join(selected_features),
    "N_Selected_Features": len(selected_features),

    "Final_Pipeline_File": "O1_best_tuned_final_pipeline.joblib"
}])

best_info.to_csv(
    OUT_DIR / "O1_best_tuned_model_information.csv",
    index=False
)


# ============================================================
# 19. Finish
# ============================================================

print("\nDone. Nested hyperparameter tuning completed.")
print("\nMain output folder:")
print(OUT_DIR)

print("\nObserved-vs-predicted figures:")
print(PRED_FIG_DIR)

print("\nImportant files:")
print("1. O1_nested_hyperparameter_tuning_summary.csv")
print("2. O1_nested_hyperparameter_tuning_all_outer_test_predictions.csv")
print("3. O1_all_models_mean_repeated_outer_test_predictions.csv")
print("4. O1_all_models_mean_repeated_outer_test_prediction_metrics.csv")
print("5. O1_best_tuned_model_information.csv")
print("6. O1_best_tuned_final_pipeline.joblib")


# ============================================================
# Additional section:
# Important predictors from the best RF, XGBoost, and GradientBoosting models
# ============================================================

MODEL_IMPORTANCE_DIR = OUT_DIR / "ImpPred_BestByAlgorithm"
MODEL_IMPORTANCE_DIR.mkdir(parents=True, exist_ok=True)

all_algorithm_importances = []
all_algorithm_best_info = []

for algorithm_name in ["RF", "XGBoost", "GradientBoosting"]:

    print("\n=====================================")
    print("Final importance analysis for:", algorithm_name)
    print("=====================================")

    # --------------------------------------------------------
    # 1. Find best feature set for this algorithm
    # --------------------------------------------------------

    algorithm_summary = summary_df[
        summary_df["Model"] == algorithm_name
    ].copy()

    if algorithm_summary.empty:
        print(f"No results found for {algorithm_name}. Skipping.")
        continue

    algorithm_best_row = algorithm_summary.sort_values(
        "R2_mean",
        ascending=False
    ).iloc[0]

    algorithm_best_feature_set = algorithm_best_row["Feature_Set"]

    print("Best feature set:", algorithm_best_feature_set)
    print("Nested CV RMSE:", algorithm_best_row["RMSE_mean"])
    print("Nested CV R2:", algorithm_best_row["R2_mean"])
    print("Nested CV MAE:", algorithm_best_row["MAE_mean"])

    # --------------------------------------------------------
    # 2. Prepare predictors for this model's best feature set
    # --------------------------------------------------------

    algorithm_best_groups = feature_sets[algorithm_best_feature_set]

    algorithm_input_cols = []

    for cols in algorithm_best_groups.values():
        algorithm_input_cols.extend(cols)

    algorithm_input_cols = list(dict.fromkeys(algorithm_input_cols))

    X_algorithm = X[algorithm_input_cols].copy()

    # --------------------------------------------------------
    # 3. Final tuning on full data for this algorithm
    # --------------------------------------------------------

    algorithm_selector = GroupWiseSelectKBest(
        groups=algorithm_best_groups,
        k_per_group=k_grid[0]
    )

    algorithm_pipe = Pipeline([
        ("selector", algorithm_selector),
        ("model", clone(models_and_params[algorithm_name]["model"]))
    ])

    algorithm_search = RandomizedSearchCV(
        estimator=algorithm_pipe,
        param_distributions=models_and_params[algorithm_name]["params"],
        n_iter=N_ITER,
        scoring="r2",
        cv=inner_cv,
        n_jobs=-1,
        random_state=2025,
        refit=True,
        error_score="raise"
    )

    algorithm_search.fit(X_algorithm, y)

    algorithm_best_pipe = algorithm_search.best_estimator_
    algorithm_best_params = algorithm_search.best_params_

    selected_features_algorithm = (
        algorithm_best_pipe
        .named_steps["selector"]
        .get_feature_names_out()
        .tolist()
    )

    final_algorithm_model = algorithm_best_pipe.named_steps["model"]

    # Save final fitted pipeline for this algorithm
    safe_algorithm_name = algorithm_name.replace(" ", "_").replace("+", "_plus_")
    safe_feature_set_name = (
        algorithm_best_feature_set
        .replace(" ", "_")
        .replace("+", "_plus_")
        .replace("/", "_")
    )

    joblib.dump(
        algorithm_best_pipe,
        MODEL_IMPORTANCE_DIR / f"Final_pipeline_best_{safe_algorithm_name}_{safe_feature_set_name}.joblib"
    )

    all_algorithm_best_info.append({
        "Model": algorithm_name,
        "Best_Feature_Set": algorithm_best_feature_set,
        "Nested_CV_R2_mean": algorithm_best_row["R2_mean"],
        "Nested_CV_R2_sd": algorithm_best_row["R2_sd"],
        "Nested_CV_RMSE_mean": algorithm_best_row["RMSE_mean"],
        "Nested_CV_RMSE_sd": algorithm_best_row["RMSE_sd"],
        "Nested_CV_MAE_mean": algorithm_best_row["MAE_mean"],
        "Nested_CV_MAE_sd": algorithm_best_row["MAE_sd"],
        "Final_Best_Params_JSON": json.dumps(algorithm_best_params, default=str),
        "Selected_Features": ", ".join(selected_features_algorithm),
        "N_Selected_Features": len(selected_features_algorithm),
        "Saved_Pipeline": f"Final_pipeline_best_{safe_algorithm_name}_{safe_feature_set_name}.joblib"
    })

    # --------------------------------------------------------
    # 4. Extract built-in feature importance
    # --------------------------------------------------------

    if not hasattr(final_algorithm_model, "feature_importances_"):
        print(f"{algorithm_name} does not support built-in feature_importances_.")
        continue

    importance_algorithm_df = pd.DataFrame({
        "Model": algorithm_name,
        "Best_Feature_Set": algorithm_best_feature_set,
        "Predictor": selected_features_algorithm,
        "Importance": final_algorithm_model.feature_importances_
    })

    def get_group_name(feature):
        if feature.startswith("CHM"):
            return "CHM"
        elif feature.startswith("SE2025") or feature.startswith("dSE"):
            return "SE"
        else:
            return "RGB-NIR"

    importance_algorithm_df["Group"] = importance_algorithm_df["Predictor"].apply(
        get_group_name
    )

    importance_algorithm_df = importance_algorithm_df.sort_values(
        "Importance",
        ascending=False
    )

    all_algorithm_importances.append(importance_algorithm_df)

    # Save individual importance table
    importance_algorithm_df.to_csv(
        MODEL_IMPORTANCE_DIR / f"Predictor_importance_best_{safe_algorithm_name}_{safe_feature_set_name}.csv",
        index=False
    )

    print("\nTop 20 important predictors:")
    print(importance_algorithm_df.head(20))

    # --------------------------------------------------------
    # 5. Plot top 10 predictors for this algorithm
    # --------------------------------------------------------

    top_n = min(10, len(importance_algorithm_df))

    plot_df = (
        importance_algorithm_df
        .head(top_n)
        .sort_values("Importance", ascending=True)
    )

    # Manuscript colours: CHM = blue, MSI/RGB-NIR = orange, SE = green
    group_colors = {
        "CHM": "#6aaed6",
        "RGB-NIR": "#e7ad72",
        "SE": "#7bd88f"
    }

    plt.figure(figsize=(8, 6))

    plt.barh(
        plot_df["Predictor"],
        plot_df["Importance"],
        color=plot_df["Group"].map(group_colors)
    )

    plt.xlabel("Built-in tree feature importance")
    plt.ylabel("Predictor")
    plt.title(
        f"Top {top_n} important predictors\n"
        f"{algorithm_name} | {algorithm_best_feature_set}"
    )

    from matplotlib.patches import Patch

    present_groups = plot_df["Group"].unique()

    legend_elements = [
        Patch(facecolor=group_colors[g], label=g)
        for g in present_groups
        if g in group_colors
    ]

    plt.legend(handles=legend_elements, title="Predictor group")

    plt.tight_layout()

    plt.savefig(
        MODEL_IMPORTANCE_DIR / f"Figure_top_{top_n}_predictors_best_{safe_algorithm_name}_{safe_feature_set_name}.png",
        dpi=300,
        bbox_inches="tight"
    )

    plt.close("all")

    # --------------------------------------------------------
    # 6. Group-level importance for this algorithm
    # --------------------------------------------------------

    group_importance_algorithm_df = (
        importance_algorithm_df
        .groupby(["Model", "Best_Feature_Set", "Group"], as_index=False)
        .agg(Total_Importance=("Importance", "sum"))
        .sort_values("Total_Importance", ascending=False)
    )

    group_importance_algorithm_df.to_csv(
        MODEL_IMPORTANCE_DIR / f"Group_importance_best_{safe_algorithm_name}_{safe_feature_set_name}.csv",
        index=False
    )

    plt.figure(figsize=(6, 5))

    plt.bar(
        group_importance_algorithm_df["Group"],
        group_importance_algorithm_df["Total_Importance"]
    )

    plt.xlabel("Predictor group")
    plt.ylabel("Total built-in feature importance")
    plt.title(
        f"Group-level predictor importance\n"
        f"{algorithm_name} | {algorithm_best_feature_set}"
    )

    plt.tight_layout()

    plt.savefig(
        MODEL_IMPORTANCE_DIR / f"Figure_group_importance_best_{safe_algorithm_name}_{safe_feature_set_name}.png",
        dpi=300,
        bbox_inches="tight"
    )

    plt.close("all")


# ============================================================
# Save combined importance outputs for all three best models
# ============================================================

if len(all_algorithm_importances) > 0:

    all_algorithm_importance_df = pd.concat(
        all_algorithm_importances,
        ignore_index=True
    )

    all_algorithm_importance_df.to_csv(
        MODEL_IMPORTANCE_DIR / "All_best_algorithm_predictor_importances.csv",
        index=False
    )

    print("\nCombined predictor importance table saved.")

else:
    print("\nNo predictor importance tables were created.")


if len(all_algorithm_best_info) > 0:

    all_algorithm_best_info_df = pd.DataFrame(all_algorithm_best_info)

    all_algorithm_best_info_df.to_csv(
        MODEL_IMPORTANCE_DIR / "Best_model_information_by_algorithm.csv",
        index=False
    )

    print("\nBest model information by algorithm saved.")


print("\nImportant predictor outputs saved to:")
print(MODEL_IMPORTANCE_DIR)

# CELL

# ============================================================
# Objective 2:
# Added-value assessment of SE predictors
# Comparing paired outer-CV performance with and without SE
# ============================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import wilcoxon
from pathlib import Path


# ------------------------------------------------------------
# 1. Output folder
# ------------------------------------------------------------

try:
    OUT_DIR
except NameError:
    OUT_DIR = Path(r"C:/Users/Dead_tree_fraction")
    OUT_DIR.mkdir(parents=True, exist_ok=True)

O2_DIR = OUT_DIR / "O2_SE_added_value"
O2_DIR.mkdir(parents=True, exist_ok=True)


# ------------------------------------------------------------
# 2. Detect outer CV split column
# ------------------------------------------------------------

if "Outer_Split" in results_df.columns:
    fold_col = "Outer_Split"
elif "Outer_CV_Run" in results_df.columns:
    fold_col = "Outer_CV_Run"
elif "CV_Fold" in results_df.columns:
    fold_col = "CV_Fold"
else:
    raise ValueError(
        "No outer CV split column found. Expected one of: "
        "'Outer_Split', 'Outer_CV_Run', or 'CV_Fold'. "
        f"Available columns are: {results_df.columns.tolist()}"
    )

print("Using outer CV split column:", fold_col)


# ------------------------------------------------------------
# 3. Define model order
# ------------------------------------------------------------

if "model_order" not in globals():
    model_order = ["RF", "GradientBoosting", "XGBoost"]


# ------------------------------------------------------------
# 4. Define SE added-value comparisons
# ------------------------------------------------------------

se_comparisons = {
    "CHM_to_CHM+SE": {
        "Without_SE": "CHM",
        "With_SE": "CHM+SE"
    },
    "RGBNIR_to_RGBNIR+SE": {
        "Without_SE": "RGBNIR",
        "With_SE": "RGBNIR+SE"
    },
    "CHM+RGBNIR_to_CHM+RGBNIR+SE": {
        "Without_SE": "CHM+RGBNIR",
        "With_SE": "CHM+RGBNIR+SE"
    }
}


# ------------------------------------------------------------
# 5. Paired SE added-value calculation
# ------------------------------------------------------------

added_value_rows = []
paired_delta_rows = []

for model_name in model_order:

    for comparison_name, comparison in se_comparisons.items():

        without_set = comparison["Without_SE"]
        with_set = comparison["With_SE"]

        without_df = results_df[
            (results_df["Feature_Set"] == without_set) &
            (results_df["Model"] == model_name)
        ][[fold_col, "R2", "RMSE", "MAE"]].copy()

        with_df = results_df[
            (results_df["Feature_Set"] == with_set) &
            (results_df["Model"] == model_name)
        ][[fold_col, "R2", "RMSE", "MAE"]].copy()

        if without_df.empty or with_df.empty:
            print(f"Skipping missing comparison: {model_name} - {comparison_name}")
            continue

        paired_df = without_df.merge(
            with_df,
            on=fold_col,
            suffixes=("_without_SE", "_with_SE")
        )

        if paired_df.shape[0] != without_df.shape[0] or paired_df.shape[0] != with_df.shape[0]:
            raise ValueError(
                f"Fold pairing mismatch for {model_name} - {comparison_name}: "
                f"without SE = {without_df.shape[0]}, "
                f"with SE = {with_df.shape[0]}, "
                f"paired = {paired_df.shape[0]}"
            )

        paired_df["Model"] = model_name
        paired_df["Comparison"] = comparison_name
        paired_df["Without_SE"] = without_set
        paired_df["With_SE"] = with_set

        paired_df["Delta_R2"] = (
            paired_df["R2_with_SE"] -
            paired_df["R2_without_SE"]
        )

        paired_df["Delta_RMSE"] = (
            paired_df["RMSE_without_SE"] -
            paired_df["RMSE_with_SE"]
        )

        paired_df["Delta_MAE"] = (
            paired_df["MAE_without_SE"] -
            paired_df["MAE_with_SE"]
        )

        paired_df["RMSE_percent_improvement"] = (
            paired_df["Delta_RMSE"] /
            paired_df["RMSE_without_SE"]
        ) * 100

        paired_df["MAE_percent_improvement"] = (
            paired_df["Delta_MAE"] /
            paired_df["MAE_without_SE"]
        ) * 100

        # Store split-level deltas
        paired_delta_rows.append(
            paired_df[
                [
                    fold_col,
                    "Model",
                    "Comparison",
                    "Without_SE",
                    "With_SE",
                    "R2_without_SE",
                    "R2_with_SE",
                    "Delta_R2",
                    "RMSE_without_SE",
                    "RMSE_with_SE",
                    "Delta_RMSE",
                    "RMSE_percent_improvement",
                    "MAE_without_SE",
                    "MAE_with_SE",
                    "Delta_MAE",
                    "MAE_percent_improvement"
                ]
            ]
        )

        # Wilcoxon paired tests
        try:
            p_r2 = wilcoxon(
                paired_df["R2_without_SE"],
                paired_df["R2_with_SE"],
                zero_method="wilcox"
            ).pvalue
        except ValueError:
            p_r2 = np.nan

        try:
            p_rmse = wilcoxon(
                paired_df["RMSE_without_SE"],
                paired_df["RMSE_with_SE"],
                zero_method="wilcox"
            ).pvalue
        except ValueError:
            p_rmse = np.nan

        try:
            p_mae = wilcoxon(
                paired_df["MAE_without_SE"],
                paired_df["MAE_with_SE"],
                zero_method="wilcox"
            ).pvalue
        except ValueError:
            p_mae = np.nan

        added_value_rows.append({
            "Model": model_name,
            "Comparison": comparison_name,
            "Without_SE": without_set,
            "With_SE": with_set,

            "N_Paired_Outer_Splits": paired_df.shape[0],

            "R2_without_SE_mean": paired_df["R2_without_SE"].mean(),
            "R2_with_SE_mean": paired_df["R2_with_SE"].mean(),
            "Delta_R2_mean": paired_df["Delta_R2"].mean(),
            "Delta_R2_sd": paired_df["Delta_R2"].std(ddof=1),
            "Delta_R2_median": paired_df["Delta_R2"].median(),
            "Wilcoxon_p_R2": p_r2,

            "RMSE_without_SE_mean": paired_df["RMSE_without_SE"].mean(),
            "RMSE_with_SE_mean": paired_df["RMSE_with_SE"].mean(),
            "Delta_RMSE_mean": paired_df["Delta_RMSE"].mean(),
            "Delta_RMSE_sd": paired_df["Delta_RMSE"].std(ddof=1),
            "Delta_RMSE_median": paired_df["Delta_RMSE"].median(),
            "RMSE_percent_improvement_mean": paired_df["RMSE_percent_improvement"].mean(),
            "RMSE_percent_improvement_median": paired_df["RMSE_percent_improvement"].median(),
            "Wilcoxon_p_RMSE": p_rmse,

            "MAE_without_SE_mean": paired_df["MAE_without_SE"].mean(),
            "MAE_with_SE_mean": paired_df["MAE_with_SE"].mean(),
            "Delta_MAE_mean": paired_df["Delta_MAE"].mean(),
            "Delta_MAE_sd": paired_df["Delta_MAE"].std(ddof=1),
            "Delta_MAE_median": paired_df["Delta_MAE"].median(),
            "MAE_percent_improvement_mean": paired_df["MAE_percent_improvement"].mean(),
            "MAE_percent_improvement_median": paired_df["MAE_percent_improvement"].median(),
            "Wilcoxon_p_MAE": p_mae
        })


# ------------------------------------------------------------
# 6. Save Objective 2 results
# ------------------------------------------------------------

added_value_df = pd.DataFrame(added_value_rows)

if len(paired_delta_rows) > 0:
    paired_delta_df = pd.concat(paired_delta_rows, ignore_index=True)
else:
    paired_delta_df = pd.DataFrame()

print("\nObjective 2: SE added-value summary")
print(added_value_df)

added_value_df.to_csv(
    O2_DIR / "O2_SE_added_value_summary.csv",
    index=False
)

paired_delta_df.to_csv(
    O2_DIR / "O2_SE_added_value_paired_outer_split_deltas.csv",
    index=False
)

with pd.ExcelWriter(O2_DIR / "O2_SE_added_value_results.xlsx") as writer:
    added_value_df.to_excel(writer, sheet_name="Summary", index=False)
    paired_delta_df.to_excel(writer, sheet_name="Paired_Deltas", index=False)

print("\nSaved Objective 2 files to:")
print(O2_DIR)

# CELL

# ============================================================
# Figures: Added value of SE predictors
# ============================================================

from pathlib import Path
import numpy as np
import matplotlib.pyplot as plt

# Output folder
try:
    O2_DIR
except NameError:
    O2_DIR = Path(".")
else:
    O2_DIR = Path(O2_DIR)

# Detect correct RMSE improvement column
if "RMSE_percent_improvement_mean" in added_value_df.columns:
    rmse_improvement_col = "RMSE_percent_improvement_mean"
elif "RMSE_percent_improvement" in added_value_df.columns:
    rmse_improvement_col = "RMSE_percent_improvement"
else:
    raise ValueError(
        "No RMSE improvement column found. Expected "
        "'RMSE_percent_improvement_mean' or 'RMSE_percent_improvement'."
    )

# Sort for cleaner plotting
plot_df = added_value_df.copy()
plot_df["Plot_Label"] = plot_df["Comparison"] + "\n" + plot_df["Model"]

model_order = ["RF", "GradientBoosting", "XGBoost"]
plot_df["Model"] = pd.Categorical(plot_df["Model"], categories=model_order, ordered=True)

plot_df = plot_df.sort_values(["Comparison", "Model"]).reset_index(drop=True)

x = np.arange(len(plot_df))


# ============================================================
# Figure 1: RMSE improvement from adding SE predictors
# ============================================================

plt.figure(figsize=(12, 6))

plt.bar(
    x,
    plot_df[rmse_improvement_col]
)

plt.axhline(0, linestyle="--", linewidth=1)

plt.xticks(x, plot_df["Plot_Label"], rotation=45, ha="right")
plt.ylabel("RMSE improvement after adding SE (%)")
plt.xlabel("Comparison and model")
plt.title("Added value of SE predictors based on RMSE reduction")

plt.tight_layout()

plt.savefig(
    O2_DIR / "Figure_O2_SE_added_value_RMSE_improvement.png",
    dpi=300,
    bbox_inches="tight"
)

plt.close("all")


# ============================================================
# Figure 2: R² improvement from adding SE predictors
# ============================================================

if "Delta_R2_mean" not in plot_df.columns:
    raise ValueError("Column 'Delta_R2_mean' was not found in added_value_df.")

plt.figure(figsize=(12, 6))

plt.bar(
    x,
    plot_df["Delta_R2_mean"]
)

plt.axhline(0, linestyle="--", linewidth=1)

plt.xticks(x, plot_df["Plot_Label"], rotation=45, ha="right")
plt.ylabel("Change in cross-validated R² after adding SE")
plt.xlabel("Comparison and model")
plt.title("Added value of SE predictors based on R² improvement")

plt.tight_layout()

plt.savefig(
    O2_DIR / "Figure_O2_SE_added_value_R2_improvement.png",
    dpi=300,
    bbox_inches="tight"
)

plt.close("all")

# CELL

# ============================================================
# Objective 3:
# Label-efficiency evaluation
# Assess whether SE predictors improve performance
# under reduced training-data scenarios
# ============================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

from pathlib import Path
from sklearn.base import clone
from sklearn.model_selection import StratifiedShuffleSplit
from sklearn.pipeline import Pipeline
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error


# ------------------------------------------------------------
# 1. Output folder
# ------------------------------------------------------------

try:
    OUT_DIR
except NameError:
    OUT_DIR = Path(r"C:/Users/Dead_tree_fraction")
    OUT_DIR.mkdir(parents=True, exist_ok=True)

O3_DIR = OUT_DIR / "O3_label_efficiency"
O3_DIR.mkdir(parents=True, exist_ok=True)


# ------------------------------------------------------------
# 2. Training-data proportions
# ------------------------------------------------------------

train_proportions = [0.2, 0.4, 0.6, 0.8, 1.0]
n_repeats = 30


# ------------------------------------------------------------
# 3. Create target bins for stratified splitting
# ------------------------------------------------------------

def make_regression_bins(y, n_bins=5):
    """
    Create quantile bins for stratified regression splitting.
    Automatically reduces number of bins if needed.
    """

    y_series = pd.Series(y).reset_index(drop=True)

    max_bins = min(n_bins, y_series.nunique())

    for q in range(max_bins, 1, -1):
        try:
            bins = pd.qcut(
                y_series,
                q=q,
                labels=False,
                duplicates="drop"
            )

            if pd.Series(bins).isna().any():
                continue

            return pd.Series(bins, index=y.index)

        except Exception:
            continue

    raise ValueError("Could not create valid target bins for stratified splitting.")


y_bins = make_regression_bins(y, n_bins=5)


# ------------------------------------------------------------
# 4. Main comparison sets for label efficiency
# ------------------------------------------------------------

label_efficiency_feature_sets = {
    "CHM+RGBNIR": feature_sets["CHM+RGBNIR"],
    "CHM+RGBNIR+SE": feature_sets["CHM+RGBNIR+SE"]
}


# ------------------------------------------------------------
# 5. Select models
# ------------------------------------------------------------

# Use the same base models from Objective 1
label_efficiency_models = {
    model_name: model_info["model"]
    for model_name, model_info in models_and_params.items()
}

# Option: only use the best overall model
# label_efficiency_models = {
#     best_model_name: models_and_params[best_model_name]["model"]
# }


# ------------------------------------------------------------
# 6. Fixed k-per-group for label-efficiency experiment
# ------------------------------------------------------------

# Recommended: use a fixed selector size to isolate the effect of sample size.
# You can change this to another entry from k_grid.
fixed_k_per_group = {"CHM": 5, "RGBNIR": 25, "SE": 25}

# Alternative:
# fixed_k_per_group = k_grid[-1]


# ------------------------------------------------------------
# 7. Prepare fixed independent test split
# ------------------------------------------------------------

test_splitter = StratifiedShuffleSplit(
    n_splits=1,
    test_size=0.2,
    random_state=42
)

train_full_idx, test_idx = next(
    test_splitter.split(X, y_bins)
)

X_train_full = X.iloc[train_full_idx].copy()
y_train_full = y.iloc[train_full_idx].copy()
y_bins_train_full = y_bins.iloc[train_full_idx].copy()

X_test_full = X.iloc[test_idx].copy()
y_test = y.iloc[test_idx].copy()

print("Full training pool size:", X_train_full.shape[0])
print("Independent test size:", X_test_full.shape[0])


# ------------------------------------------------------------
# 8. Run label-efficiency experiment
# ------------------------------------------------------------

label_eff_rows = []

for train_prop in train_proportions:

    print("\nTraining proportion:", train_prop)

    # At 100%, there is only one unique training subset.
    # Repeating it 30 times would produce duplicate results for deterministic models.
    if train_prop == 1.0:
        repeat_range = range(1)
    else:
        repeat_range = range(n_repeats)

    for repeat in repeat_range:

        if train_prop < 1.0:

            sub_splitter = StratifiedShuffleSplit(
                n_splits=1,
                train_size=train_prop,
                random_state=1000 + repeat
            )

            sub_train_idx_relative, _ = next(
                sub_splitter.split(X_train_full, y_bins_train_full)
            )

            sub_train_idx = X_train_full.index[sub_train_idx_relative]

        else:
            sub_train_idx = X_train_full.index

        for feature_name, groups in label_efficiency_feature_sets.items():

            selected_input_cols = []

            for cols in groups.values():
                selected_input_cols.extend(cols)

            selected_input_cols = list(dict.fromkeys(selected_input_cols))

            X_train_sub = X.loc[sub_train_idx, selected_input_cols].copy()
            y_train_sub = y.loc[sub_train_idx].copy()

            X_test_sub = X_test_full[selected_input_cols].copy()

            for model_name, model in label_efficiency_models.items():

                selector = GroupWiseSelectKBest(
                    groups=groups,
                    k_per_group=fixed_k_per_group
                )

                pipe = Pipeline([
                    ("selector", selector),
                    ("model", clone(model))
                ])

                pipe.fit(X_train_sub, y_train_sub)

                y_pred = pipe.predict(X_test_sub)

                r2 = r2_score(y_test, y_pred)
                rmse = np.sqrt(mean_squared_error(y_test, y_pred))
                mae = mean_absolute_error(y_test, y_pred)

                label_eff_rows.append({
                    "Objective": "O3_Label_efficiency",
                    "Train_Proportion": train_prop,
                    "Train_N": len(y_train_sub),
                    "Feature_Set": feature_name,
                    "Model": model_name,
                    "Repeat": repeat + 1,
                    "R2": r2,
                    "RMSE": rmse,
                    "MAE": mae
                })

                print(
                    f"{feature_name} | {model_name} | "
                    f"Train prop={train_prop} | Repeat={repeat + 1} | "
                    f"RMSE={rmse:.4f}, R2={r2:.4f}"
                )


# ------------------------------------------------------------
# 9. Save raw label-efficiency results
# ------------------------------------------------------------

label_eff_df = pd.DataFrame(label_eff_rows)

label_eff_df.to_csv(
    O3_DIR / "O3_label_efficiency_results.csv",
    index=False
)

print("\nLabel-efficiency results:")
print(label_eff_df.head())

print("\nSaved:")
print(O3_DIR / "O3_label_efficiency_results.csv")


# ------------------------------------------------------------
# 10. Summary table
# ------------------------------------------------------------

label_eff_summary_df = (
    label_eff_df
    .groupby(["Train_Proportion", "Feature_Set", "Model"])
    .agg(
        R2_mean=("R2", "mean"),
        R2_sd=("R2", "std"),
        RMSE_mean=("RMSE", "mean"),
        RMSE_sd=("RMSE", "std"),
        MAE_mean=("MAE", "mean"),
        MAE_sd=("MAE", "std"),
        Train_N_mean=("Train_N", "mean")
    )
    .reset_index()
)

label_eff_summary_df.to_csv(
    O3_DIR / "O3_label_efficiency_summary.csv",
    index=False
)

print("\nLabel-efficiency summary:")
print(label_eff_summary_df)


# ------------------------------------------------------------
# 11. Calculate SE added value at each training proportion
# ------------------------------------------------------------

label_eff_added_rows = []

for model_name in label_eff_df["Model"].unique():

    for train_prop in train_proportions:

        without_se_df = label_eff_df[
            (label_eff_df["Model"] == model_name) &
            (label_eff_df["Train_Proportion"] == train_prop) &
            (label_eff_df["Feature_Set"] == "CHM+RGBNIR")
        ].copy()

        with_se_df = label_eff_df[
            (label_eff_df["Model"] == model_name) &
            (label_eff_df["Train_Proportion"] == train_prop) &
            (label_eff_df["Feature_Set"] == "CHM+RGBNIR+SE")
        ].copy()

        paired_df = without_se_df.merge(
            with_se_df,
            on=["Model", "Train_Proportion", "Repeat"],
            suffixes=("_without_SE", "_with_SE")
        )

        if paired_df.empty:
            continue

        paired_df["Delta_R2"] = (
            paired_df["R2_with_SE"] -
            paired_df["R2_without_SE"]
        )

        paired_df["Delta_RMSE"] = (
            paired_df["RMSE_without_SE"] -
            paired_df["RMSE_with_SE"]
        )

        paired_df["Delta_MAE"] = (
            paired_df["MAE_without_SE"] -
            paired_df["MAE_with_SE"]
        )

        paired_df["RMSE_percent_improvement"] = (
            paired_df["Delta_RMSE"] /
            paired_df["RMSE_without_SE"]
        ) * 100

        paired_df["MAE_percent_improvement"] = (
            paired_df["Delta_MAE"] /
            paired_df["MAE_without_SE"]
        ) * 100

        label_eff_added_rows.append({
            "Model": model_name,
            "Train_Proportion": train_prop,
            "N_Paired_Repeats": paired_df.shape[0],

            "Delta_R2_mean": paired_df["Delta_R2"].mean(),
            "Delta_R2_sd": paired_df["Delta_R2"].std(ddof=1),

            "Delta_RMSE_mean": paired_df["Delta_RMSE"].mean(),
            "Delta_RMSE_sd": paired_df["Delta_RMSE"].std(ddof=1),
            "RMSE_percent_improvement_mean": paired_df["RMSE_percent_improvement"].mean(),
            "RMSE_percent_improvement_sd": paired_df["RMSE_percent_improvement"].std(ddof=1),

            "Delta_MAE_mean": paired_df["Delta_MAE"].mean(),
            "Delta_MAE_sd": paired_df["Delta_MAE"].std(ddof=1),
            "MAE_percent_improvement_mean": paired_df["MAE_percent_improvement"].mean(),
            "MAE_percent_improvement_sd": paired_df["MAE_percent_improvement"].std(ddof=1)
        })

label_eff_added_df = pd.DataFrame(label_eff_added_rows)

label_eff_added_df.to_csv(
    O3_DIR / "O3_label_efficiency_SE_added_value_by_training_size.csv",
    index=False
)

print("\nSE added value by training proportion:")
print(label_eff_added_df)


# ------------------------------------------------------------
# 12. Plot RMSE learning curves
# ------------------------------------------------------------

for model_name in label_eff_summary_df["Model"].unique():

    plot_df = label_eff_summary_df[
        label_eff_summary_df["Model"] == model_name
    ].copy()

    plt.figure(figsize=(8, 6))

    for feature_name in ["CHM+RGBNIR", "CHM+RGBNIR+SE"]:

        data = plot_df[
            plot_df["Feature_Set"] == feature_name
        ].sort_values("Train_Proportion")

        plt.errorbar(
            data["Train_Proportion"] * 100,
            data["RMSE_mean"],
            yerr=data["RMSE_sd"],
            marker="o",
            capsize=4,
            label=feature_name
        )

    plt.xlabel("Training data used (%)")
    plt.ylabel("Test RMSE")
    plt.title(f"Label-efficiency learning curve: RMSE\n{model_name}")
    plt.legend(title="Predictor set")
    plt.tight_layout()

    safe_model_name = model_name.replace(" ", "_").replace("+", "_plus_")

    plt.savefig(
        O3_DIR / f"Figure_O3_label_efficiency_RMSE_{safe_model_name}.png",
        dpi=300,
        bbox_inches="tight"
    )

    plt.close("all")


# ------------------------------------------------------------
# 13. Plot SE added-value by training size
# ------------------------------------------------------------

for model_name in label_eff_added_df["Model"].unique():

    plot_df = label_eff_added_df[
        label_eff_added_df["Model"] == model_name
    ].sort_values("Train_Proportion")

    plt.figure(figsize=(8, 6))

    plt.errorbar(
        plot_df["Train_Proportion"] * 100,
        plot_df["RMSE_percent_improvement_mean"],
        yerr=plot_df["RMSE_percent_improvement_sd"],
        marker="o",
        capsize=4
    )

    plt.axhline(0, linestyle="--", linewidth=1)

    plt.xlabel("Training data used (%)")
    plt.ylabel("RMSE improvement from adding SE (%)")
    plt.title(f"SE added value under reduced training data\n{model_name}")

    plt.tight_layout()

    safe_model_name = model_name.replace(" ", "_").replace("+", "_plus_")

    plt.savefig(
        O3_DIR / f"Figure_O3_SE_added_value_by_training_size_{safe_model_name}.png",
        dpi=300,
        bbox_inches="tight"
    )

    plt.close("all")

# CELL

# ============================================================
# O3 summary table
# ============================================================

label_eff_summary = (
    label_eff_df
    .groupby(["Train_Proportion", "Feature_Set", "Model"])
    .agg(
        Train_N_mean=("Train_N", "mean"),
        R2_mean=("R2", "mean"),
        R2_sd=("R2", "std"),
        RMSE_mean=("RMSE", "mean"),
        RMSE_sd=("RMSE", "std"),
        MAE_mean=("MAE", "mean"),
        MAE_sd=("MAE", "std")
    )
    .reset_index()
)

label_eff_summary.to_csv(
    "O3_label_efficiency_summary.csv",
    index=False
)

print(label_eff_summary)


import os

output_folder = r"C:/Users/Dead_tree_fraction"
os.makedirs(output_folder, exist_ok=True)

excel_path = os.path.join(output_folder, "O3_label_efficiency_summary.xlsx")

label_eff_summary.to_excel(excel_path, index=False)

print("Saved Excel file:")
print(excel_path)

# CELL

# ============================================================
# Figure: label efficiency based on RMSE
# ============================================================

for model_name in label_efficiency_models.keys():

    plt.figure(figsize=(7, 5))

    for feature_name in label_efficiency_feature_sets.keys():

        plot_df = label_eff_summary[
            (label_eff_summary["Model"] == model_name) &
            (label_eff_summary["Feature_Set"] == feature_name)
        ].sort_values("Train_Proportion")

        plt.errorbar(
            plot_df["Train_Proportion"] * 100,
            plot_df["RMSE_mean"],
            yerr=plot_df["RMSE_sd"],
            marker="o",
            capsize=4,
            label=feature_name
        )

    plt.xlabel("Training data used (%)")
    plt.ylabel("RMSE")
    plt.title(f"Label efficiency based on RMSE: {model_name}")
    plt.legend(title="Predictor set")
    plt.tight_layout()

    plt.savefig(
        f"O3_label_efficiency_RMSE_{model_name}.png",
        dpi=300,
        bbox_inches="tight"
    )

    plt.close("all")

# CELL

# ============================================================
# Figure: label efficiency based on R2
# ============================================================

from pathlib import Path
import matplotlib.pyplot as plt

# Output folder
try:
    O3_DIR
except NameError:
    O3_DIR = Path(".")
else:
    O3_DIR = Path(O3_DIR)

for model_name in label_efficiency_models.keys():

    plt.figure(figsize=(7, 5))

    has_data = False

    for feature_name in label_efficiency_feature_sets.keys():

        plot_df = label_eff_summary_df[
            (label_eff_summary_df["Model"] == model_name) &
            (label_eff_summary_df["Feature_Set"] == feature_name)
        ].sort_values("Train_Proportion")

        if plot_df.empty:
            continue

        has_data = True

        plt.errorbar(
            plot_df["Train_Proportion"] * 100,
            plot_df["R2_mean"],
            yerr=plot_df["R2_sd"],
            marker="o",
            capsize=4,
            label=feature_name
        )

    if not has_data:
        plt.close()
        continue

    plt.xlabel("Training data used (%)")
    plt.ylabel("Test R²")
    plt.title(f"Label efficiency based on R²\n{model_name}")
    plt.legend(title="Predictor set")
    plt.tight_layout()

    safe_model_name = model_name.replace(" ", "_").replace("+", "_plus_")

    plt.savefig(
        O3_DIR / f"Figure_O3_label_efficiency_R2_{safe_model_name}.png",
        dpi=300,
        bbox_inches="tight"
    )

    plt.close("all")

# CELL

# ============================================================
# SE gain by training-data proportion
# ============================================================

gain_rows = []

for model_name in label_efficiency_models.keys():

    for train_prop in train_proportions:

        base_df = label_eff_df[
            (label_eff_df["Model"] == model_name) &
            (label_eff_df["Train_Proportion"] == train_prop) &
            (label_eff_df["Feature_Set"] == "CHM+RGBNIR")
        ].sort_values("Repeat")

        se_df = label_eff_df[
            (label_eff_df["Model"] == model_name) &
            (label_eff_df["Train_Proportion"] == train_prop) &
            (label_eff_df["Feature_Set"] == "CHM+RGBNIR+SE")
        ].sort_values("Repeat")

        delta_rmse = base_df["RMSE"].values - se_df["RMSE"].values
        delta_r2 = se_df["R2"].values - base_df["R2"].values
        delta_mae = base_df["MAE"].values - se_df["MAE"].values

        gain_rows.append({
            "Model": model_name,
            "Train_Proportion": train_prop,
            "Delta_RMSE_mean": np.mean(delta_rmse),
            "Delta_RMSE_sd": np.std(delta_rmse),
            "RMSE_improvement_percent": (
                np.mean(delta_rmse) / base_df["RMSE"].mean()
            ) * 100,
            "Delta_R2_mean": np.mean(delta_r2),
            "Delta_R2_sd": np.std(delta_r2),
            "Delta_MAE_mean": np.mean(delta_mae),
            "Delta_MAE_sd": np.std(delta_mae)
        })

label_eff_gain_df = pd.DataFrame(gain_rows)

label_eff_gain_df.to_csv(
    "O3_SE_gain_by_training_fraction.csv",
    index=False
)

print(label_eff_gain_df)

# CELL

# ============================================================
# Figure: SE improvement under reduced labels
# ============================================================

plt.figure(figsize=(8, 5))

for model_name in label_efficiency_models.keys():

    plot_df = label_eff_gain_df[
        label_eff_gain_df["Model"] == model_name
    ].sort_values("Train_Proportion")

    plt.plot(
        plot_df["Train_Proportion"] * 100,
        plot_df["RMSE_improvement_percent"],
        marker="o",
        label=model_name
    )

plt.axhline(0, linestyle="--", linewidth=1)

plt.xlabel("Training data used (%)", fontsize=10)
plt.ylabel("RMSE improvement from SE (%)", fontsize=10)
plt.title("", fontsize=10)

plt.xticks(fontsize=10)
plt.yticks(fontsize=10)

plt.legend(title="Model", fontsize=10, title_fontsize=10)

plt.tight_layout()

plt.savefig(
    "O3_SE_improvement_under_reduced_labels.png",
    dpi=300,
    bbox_inches="tight"
)

plt.close("all")

# CELL

# Switch bare SHAP/Objective-5 output filenames into their own folder.
O4_DIR = OUT_DIR / "O4_SHAP_and_complementarity"
O4_DIR.mkdir(parents=True, exist_ok=True)
os.chdir(O4_DIR)
print("Objective 4/5 outputs will be saved to:", O4_DIR)

# ============================================================
# Objective 4:
# Model interpretability using SHAP
# Evaluate importance of individual predictors and groups
# ============================================================

import shap
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.base import clone
from sklearn.pipeline import Pipeline


# ------------------------------------------------------------
# 1. Rebuild and fit best model from Objective 1
# ------------------------------------------------------------

best_groups = feature_sets[best_feature_set]

best_input_cols = []
for cols in best_groups.values():
    best_input_cols.extend(cols)

best_input_cols = list(dict.fromkeys(best_input_cols))

X_best = X[best_input_cols]

# Use the best tuned k_per_group if available.
# Otherwise use the first/default k_grid setting.
if "best_final_params" in globals() and "selector__k_per_group" in best_final_params:
    best_k_per_group = best_final_params["selector__k_per_group"]
else:
    best_k_per_group = k_grid[0]

best_selector = GroupWiseSelectKBest(
    groups=best_groups,
    k_per_group=best_k_per_group
)

# Use current Objective 1 model dictionary
best_model = clone(models_and_params[best_model_name]["model"])

best_pipe = Pipeline([
    ("selector", best_selector),
    ("model", best_model)
])

best_pipe.fit(X_best, y)


# ------------------------------------------------------------
# 2. Get selected predictors
# ------------------------------------------------------------

selected_features = best_pipe.named_steps["selector"].get_feature_names_out()

X_selected = best_pipe.named_steps["selector"].transform(X_best)
X_selected = pd.DataFrame(X_selected, columns=selected_features)

final_model = best_pipe.named_steps["model"]

print("Best model:", best_model_name)
print("Best feature set:", best_feature_set)
print("k_per_group used:", best_k_per_group)
print("Number of selected predictors:", len(selected_features))


# ------------------------------------------------------------
# 3. Compute SHAP values
# ------------------------------------------------------------

explainer = shap.TreeExplainer(final_model)
shap_values = explainer.shap_values(X_selected)

if isinstance(shap_values, list):
    shap_matrix = shap_values[0]
else:
    shap_matrix = np.array(shap_values)

if len(shap_matrix.shape) == 3:
    shap_matrix = shap_matrix[:, :, 0]


# ------------------------------------------------------------
# 4. SHAP importance table
# ------------------------------------------------------------

shap_importance_df = pd.DataFrame({
    "Predictor": selected_features,
    "Mean_abs_SHAP": np.abs(shap_matrix).mean(axis=0),
    "SD_abs_SHAP": np.abs(shap_matrix).std(axis=0)
})


def get_group_name(feature):
    if feature.startswith("CHM"):
        return "CHM"
    elif feature.startswith("SE2025") or feature.startswith("dSE"):
        return "SE"
    else:
        return "RGB-NIR"


shap_importance_df["Group"] = shap_importance_df["Predictor"].apply(get_group_name)

shap_importance_df = shap_importance_df.sort_values(
    "Mean_abs_SHAP",
    ascending=False
)

shap_importance_df.to_csv(
    "O4_SHAP_predictor_importance.csv",
    index=False
)

print("\nTop SHAP predictors:")
print(shap_importance_df.head(20))


# ------------------------------------------------------------
# 5. Top 20 SHAP predictors figure
# ------------------------------------------------------------

top_n = min(20, len(shap_importance_df))

plot_df = (
    shap_importance_df
    .head(top_n)
    .sort_values("Mean_abs_SHAP", ascending=True)
)

# Manuscript colours: CHM = blue, MSI/RGB-NIR = orange, SE = green
group_colors = {
    "CHM": "#6aaed6",
    "RGB-NIR": "#e7ad72",
    "SE": "#7bd88f"
}

plt.figure(figsize=(9, 8))

plt.barh(
    plot_df["Predictor"],
    plot_df["Mean_abs_SHAP"],
    xerr=plot_df["SD_abs_SHAP"],
    color=plot_df["Group"].map(group_colors),
    capsize=3
)

plt.xlabel("Mean absolute SHAP value")
plt.ylabel("Predictor")
plt.title(f"Top {top_n} SHAP predictors\n{best_model_name} | {best_feature_set}")

from matplotlib.patches import Patch

present_groups = plot_df["Group"].unique()

legend_elements = [
    Patch(facecolor=group_colors[g], label=g)
    for g in present_groups
    if g in group_colors
]

plt.legend(handles=legend_elements, title="Predictor group")

plt.tight_layout()
plt.savefig(
    "Figure_O4_top20_SHAP_predictors.png",
    dpi=300,
    bbox_inches="tight"
)

plt.close("all")


# ------------------------------------------------------------
# 6. Group-level SHAP importance
# ------------------------------------------------------------

group_shap_df = (
    shap_importance_df
    .groupby("Group", as_index=False)
    .agg(
        Total_Mean_abs_SHAP=("Mean_abs_SHAP", "sum"),
        Mean_Mean_abs_SHAP=("Mean_abs_SHAP", "mean"),
        N_Predictors=("Predictor", "count")
    )
    .sort_values("Total_Mean_abs_SHAP", ascending=False)
)

group_shap_df.to_csv(
    "O4_SHAP_group_importance.csv",
    index=False
)

print("\nGroup-level SHAP importance:")
print(group_shap_df)


plt.figure(figsize=(6, 5))

plt.bar(
    group_shap_df["Group"],
    group_shap_df["Total_Mean_abs_SHAP"],
    color=group_shap_df["Group"].map(group_colors)
)

plt.xlabel("Predictor group")
plt.ylabel("Total mean absolute SHAP value")
plt.title(f"Group-level SHAP importance\n{best_model_name} | {best_feature_set}")

plt.tight_layout()
plt.savefig(
    "Figure_O4_group_level_SHAP_importance.png",
    dpi=300,
    bbox_inches="tight"
)

plt.close("all")


# ------------------------------------------------------------
# 7. SHAP summary bar plot
# ------------------------------------------------------------

plt.figure()

shap.summary_plot(
    shap_matrix,
    X_selected,
    plot_type="bar",
    max_display=20,
    show=False
)

plt.title(f"SHAP summary bar plot\n{best_model_name} | {best_feature_set}")

plt.tight_layout()

plt.savefig(
    "Figure_O4_SHAP_summary_bar.png",
    dpi=300,
    bbox_inches="tight"
)

plt.close("all")


# ============================================================
# Objective 4B:
# Algorithm-specific SHAP importance for RF, GradientBoosting,
# and XGBoost using CHM + RGBNIR + SE models
# ============================================================

import joblib
from pathlib import Path

ALGO_SHAP_DIR = O4_DIR / "O4_algorithm_specific_SHAP"
ALGO_SHAP_DIR.mkdir(parents=True, exist_ok=True)

print("Algorithm-specific SHAP outputs will be saved to:", ALGO_SHAP_DIR)

pipeline_dir = OUT_DIR / "ImpPred_BestByAlgorithm"

pipeline_files = {
    "RF": pipeline_dir / "Final_pipeline_best_RF_CHM_plus_RGBNIR_plus_SE.joblib",
    "GradientBoosting": pipeline_dir / "Final_pipeline_best_GradientBoosting_CHM_plus_RGBNIR_plus_SE.joblib",
    "XGBoost": pipeline_dir / "Final_pipeline_best_XGBoost_CHM_plus_RGBNIR_plus_SE.joblib",
}

for model_name, pipeline_file in pipeline_files.items():
    if not pipeline_file.exists():
        raise FileNotFoundError(f"Missing fitted pipeline for {model_name}: {pipeline_file}")


def clean_model_name_for_plot(model_name):
    if model_name == "GradientBoosting":
        return "GB"
    return model_name


def get_group_name_for_shap(feature):
    if str(feature).startswith("CHM"):
        return "CHM"
    elif str(feature).startswith("SE2025") or str(feature).startswith("dSE"):
        return "SE"
    else:
        return "RGB-NIR"


all_algorithm_shap_tables = []
all_algorithm_shap_group_tables = []
all_algorithm_shap_model_info = []

for algorithm_name, pipeline_file in pipeline_files.items():

    print("\n=====================================")
    print("Algorithm-specific SHAP:", algorithm_name)
    print("Pipeline:", pipeline_file)
    print("=====================================")

    fitted_pipe = joblib.load(pipeline_file)

    selector = fitted_pipe.named_steps["selector"]
    fitted_model = fitted_pipe.named_steps["model"]

    selected_features_algorithm = selector.get_feature_names_out().tolist()

    X_selected_algorithm = selector.transform(X)
    X_selected_algorithm = pd.DataFrame(
        X_selected_algorithm,
        columns=selected_features_algorithm,
        index=X.index
    )

    explainer_algorithm = shap.TreeExplainer(fitted_model)
    shap_values_algorithm = explainer_algorithm.shap_values(X_selected_algorithm)

    if isinstance(shap_values_algorithm, list):
        shap_matrix_algorithm = shap_values_algorithm[0]
    else:
        shap_matrix_algorithm = np.array(shap_values_algorithm)

    if len(shap_matrix_algorithm.shape) == 3:
        shap_matrix_algorithm = shap_matrix_algorithm[:, :, 0]

    if shap_matrix_algorithm.shape[1] != len(selected_features_algorithm):
        raise ValueError(
            f"SHAP matrix has {shap_matrix_algorithm.shape[1]} columns but "
            f"{len(selected_features_algorithm)} selected features were found for {algorithm_name}."
        )

    model_plot_name = clean_model_name_for_plot(algorithm_name)

    shap_algorithm_df = pd.DataFrame({
        "Model": algorithm_name,
        "Model_plot": model_plot_name,
        "Feature_Set": "CHM+RGBNIR+SE",
        "Predictor": selected_features_algorithm,
        "Mean_abs_SHAP": np.abs(shap_matrix_algorithm).mean(axis=0),
        "SD_abs_SHAP": np.abs(shap_matrix_algorithm).std(axis=0)
    })

    shap_algorithm_df["Group"] = shap_algorithm_df["Predictor"].apply(get_group_name_for_shap)

    shap_algorithm_df = shap_algorithm_df.sort_values(
        "Mean_abs_SHAP",
        ascending=False
    )

    shap_algorithm_df.to_csv(
        ALGO_SHAP_DIR / f"SHAP_predictor_importance_{model_plot_name}_CHM_plus_RGBNIR_plus_SE.csv",
        index=False
    )

    shap_algorithm_df.head(20).to_csv(
        ALGO_SHAP_DIR / f"SHAP_top20_predictor_importance_{model_plot_name}_CHM_plus_RGBNIR_plus_SE.csv",
        index=False
    )

    group_algorithm_df = (
        shap_algorithm_df
        .groupby(["Model", "Model_plot", "Feature_Set", "Group"], as_index=False)
        .agg(
            Total_Mean_abs_SHAP=("Mean_abs_SHAP", "sum"),
            Mean_Mean_abs_SHAP=("Mean_abs_SHAP", "mean"),
            N_Predictors=("Predictor", "count")
        )
        .sort_values("Total_Mean_abs_SHAP", ascending=False)
    )

    group_algorithm_df.to_csv(
        ALGO_SHAP_DIR / f"SHAP_group_importance_{model_plot_name}_CHM_plus_RGBNIR_plus_SE.csv",
        index=False
    )

    all_algorithm_shap_tables.append(shap_algorithm_df)
    all_algorithm_shap_group_tables.append(group_algorithm_df)

    all_algorithm_shap_model_info.append({
        "Model": algorithm_name,
        "Model_plot": model_plot_name,
        "Feature_Set": "CHM+RGBNIR+SE",
        "Selected_Features": ", ".join(selected_features_algorithm),
        "N_Selected_Features": len(selected_features_algorithm),
        "Pipeline": str(pipeline_file)
    })


all_algorithm_shap_df = pd.concat(
    all_algorithm_shap_tables,
    ignore_index=True
)

all_algorithm_shap_df.to_csv(
    ALGO_SHAP_DIR / "All_algorithm_specific_SHAP_predictor_importance.csv",
    index=False
)

all_algorithm_group_shap_df = pd.concat(
    all_algorithm_shap_group_tables,
    ignore_index=True
)

all_algorithm_group_shap_df.to_csv(
    ALGO_SHAP_DIR / "All_algorithm_specific_SHAP_group_importance.csv",
    index=False
)

pd.DataFrame(all_algorithm_shap_model_info).to_csv(
    ALGO_SHAP_DIR / "All_algorithm_specific_SHAP_model_information.csv",
    index=False
)


# ------------------------------------------------------------
# Overall algorithm-averaged SHAP ranking
# Missing predictors are treated as zero in algorithms where
# they were not selected.
# ------------------------------------------------------------

model_cols = ["RF", "GB", "XGBoost"]

overall_shap_wide = (
    all_algorithm_shap_df
    .pivot_table(
        index=["Predictor", "Group"],
        columns="Model_plot",
        values="Mean_abs_SHAP",
        aggfunc="mean"
    )
    .reset_index()
)

for model_col in model_cols:
    if model_col not in overall_shap_wide.columns:
        overall_shap_wide[model_col] = 0.0

overall_shap_wide[model_cols] = overall_shap_wide[model_cols].fillna(0.0)

overall_shap_wide["Mean_abs_SHAP_overall"] = overall_shap_wide[model_cols].mean(axis=1)
overall_shap_wide["SD_abs_SHAP_across_models"] = overall_shap_wide[model_cols].std(axis=1)

overall_shap_df = overall_shap_wide.sort_values(
    "Mean_abs_SHAP_overall",
    ascending=False
)

overall_shap_df.to_csv(
    ALGO_SHAP_DIR / "Overall_mean_SHAP_importance_across_algorithms.csv",
    index=False
)

overall_shap_df.head(20).to_csv(
    ALGO_SHAP_DIR / "Table_07a_SHAP_top20_predictor_importance_across_algorithms.csv",
    index=False
)

overall_group_shap_df = (
    overall_shap_df
    .groupby("Group", as_index=False)
    .agg(
        Total_Mean_abs_SHAP_overall=("Mean_abs_SHAP_overall", "sum"),
        Mean_Mean_abs_SHAP_overall=("Mean_abs_SHAP_overall", "mean"),
        N_Predictors=("Predictor", "count")
    )
    .sort_values("Total_Mean_abs_SHAP_overall", ascending=False)
)

overall_group_shap_df.to_csv(
    ALGO_SHAP_DIR / "Overall_SHAP_group_importance_across_algorithms.csv",
    index=False
)

print("\nAlgorithm-specific SHAP analysis completed.")
print("Outputs saved to:", ALGO_SHAP_DIR)


# CELL

# ============================================================
# Objective 5:
# Complementarity analysis
# SHAP dependence and interaction analysis
# Do SE predictors complement CHM and RGB-NIR predictors?
# ============================================================

import shap
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt


# ------------------------------------------------------------
# 1. Use outputs from Objective 4
# Required existing objects:
# shap_matrix
# X_selected
# shap_importance_df
# selected_features
# final_model
# best_model_name
# best_feature_set
# ------------------------------------------------------------

print("Running Objective 5 for:")
print("Model:", best_model_name)
print("Feature set:", best_feature_set)


# ------------------------------------------------------------
# 2. Identify top predictors by group
# ------------------------------------------------------------

top_se_features = (
    shap_importance_df[shap_importance_df["Group"] == "SE"]
    .sort_values("Mean_abs_SHAP", ascending=False)
    .head(5)["Predictor"]
    .tolist()
)

top_chm_features = (
    shap_importance_df[shap_importance_df["Group"] == "CHM"]
    .sort_values("Mean_abs_SHAP", ascending=False)
    .head(5)["Predictor"]
    .tolist()
)

top_rgbnir_features = (
    shap_importance_df[shap_importance_df["Group"] == "RGB-NIR"]
    .sort_values("Mean_abs_SHAP", ascending=False)
    .head(5)["Predictor"]
    .tolist()
)

print("Top SE predictors:", top_se_features)
print("Top CHM predictors:", top_chm_features)
print("Top RGB-NIR predictors:", top_rgbnir_features)


# ------------------------------------------------------------
# 3. SHAP dependence plots for top SE predictors
# ------------------------------------------------------------

for se_feature in top_se_features:

    plt.figure(figsize=(6.5, 5))

    shap.dependence_plot(
        se_feature,
        shap_matrix,
        X_selected,
        interaction_index="auto",
        show=False
    )

    plt.title(f"SHAP dependence plot\n{se_feature}")
    plt.tight_layout()

    plt.savefig(
        f"Figure_O5_SHAP_dependence_{se_feature}.png",
        dpi=300,
        bbox_inches="tight"
    )

    plt.close("all")


# ------------------------------------------------------------
# 4. SE × CHM dependence plots
# ------------------------------------------------------------

for se_feature in top_se_features[:3]:

    for chm_feature in top_chm_features[:3]:

        plt.figure(figsize=(6.5, 5))

        shap.dependence_plot(
            se_feature,
            shap_matrix,
            X_selected,
            interaction_index=chm_feature,
            show=False
        )

        plt.title(f"SHAP dependence: {se_feature} × {chm_feature}")
        plt.tight_layout()

        plt.savefig(
            f"Figure_O5_dependence_{se_feature}_x_{chm_feature}.png",
            dpi=300,
            bbox_inches="tight"
        )

        plt.close("all")


# ------------------------------------------------------------
# 5. SE × RGB-NIR dependence plots
# ------------------------------------------------------------

for se_feature in top_se_features[:3]:

    for rgb_feature in top_rgbnir_features[:3]:

        plt.figure(figsize=(6.5, 5))

        shap.dependence_plot(
            se_feature,
            shap_matrix,
            X_selected,
            interaction_index=rgb_feature,
            show=False
        )

        plt.title(f"SHAP dependence: {se_feature} × {rgb_feature}")
        plt.tight_layout()

        plt.savefig(
            f"Figure_O5_dependence_{se_feature}_x_{rgb_feature}.png",
            dpi=300,
            bbox_inches="tight"
        )

        plt.close("all")


# ------------------------------------------------------------
# 6. Approximate complementarity using SHAP correlation
# ------------------------------------------------------------

complementarity_rows = []

for se_feature in top_se_features:

    se_idx = list(X_selected.columns).index(se_feature)
    se_shap = shap_matrix[:, se_idx]

    for other_feature in top_chm_features + top_rgbnir_features:

        other_idx = list(X_selected.columns).index(other_feature)
        other_shap = shap_matrix[:, other_idx]

        corr = np.corrcoef(se_shap, other_shap)[0, 1]

        complementarity_rows.append({
            "SE_Feature": se_feature,
            "Other_Feature": other_feature,
            "Other_Group": (
                "CHM" if other_feature in top_chm_features else "RGB-NIR"
            ),
            "SHAP_Correlation": corr,
            "Abs_SHAP_Correlation": abs(corr)
        })


complementarity_df = pd.DataFrame(complementarity_rows)

complementarity_df = complementarity_df.sort_values(
    "Abs_SHAP_Correlation",
    ascending=False
)

complementarity_df.to_csv(
    "O5_SHAP_complementarity_correlation.csv",
    index=False
)

print("\nSHAP complementarity correlation table:")
print(complementarity_df)


# ------------------------------------------------------------
# 7. Plot SHAP complementarity correlation
# ------------------------------------------------------------

plot_df = complementarity_df.copy()
plot_df["Pair"] = plot_df["SE_Feature"] + " × " + plot_df["Other_Feature"]

plot_df = plot_df.sort_values("Abs_SHAP_Correlation", ascending=True)

plt.figure(figsize=(10, 8))

plt.barh(
    plot_df["Pair"],
    plot_df["Abs_SHAP_Correlation"]
)

plt.xlabel("Absolute correlation between SHAP values")
plt.ylabel("Feature pair")
plt.title("Complementarity between SE and CHM/RGB-NIR predictors")

plt.tight_layout()

plt.savefig(
    "Figure_O5_SHAP_complementarity_correlation.png",
    dpi=300,
    bbox_inches="tight"
)

plt.close("all")


# ------------------------------------------------------------
# 8. Optional: SHAP interaction values for tree models
# ------------------------------------------------------------

try:
    explainer_tree = shap.TreeExplainer(final_model)
    shap_interaction_values = explainer_tree.shap_interaction_values(X_selected)

    if isinstance(shap_interaction_values, list):
        shap_interaction_values = shap_interaction_values[0]

    interaction_rows = []

    feature_list = list(X_selected.columns)

    for se_feature in top_se_features:
        se_idx = feature_list.index(se_feature)

        for other_feature in top_chm_features + top_rgbnir_features:
            other_idx = feature_list.index(other_feature)

            mean_abs_interaction = np.abs(
                shap_interaction_values[:, se_idx, other_idx]
            ).mean()

            interaction_rows.append({
                "SE_Feature": se_feature,
                "Other_Feature": other_feature,
                "Other_Group": (
                    "CHM" if other_feature in top_chm_features else "RGB-NIR"
                ),
                "Mean_abs_SHAP_Interaction": mean_abs_interaction
            })

    interaction_df = pd.DataFrame(interaction_rows)

    interaction_df = interaction_df.sort_values(
        "Mean_abs_SHAP_Interaction",
        ascending=False
    )

    interaction_df.to_csv(
        "O5_SHAP_interaction_values.csv",
        index=False
    )

    print("\nSHAP interaction table:")
    print(interaction_df)


    # --------------------------------------------------------
    # 9. SHAP interaction figure
    # --------------------------------------------------------

    plot_df = interaction_df.copy()
    plot_df["Pair"] = plot_df["SE_Feature"] + " × " + plot_df["Other_Feature"]

    plot_df = plot_df.sort_values(
        "Mean_abs_SHAP_Interaction",
        ascending=True
    )

    plt.figure(figsize=(10, 8))

    plt.barh(
        plot_df["Pair"],
        plot_df["Mean_abs_SHAP_Interaction"]
    )

    plt.xlabel("Mean absolute SHAP interaction value")
    plt.ylabel("Feature pair")
    plt.title("SHAP interaction strength: SE with CHM/MSI")

    plt.tight_layout()

    plt.savefig(
        "Figure_O5_SHAP_interaction_strength.png",
        dpi=300,
        bbox_inches="tight"
    )

    plt.close("all")

except Exception as e:
    print("SHAP interaction values could not be calculated.")
    print("Reason:", e)

# CELL

# ============================================================
# Objective 7:
# Plot-specific SE-benefit analysis
# Under which plot-level forest structural or spectral variables
# do SE predictors provide the greatest improvement?
# ============================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

from pathlib import Path
from sklearn.base import clone
from sklearn.pipeline import Pipeline
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
from matplotlib.lines import Line2D


# ------------------------------------------------------------
# 0. Output folder
# ------------------------------------------------------------

try:
    OUT_DIR
except NameError:
    OUT_DIR = Path(r"C:\Users\results\O1_nested_tuning_outputs")
    OUT_DIR.mkdir(parents=True, exist_ok=True)

O7_DIR = OUT_DIR / "O7_plot_specific_SE_benefit"
O7_DIR.mkdir(parents=True, exist_ok=True)


def display_feature_set_name(name):
    """Use manuscript labels in figures only; keep internal names unchanged."""
    return str(name).replace("RGBNIR", "MSI")


def display_model_name(name):
    """Use compact manuscript model labels in figures only."""
    return str(name).replace("GradientBoosting", "GB")


def safe_name(name):
    """Make a string safe for filenames."""
    return (
        str(name)
        .replace("/", "_")
        .replace("\\", "_")
        .replace(" ", "_")
        .replace("+", "_plus_")
    )


def pretty_name(x):
    """Make variable names cleaner for manuscript figures."""
    return str(x).replace("RGBNIR", "MSI").replace("_", " ")


def make_bin_labels(n_bins):
    """Return readable bin labels for qcut output."""
    if n_bins == 4:
        return ["Low", "Medium-low", "Medium-high", "High"]
    if n_bins == 3:
        return ["Low", "Medium", "High"]
    if n_bins == 2:
        return ["Low", "High"]
    return [f"Bin {i + 1}" for i in range(n_bins)]


# ------------------------------------------------------------
# 1. Choose baseline and SE-enhanced feature sets
# ------------------------------------------------------------

baseline_feature_set = "CHM+RGBNIR"
se_feature_set = "CHM+RGBNIR+SE"

if baseline_feature_set not in feature_sets:
    raise ValueError(f"{baseline_feature_set} not found in feature_sets.")

if se_feature_set not in feature_sets:
    raise ValueError(f"{se_feature_set} not found in feature_sets.")


# ------------------------------------------------------------
# 2. Select models
# ------------------------------------------------------------

# Use all models from Objective 1.
plot_models = {
    model_name: model_info["model"]
    for model_name, model_info in models_and_params.items()
}

# Option: only use best overall model.
# plot_models = {
#     best_model_name: models_and_params[best_model_name]["model"]
# }


# ------------------------------------------------------------
# 3. Fixed k-per-group for this analysis
# ------------------------------------------------------------

# Use tuned k_per_group if available; otherwise use a reasonable default.
if "best_final_params" in globals() and "selector__k_per_group" in best_final_params:
    plot_k_per_group = best_final_params["selector__k_per_group"]
else:
    plot_k_per_group = k_grid[-1]

print("Using k_per_group:", plot_k_per_group)


# ------------------------------------------------------------
# 4. Create or reuse repeated CV splitter
# ------------------------------------------------------------

# Use the same outer CV logic from Objective 1 if available.
try:
    plot_cv = outer_cv
except NameError:
    plot_cv = RepeatedStratifiedKFoldReg(
        n_splits=5,
        n_repeats=10,
        n_bins=5,
        random_state=42
    )


# ------------------------------------------------------------
# 5. Create repeated-CV predictions for baseline vs SE model
# ------------------------------------------------------------

plot_prediction_rows = []

for model_name, model in plot_models.items():

    print("\nModel:", model_name)

    for feature_name in [baseline_feature_set, se_feature_set]:

        print("Feature set:", feature_name)

        groups = feature_sets[feature_name]

        selected_input_cols = []

        for cols in groups.values():
            selected_input_cols.extend(cols)

        selected_input_cols = list(dict.fromkeys(selected_input_cols))

        X_sub = X[selected_input_cols].copy()

        selector = GroupWiseSelectKBest(
            groups=groups,
            k_per_group=plot_k_per_group
        )

        pipe = Pipeline([
            ("selector", selector),
            ("model", clone(model))
        ])

        for fold_id, (train_idx, test_idx) in enumerate(
            plot_cv.split(X_sub, y),
            start=1
        ):

            pipe_fold = clone(pipe)

            X_train = X_sub.iloc[train_idx].copy()
            X_test = X_sub.iloc[test_idx].copy()

            y_train = y.iloc[train_idx].copy()
            y_test = y.iloc[test_idx].copy()

            pipe_fold.fit(X_train, y_train)

            y_pred = pipe_fold.predict(X_test)

            for idx, obs, pred in zip(X_test.index, y_test, y_pred):

                plot_prediction_rows.append({
                    "Index": idx,
                    "CV_Run": fold_id,
                    "Model": model_name,
                    "Feature_Set": feature_name,
                    "Observed": obs,
                    "Predicted": pred,
                    "Abs_Error": abs(obs - pred),
                    "Squared_Error": (obs - pred) ** 2
                })


plot_pred_df = pd.DataFrame(plot_prediction_rows)

plot_pred_df.to_csv(
    O7_DIR / "O7_plot_specific_predictions.csv",
    index=False
)


# ------------------------------------------------------------
# 6. Convert long predictions to baseline vs SE comparison
# ------------------------------------------------------------

plot_compare_df = plot_pred_df.pivot_table(
    index=["Index", "CV_Run", "Model", "Observed"],
    columns="Feature_Set",
    values=["Predicted", "Abs_Error", "Squared_Error"],
    aggfunc="first"
)

plot_compare_df.columns = [
    "_".join(col).strip()
    for col in plot_compare_df.columns.values
]

plot_compare_df = plot_compare_df.reset_index()

required_cols = [
    f"Abs_Error_{baseline_feature_set}",
    f"Abs_Error_{se_feature_set}",
    f"Squared_Error_{baseline_feature_set}",
    f"Squared_Error_{se_feature_set}"
]

missing_cols = [
    col for col in required_cols
    if col not in plot_compare_df.columns
]

if missing_cols:
    raise ValueError(f"Missing columns after pivot: {missing_cols}")

plot_compare_df["SE_Abs_Error_Improvement"] = (
    plot_compare_df[f"Abs_Error_{baseline_feature_set}"]
    - plot_compare_df[f"Abs_Error_{se_feature_set}"]
)

plot_compare_df["SE_Squared_Error_Improvement"] = (
    plot_compare_df[f"Squared_Error_{baseline_feature_set}"]
    - plot_compare_df[f"Squared_Error_{se_feature_set}"]
)

plot_compare_df["SE_Better"] = (
    plot_compare_df["SE_Abs_Error_Improvement"] > 0
)

plot_compare_df.to_csv(
    O7_DIR / "O7_SE_error_improvement_per_plot_CV.csv",
    index=False
)


# ------------------------------------------------------------
# 7. Average repeated-CV values to one record per plot and model
# ------------------------------------------------------------

# This avoids treating repeated predictions for the same plot
# as independent observations in plot-level analyses.
plot_summary_df = (
    plot_compare_df
    .groupby(["Index", "Model"], as_index=False)
    .agg(
        Observed=("Observed", "first"),
        Mean_SE_Abs_Error_Improvement=("SE_Abs_Error_Improvement", "mean"),
        Mean_SE_Squared_Error_Improvement=("SE_Squared_Error_Improvement", "mean"),
        Percent_CV_Runs_SE_Better=("SE_Better", "mean"),
        N_CV_Runs=("SE_Better", "count")
    )
)

plot_summary_df["SE_Abs_Error_Improvement"] = (
    plot_summary_df["Mean_SE_Abs_Error_Improvement"]
)

plot_summary_df["SE_Squared_Error_Improvement"] = (
    plot_summary_df["Mean_SE_Squared_Error_Improvement"]
)

plot_summary_df["SE_Better"] = (
    plot_summary_df["Percent_CV_Runs_SE_Better"] > 0.5
)

plot_summary_df.to_csv(
    O7_DIR / "O7_SE_error_improvement_per_plot_mean.csv",
    index=False
)

# Use the plot-level aggregated table for all analyses below.
plot_compare_df = plot_summary_df.copy()


# ------------------------------------------------------------
# 8. Add plot-level variables
# ------------------------------------------------------------

# Use all CHM and MSI/RGB-NIR predictors as potential plot-level variables.
# These describe each plot structurally or spectrally.
plot_vars = list(dict.fromkeys(CHM_cols + RGBNIR_cols))

# Add observed target as a plot-level variable.
plot_compare_df["Dead_F_observed"] = plot_compare_df["Observed"]
plot_vars.append("Dead_F_observed")

plot_vars = list(dict.fromkeys(plot_vars))

# Add plot-level variable values from the original X matrix.
for col in plot_vars:
    if col in X.columns:
        plot_compare_df[col] = plot_compare_df["Index"].map(X[col])
    elif col == "Dead_F_observed":
        pass


# ------------------------------------------------------------
# 9. Correlation between plot-level variables and SE improvement
# ------------------------------------------------------------

plot_corr_rows = []

for model_name in plot_compare_df["Model"].unique():

    model_df = plot_compare_df[
        plot_compare_df["Model"] == model_name
    ].copy()

    for var in plot_vars:

        if var not in model_df.columns:
            continue

        temp = model_df[[var, "SE_Abs_Error_Improvement"]].dropna()

        if temp[var].nunique() < 3:
            continue

        corr = temp[var].corr(
            temp["SE_Abs_Error_Improvement"],
            method="spearman"
        )

        if pd.isna(corr):
            continue

        plot_corr_rows.append({
            "Model": model_name,
            "Plot_Variable": var,
            "Spearman_Correlation_with_SE_Improvement": corr,
            "Abs_Correlation": abs(corr)
        })


plot_corr_df = pd.DataFrame(plot_corr_rows)

if plot_corr_df.empty:
    raise ValueError("No valid plot-level correlations were calculated.")

plot_corr_df = plot_corr_df.sort_values(
    "Abs_Correlation",
    ascending=False
)

plot_corr_df.to_csv(
    O7_DIR / "O7_plot_variable_correlations_with_SE_improvement.csv",
    index=False
)

print("\nTop plot-level variables associated with SE improvement:")
print(plot_corr_df.head(20))


# ------------------------------------------------------------
# 10. Bin-based plot-variable analysis
# ------------------------------------------------------------

top_plot_vars = (
    plot_corr_df
    .head(6)["Plot_Variable"]
    .tolist()
)

bin_summary_rows = []

for model_name in plot_compare_df["Model"].unique():

    model_df = plot_compare_df[
        plot_compare_df["Model"] == model_name
    ].copy()

    for var in top_plot_vars:

        if var not in model_df.columns:
            continue

        temp = model_df[
            [
                var,
                "SE_Abs_Error_Improvement",
                "SE_Better",
                "Percent_CV_Runs_SE_Better"
            ]
        ].dropna().copy()

        if temp[var].nunique() < 5:
            continue

        try:
            preliminary_bins = pd.qcut(
                temp[var],
                q=4,
                duplicates="drop"
            )
        except ValueError:
            continue

        n_bins = len(preliminary_bins.cat.categories)

        if n_bins < 2:
            continue

        bin_labels = make_bin_labels(n_bins)

        try:
            temp["Plot_Variable_Bin"] = pd.qcut(
                temp[var],
                q=4,
                labels=bin_labels,
                duplicates="drop"
            )
        except ValueError:
            continue

        summary = (
            temp
            .groupby("Plot_Variable_Bin", observed=False)
            .agg(
                Mean_SE_Abs_Error_Improvement=("SE_Abs_Error_Improvement", "mean"),
                SD_SE_Abs_Error_Improvement=("SE_Abs_Error_Improvement", "std"),
                Percent_Plots_SE_Better_Most_CV_Runs=("SE_Better", "mean"),
                Mean_Percent_CV_Runs_SE_Better=("Percent_CV_Runs_SE_Better", "mean"),
                N_Plots=("SE_Better", "count")
            )
            .reset_index()
        )

        summary["Percent_Plots_SE_Better_Most_CV_Runs"] = (
            summary["Percent_Plots_SE_Better_Most_CV_Runs"] * 100
        )

        summary["Mean_Percent_CV_Runs_SE_Better"] = (
            summary["Mean_Percent_CV_Runs_SE_Better"] * 100
        )

        summary["Model"] = model_name
        summary["Plot_Variable"] = var

        bin_summary_rows.append(summary)


if len(bin_summary_rows) > 0:
    bin_summary_df = pd.concat(bin_summary_rows, ignore_index=True)
else:
    bin_summary_df = pd.DataFrame()

bin_summary_df.to_csv(
    O7_DIR / "O7_plot_variable_bin_summary.csv",
    index=False
)

print("\nPlot-variable bin summary:")
print(bin_summary_df.head())


# ------------------------------------------------------------
# 11. Figure: SE improvement across plot-variable bins
# ------------------------------------------------------------

if not bin_summary_df.empty:

    for model_name in bin_summary_df["Model"].unique():

        model_bin_df = bin_summary_df[
            bin_summary_df["Model"] == model_name
        ]

        for var in model_bin_df["Plot_Variable"].unique():

            plot_df = model_bin_df[
                model_bin_df["Plot_Variable"] == var
            ].copy()

            plt.figure(figsize=(6.5, 5))

            plt.bar(
                plot_df["Plot_Variable_Bin"].astype(str),
                plot_df["Mean_SE_Abs_Error_Improvement"],
                yerr=plot_df["SD_SE_Abs_Error_Improvement"],
                capsize=4
            )

            plt.axhline(0, linestyle="--", linewidth=1)

            plt.xlabel(f"{pretty_name(var)} quartile")
            plt.ylabel("Mean absolute-error improvement from SE")
            plt.title(
                f"SE benefit across plot-level {pretty_name(var)}\n"
                f"{display_model_name(model_name)}"
            )

            plt.tight_layout()

            safe_model = safe_name(model_name)
            safe_var = safe_name(var)

            plt.savefig(
                O7_DIR / f"Figure_O7_SE_improvement_plot_bins_{safe_model}_{safe_var}.png",
                dpi=300,
                bbox_inches="tight"
            )

            plt.close("all")


# ------------------------------------------------------------
# 12. Figure: top plot-variable correlations
# ------------------------------------------------------------

top_corr_plot = plot_corr_df.head(20).copy()

top_corr_plot["Model_plot"] = top_corr_plot["Model"].apply(display_model_name)
top_corr_plot["Plot_Label"] = top_corr_plot["Plot_Variable"].apply(pretty_name)

top_corr_plot["Label"] = (
    top_corr_plot["Model_plot"] + " | " + top_corr_plot["Plot_Label"]
)

top_corr_plot = top_corr_plot.sort_values(
    "Abs_Correlation",
    ascending=True
)

plt.figure(figsize=(9, 8))

plt.barh(
    top_corr_plot["Label"],
    top_corr_plot["Spearman_Correlation_with_SE_Improvement"]
)

plt.axvline(0, linestyle="--", linewidth=1)

plt.xlabel("Spearman correlation with SE error improvement")
plt.ylabel("Model and plot-level variable")
plt.title("Plot-level variables associated with stronger SE benefit")

plt.tight_layout()

plt.savefig(
    O7_DIR / "Figure_O7_top_plot_variable_correlations_SE_improvement.png",
    dpi=300,
    bbox_inches="tight"
)

plt.close("all")


# ------------------------------------------------------------
# 13. Best plot-level SE benefit table
# ------------------------------------------------------------

if not bin_summary_df.empty:

    best_plot_variable_table = (
        bin_summary_df
        .sort_values("Mean_SE_Abs_Error_Improvement", ascending=False)
    )

    best_plot_variable_table.to_csv(
        O7_DIR / "O7_best_plot_variables_for_SE_benefit.csv",
        index=False
    )

    print("\nBest plot-level variables for SE benefit:")
    print(best_plot_variable_table.head(20))

else:
    best_plot_variable_table = pd.DataFrame()
    print("\nNo bin-based plot-variable table was created.")


print("\nObjective 7 outputs saved to:")
print(O7_DIR)


# ============================================================
# Objective 7 alternative manuscript figure:
# Split dot plot of negative and positive plot-level SE-benefit associations
# ============================================================

# -----------------------------
# 14. File paths
# -----------------------------
base_dir = O7_DIR

corr_file = base_dir / "O7_plot_variable_correlations_with_SE_improvement.csv"

outdir = base_dir / "manuscript_figures"
outdir.mkdir(parents=True, exist_ok=True)

# -----------------------------
# 15. Load data
# -----------------------------
corr_df = pd.read_csv(corr_file)
corr_df.columns = corr_df.columns.str.strip()

# -----------------------------
# 16. Clean model names and labels
# -----------------------------
corr_df["Model"] = corr_df["Model"].replace({
    "GradientBoosting": "GB",
    "Gradient_Boosting": "GB"
})

model_order = ["RF", "GB", "XGBoost"]

model_colors = {
    "RF": "#1f77b4",
    "GB": "#ff7f0e",
    "XGBoost": "#2ca02c"
}

model_markers = {
    "RF": "o",
    "GB": "s",
    "XGBoost": "^"
}

corr_col = "Spearman_Correlation_with_SE_Improvement"

corr_df["Plot_Label"] = corr_df["Plot_Variable"].apply(pretty_name)
corr_df["Full_Label"] = corr_df["Model"] + " | " + corr_df["Plot_Label"]

# -----------------------------
# 17. Select strongest positive and negative associations
# -----------------------------
n_show = 10

neg_df = (
    corr_df[corr_df[corr_col] < 0]
    .sort_values(corr_col, ascending=True)
    .head(n_show)
    .copy()
)

pos_df = (
    corr_df[corr_df[corr_col] > 0]
    .sort_values(corr_col, ascending=False)
    .head(n_show)
    .copy()
)

# Sort for display.
neg_df = neg_df.sort_values(corr_col, ascending=False).reset_index(drop=True)
pos_df = pos_df.sort_values(corr_col, ascending=True).reset_index(drop=True)

# -----------------------------
# 18. Plot
# -----------------------------
fig, axes = plt.subplots(
    nrows=1,
    ncols=2,
    figsize=(12.5, 5.8),
    sharex=False
)


def plot_correlation_side(ax, side_df, panel_label, side_title, value_side):
    """Plot one side of the split dot plot."""

    y = np.arange(len(side_df))

    if side_df.empty:
        ax.axvline(0, linestyle="--", linewidth=1, color="black")
        ax.set_yticks([])
        ax.set_xlabel("Spearman correlation", fontsize=9)
        ax.text(
            0.5,
            0.5,
            "No associations",
            transform=ax.transAxes,
            ha="center",
            va="center",
            fontsize=9
        )
        ax.set_xlim(-0.1, 0.1)

    else:
        for i, row in side_df.iterrows():

            model_key = row["Model"]

            ax.scatter(
                row[corr_col],
                i,
                color=model_colors.get(model_key, "black"),
                marker=model_markers.get(model_key, "o"),
                s=55
            )

            if value_side == "left":
                text_x = row[corr_col] - 0.006
                ha = "right"
            else:
                text_x = row[corr_col] + 0.006
                ha = "left"

            ax.text(
                text_x,
                i,
                f"{row[corr_col]:.2f}",
                ha=ha,
                va="center",
                fontsize=8
            )

        ax.axvline(0, linestyle="--", linewidth=1, color="black")
        ax.set_yticks(y)
        ax.set_yticklabels(side_df["Full_Label"], fontsize=8.5)
        ax.set_xlabel("Spearman correlation", fontsize=9)
        ax.grid(axis="x", alpha=0.25)

        if value_side == "left":
            ax.set_xlim(min(side_df[corr_col].min() - 0.04, -0.22), 0.02)
        else:
            ax.set_xlim(-0.02, max(side_df[corr_col].max() + 0.04, 0.14))

    ax.text(
        0.01,
        1.03,
        panel_label,
        transform=ax.transAxes,
        fontsize=11,
        fontweight="bold",
        ha="left",
        va="bottom"
    )

    ax.text(
        1.08,
        0.5,
        side_title,
        transform=ax.transAxes,
        rotation=-90,
        fontsize=9,
        ha="center",
        va="center"
    )


plot_correlation_side(
    axes[0],
    neg_df,
    "(a)",
    "Plot-level variables associated with weaker SE benefit",
    "left"
)

plot_correlation_side(
    axes[1],
    pos_df,
    "(b)",
    "Plot-level variables associated with stronger SE benefit",
    "right"
)


# -----------------------------
# 19. Shared legend
# -----------------------------
legend_handles = [
    Line2D(
        [0], [0],
        marker=model_markers[m],
        color="w",
        markerfacecolor=model_colors[m],
        markeredgecolor=model_colors[m],
        markersize=7,
        label=m
    )
    for m in model_order
    if m in model_markers
]

fig.legend(
    handles=legend_handles,
    loc="upper center",
    ncol=3,
    frameon=False,
    bbox_to_anchor=(0.5, 1.02),
    fontsize=9
)

# Footnote
fig.text(
    0.5,
    0.01,
    (
        "Positive correlations indicate greater SE benefit at higher plot-level variable values; "
        "negative correlations indicate weaker SE benefit."
    ),
    ha="center",
    fontsize=8
)

# Leave some extra right margin so vertical titles are not cut off.
plt.tight_layout(rect=(0, 0.05, 0.95, 0.96))

# -----------------------------
# 20. Save
# -----------------------------
fig.savefig(
    outdir / "Figure_O7_split_dotplot_plot_variable_associations.png",
    dpi=300,
    bbox_inches="tight"
)
fig.savefig(
    outdir / "Figure_O7_split_dotplot_plot_variable_associations.svg",
    bbox_inches="tight"
)
fig.savefig(
    outdir / "Figure_O7_split_dotplot_plot_variable_associations.pdf",
    bbox_inches="tight"
)

plt.close("all")

# Backward-compatible aliases for any later code that still expects
# the previous Objective 7 object names.
condition_corr_df = plot_corr_df
best_condition_table = best_plot_variable_table

# CELL

# Optional exact-format manuscript figure section 0
try:
    
    import pandas as pd
    import matplotlib.pyplot as plt
    from matplotlib.ticker import MultipleLocator, FormatStrFormatter
    from pathlib import Path
    
    # -----------------------------
    # Input files
    # -----------------------------
    interaction_file = O4_DIR / "O5_SHAP_interaction_values.csv"
    correlation_file = O4_DIR / "O5_SHAP_complementarity_correlation.csv"
    
    interaction = pd.read_csv(interaction_file)
    correlation = pd.read_csv(correlation_file)
    
    # -----------------------------
    # Build feature-pair labels
    # -----------------------------
    interaction["Feature pair"] = (
        interaction["SE_Feature"].astype(str) + " x " + interaction["Other_Feature"].astype(str)
    )
    
    correlation["Feature pair"] = (
        correlation["SE_Feature"].astype(str) + " x " + correlation["Other_Feature"].astype(str)
    )
    
    # -----------------------------
    # Sort and select top pairs
    # -----------------------------
    top_n = 40
    
    left_df = interaction.sort_values(
        "Mean_abs_SHAP_Interaction",
        ascending=False
    ).head(top_n)
    
    right_df = correlation.sort_values(
        "Abs_SHAP_Correlation",
        ascending=False
    ).head(top_n)
    
    # -----------------------------
    # A4 landscape settings
    # -----------------------------
    A4_LANDSCAPE = (11.69, 8.27)  # inches
    
    y_label_fontsize = 6.5
    axis_label_fontsize = 8
    title_fontsize = 9
    tick_fontsize = 7
    panel_label_fontsize = 11
    
    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 7,
        "axes.titlesize": title_fontsize,
        "axes.labelsize": axis_label_fontsize,
        "xtick.labelsize": tick_fontsize,
        "ytick.labelsize": y_label_fontsize,
        "axes.linewidth": 0.8,
    })
    
    bar_color = "#cd9778"
    
    fig, axes = plt.subplots(
        1, 2,
        figsize=A4_LANDSCAPE,
        dpi=300
    )
    
    # -----------------------------
    # Helper function
    # -----------------------------
    def make_panel(ax, df, value_col, title, xlabel, xlim, tick_step, tick_format):
        y = range(len(df))
    
        ax.barh(
            y,
            df[value_col],
            color=bar_color,
            edgecolor=bar_color,
            height=0.68
        )
    
        ax.set_yticks(y)
        ax.set_yticklabels(
            df["Feature pair"],
            fontsize=y_label_fontsize
        )
    
        ax.invert_yaxis()
    
        ax.set_title(title, fontsize=title_fontsize, pad=8)
        ax.set_xlabel(xlabel, fontsize=axis_label_fontsize)
        ax.set_ylabel("Feature pair", fontsize=axis_label_fontsize, labelpad=8)
    
        ax.set_xlim(xlim)
        ax.xaxis.set_major_locator(MultipleLocator(tick_step))
        ax.xaxis.set_major_formatter(FormatStrFormatter(tick_format))
    
        ax.xaxis.grid(True, color="#e6e6e6", linewidth=0.6)
        ax.set_axisbelow(True)
    
        ax.tick_params(axis="y", length=3, pad=2, labelsize=y_label_fontsize)
        ax.tick_params(axis="x", length=3, labelsize=tick_fontsize)
    
    # -----------------------------
    # Automatic x-axis limits
    # Extra space prevents longest bars from touching the border
    # -----------------------------
    left_xmax = left_df["Mean_abs_SHAP_Interaction"].max() * 1.10
    right_xmax = right_df["Abs_SHAP_Correlation"].max() * 1.08
    
    # -----------------------------
    # Panel A
    # -----------------------------
    make_panel(
        axes[0],
        left_df,
        "Mean_abs_SHAP_Interaction",
        "SHAP interaction strength: SE with CHM/MSI",
        "Mean absolute SHAP interaction value",
        xlim=(0, left_xmax),
        tick_step=0.004,
        tick_format="%.3f"
    )
    
    # -----------------------------
    # Panel B
    # -----------------------------
    make_panel(
        axes[1],
        right_df,
        "Abs_SHAP_Correlation",
        "Complementarity between SE and CHM/MSI predictors",
        "Absolute correlation between SHAP values",
        xlim=(0, right_xmax),
        tick_step=0.1,
        tick_format="%.1f"
    )
    
    # -----------------------------
    # Panel labels
    # -----------------------------
    axes[0].text(
        -0.08, 1.025, "(a)",
        transform=axes[0].transAxes,
        fontsize=panel_label_fontsize,
        fontweight="bold"
    )
    
    axes[1].text(
        -0.08, 1.025, "(b)",
        transform=axes[1].transAxes,
        fontsize=panel_label_fontsize,
        fontweight="bold"
    )
    
    # -----------------------------
    # Layout for A4 landscape
    # Increased left margin helps long y-axis labels fit
    # Reduced right margin gives bars enough visual space
    # -----------------------------
    plt.subplots_adjust(
        left=0.18,
        right=0.97,
        top=0.935,
        bottom=0.085,
        wspace=0.65
    )
    
    # -----------------------------
    # Save figure
    # Do not use bbox_inches="tight" because it changes exact A4 canvas size
    # -----------------------------
    plt.savefig(
        O4_DIR / "SHAP_interaction_complementarity_A4_landscape.png",
        dpi=300
    )
    
    plt.savefig(
        O4_DIR / "SHAP_interaction_complementarity_A4_landscape.pdf"
    )
    
    plt.close("all")
    
except Exception as exc:
    print('Warning: skipped optional exact-format figure section 0:', exc)
    traceback.print_exc(limit=1)


# Optional exact-format manuscript figure section 1
try:

    import pandas as pd
    import matplotlib.pyplot as plt
    from matplotlib.patches import Patch
    from matplotlib.ticker import MultipleLocator, FormatStrFormatter
    from pathlib import Path

    # ============================================================
    # Algorithm-specific SHAP source files
    # ============================================================
    DATA_DIR = OUT_DIR / "O4_SHAP_and_complementarity" / "O4_algorithm_specific_SHAP"

    print("Using algorithm-specific SHAP data folder:")
    print(DATA_DIR)

    all_file = DATA_DIR / "All_algorithm_specific_SHAP_predictor_importance.csv"
    overall_file = DATA_DIR / "Overall_mean_SHAP_importance_across_algorithms.csv"

    if not all_file.exists():
        raise FileNotFoundError(f"File not found:\n{all_file}")
    if not overall_file.exists():
        raise FileNotFoundError(f"File not found:\n{overall_file}")

    all_df = pd.read_csv(all_file)
    overall_df = pd.read_csv(overall_file)

    if "Mean_abs_SHAP" not in all_df.columns:
        raise ValueError(f"Expected column 'Mean_abs_SHAP' in {all_file}")
    if "Mean_abs_SHAP_overall" not in overall_df.columns:
        raise ValueError(f"Expected column 'Mean_abs_SHAP_overall' in {overall_file}")

    def clean_group(group):
        if group in ["RGB-NIR", "RGBNIR", "MSI", "RGB_NIR"]:
            return "MSI"
        return group

    all_df["Plot_Group"] = all_df["Group"].apply(clean_group)
    overall_df["Plot_Group"] = overall_df["Group"].apply(clean_group)

    if "Model_plot" not in all_df.columns:
        all_df["Model_plot"] = all_df["Model"].replace({"GradientBoosting": "GB"})

    group_colors = {
        "CHM": "#6aaed6",
        "MSI": "#e7ad72",
        "SE": "#7bd88f"
    }

    overall_top = (
        overall_df
        .sort_values("Mean_abs_SHAP_overall", ascending=False)
        .head(20)
    )

    rf_top = (
        all_df[all_df["Model_plot"] == "RF"]
        .sort_values("Mean_abs_SHAP", ascending=False)
        .head(10)
    )

    gb_top = (
        all_df[all_df["Model_plot"] == "GB"]
        .sort_values("Mean_abs_SHAP", ascending=False)
        .head(10)
    )

    xgb_top = (
        all_df[all_df["Model_plot"] == "XGBoost"]
        .sort_values("Mean_abs_SHAP", ascending=False)
        .head(10)
    )

    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 8,
        "axes.titlesize": 10,
        "axes.labelsize": 8,
        "xtick.labelsize": 8,
        "ytick.labelsize": 8,
        "axes.linewidth": 0.9,
    })

    fig, axes = plt.subplots(
        2, 2,
        figsize=(11.69, 8.27),
        dpi=300
    )

    axes = axes.flatten()

    def plot_barh(ax, df, value_col, title, panel_label, tick_step=0.02):
        df_plot = df.sort_values(value_col, ascending=True).copy()
        colors = df_plot["Plot_Group"].map(group_colors)
        xmax = float(df[value_col].max()) * 1.10 if len(df) else 0.1
        xmax = max(xmax, 0.01)

        ax.barh(
            df_plot["Predictor"],
            df_plot[value_col],
            color=colors,
            edgecolor=colors,
            height=0.70
        )

        ax.set_title(title, fontweight="bold", pad=6)
        ax.set_xlabel("Mean |SHAP value|")
        ax.set_xlim(0, xmax)
        ax.xaxis.set_major_locator(MultipleLocator(tick_step))
        ax.xaxis.set_major_formatter(FormatStrFormatter("%.2f"))
        ax.grid(axis="x", linestyle="--", linewidth=0.6, alpha=0.35)
        ax.set_axisbelow(True)
        ax.tick_params(axis="y", labelsize=8)
        ax.tick_params(axis="x", labelsize=8)
        ax.text(
            0.0, 1.04,
            panel_label,
            transform=ax.transAxes,
            fontsize=10,
            fontweight="bold",
            ha="left",
            va="bottom"
        )

    plot_barh(axes[0], overall_top, "Mean_abs_SHAP_overall", "Overall Top 20 Predictors", "(a)")
    plot_barh(axes[1], rf_top, "Mean_abs_SHAP", "RF | Top 10 Predictors", "(b)")
    plot_barh(axes[2], gb_top, "Mean_abs_SHAP", "GB | Top 10 Predictors", "(c)")
    plot_barh(axes[3], xgb_top, "Mean_abs_SHAP", "XGBoost | Top 10 Predictors", "(d)")

    legend_handles = [
        Patch(facecolor=group_colors["CHM"], edgecolor="black", label="CHM"),
        Patch(facecolor=group_colors["MSI"], edgecolor="black", label="MSI"),
        Patch(facecolor=group_colors["SE"], edgecolor="black", label="SE"),
    ]

    fig.legend(
        handles=legend_handles,
        loc="lower center",
        ncol=3,
        frameon=False,
        bbox_to_anchor=(0.5, 0.035),
        fontsize=8
    )

    plt.subplots_adjust(
        left=0.13,
        right=0.985,
        top=0.94,
        bottom=0.12,
        wspace=0.30,
        hspace=0.26
    )

    out_png = DATA_DIR / "Top_predictors_A4_exact_style.png"
    out_pdf = DATA_DIR / "Top_predictors_A4_exact_style.pdf"

    plt.savefig(out_png, dpi=300)
    plt.savefig(out_pdf)

    plt.close("all")

    print("Saved PNG to:", out_png)
    print("Saved PDF to:", out_pdf)

except Exception as exc:
    print('Warning: skipped optional exact-format figure section 1:', exc)
    traceback.print_exc(limit=1)


# Optional exact-format manuscript figure section 2
try:
    
    import pandas as pd
    import matplotlib.pyplot as plt
    from pathlib import Path
    
    # ============================================================
    # Your data folder
    # ============================================================
    DATA_DIR = OUT_DIR / "O3_label_efficiency"
    
    summary_file = DATA_DIR / "O3_label_efficiency_summary.csv"
    
    if not summary_file.exists():
        raise FileNotFoundError(
            f"File not found:\n{summary_file}\n\n"
            "Please check that O3_label_efficiency_summary.csv is inside this folder."
        )
    
    df = pd.read_csv(summary_file)
    
    # ============================================================
    # Rename feature-set labels for the figure
    # ============================================================
    feature_label_map = {
        "CHM+RGBNIR": "CHM+MSI",
        "CHM+RGBNIR+SE": "CHM+MSI+SE"
    }
    
    df["Feature_Set_Plot"] = df["Feature_Set"].replace(feature_label_map)
    
    # ============================================================
    # Model order and display names
    # ============================================================
    model_order = ["RF", "GradientBoosting", "XGBoost"]
    
    model_title_map = {
        "RF": "RF",
        "GradientBoosting": "GB",
        "XGBoost": "XGBoost"
    }
    
    feature_order = ["CHM+MSI", "CHM+MSI+SE"]
    
    # ============================================================
    # Colors matching the shown figure
    # ============================================================
    colors = {
        "CHM+MSI": "#1f77b4",
        "CHM+MSI+SE": "#ff7f0e"
    }
    
    # ============================================================
    # Metrics and axis limits
    # ============================================================
    metrics = [
        {
            "mean_col": "R2_mean",
            "sd_col": "R2_sd",
            "ylabel": "R²",
            "ylim": (-0.55, 0.70),
            "yticks": [-0.4, -0.2, 0.0, 0.2, 0.4, 0.6]
        },
        {
            "mean_col": "RMSE_mean",
            "sd_col": "RMSE_sd",
            "ylabel": "RMSE",
            "ylim": (0.11, 0.24),
            "yticks": [0.12, 0.14, 0.16, 0.18, 0.20, 0.22, 0.24]
        },
        {
            "mean_col": "MAE_mean",
            "sd_col": "MAE_sd",
            "ylabel": "MAE",
            "ylim": (0.08, 0.16),
            "yticks": [0.08, 0.09, 0.10, 0.11, 0.12, 0.13, 0.14, 0.15, 0.16]
        }
    ]
    
    # ============================================================
    # Figure style
    # ============================================================
    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 8,
        "axes.titlesize": 9,
        "axes.labelsize": 8,
        "xtick.labelsize": 8,
        "ytick.labelsize": 8,
        "axes.linewidth": 0.9,
    })
    
    # A4 landscape
    fig, axes = plt.subplots(
        3, 3,
        figsize=(11.69, 8.27),
        dpi=300
    )
    
    panel_letters = [
        "(a)", "(b)", "(c)",
        "(d)", "(e)", "(f)",
        "(g)", "(h)", "(i)"
    ]
    
    # ============================================================
    # Plot panels
    # ============================================================
    panel_i = 0
    
    for row_i, metric in enumerate(metrics):
        for col_i, model in enumerate(model_order):
    
            ax = axes[row_i, col_i]
    
            model_df = df[df["Model"] == model].copy()
    
            for feature_set in feature_order:
                plot_df = (
                    model_df[model_df["Feature_Set_Plot"] == feature_set]
                    .sort_values("Train_Proportion")
                    .copy()
                )
    
                x = plot_df["Train_Proportion"].to_numpy() * 100
                y = plot_df[metric["mean_col"]].to_numpy()
                sd = plot_df[metric["sd_col"]].fillna(0).to_numpy()
    
                ax.plot(
                    x,
                    y,
                    marker="o",
                    markersize=4.5,
                    linewidth=1.8,
                    color=colors[feature_set],
                    label=feature_set
                )
    
                ax.fill_between(
                    x,
                    y - sd,
                    y + sd,
                    color=colors[feature_set],
                    alpha=0.16,
                    linewidth=0
                )
    
            # Column titles only on first row
            if row_i == 0:
                ax.set_title(
                    model_title_map[model],
                    fontweight="bold",
                    pad=8
                )
    
            # Y-axis labels only on first column
            if col_i == 0:
                ax.set_ylabel(metric["ylabel"])
            else:
                ax.set_ylabel("")
    
            # X-axis labels only on bottom row
            if row_i == 2:
                ax.set_xlabel("Training data used (%)")
            else:
                ax.set_xticklabels([])
    
            ax.set_xlim(17, 103)
            ax.set_xticks([20, 40, 60, 80, 100])
    
            ax.set_ylim(metric["ylim"])
            ax.set_yticks(metric["yticks"])
    
            ax.grid(
                True,
                axis="both",
                color="#d9d9d9",
                linewidth=0.6,
                alpha=0.55
            )
    
            ax.set_axisbelow(True)
    
            # Panel label
            ax.text(
                0.02,
                1.04,
                panel_letters[panel_i],
                transform=ax.transAxes,
                fontsize=10,
                fontweight="bold",
                ha="left",
                va="bottom"
            )
    
            panel_i += 1
    
    # ============================================================
    # Legend
    # ============================================================
    handles, labels = axes[0, 0].get_legend_handles_labels()
    
    fig.legend(
        handles,
        labels,
        loc="upper center",
        bbox_to_anchor=(0.5, 0.985),
        ncol=2,
        frameon=False,
        fontsize=8
    )
    
    # ============================================================
    # Bottom note
    # ============================================================
    fig.text(
        0.5,
        0.035,
        "Lines show mean performance across training proportions; shaded bands indicate ±1 SD.",
        ha="center",
        va="center",
        fontsize=8
    )
    
    # ============================================================
    # A4 layout
    # Do not use bbox_inches='tight' because it changes A4 size
    # ============================================================
    plt.subplots_adjust(
        left=0.075,
        right=0.985,
        top=0.92,
        bottom=0.10,
        wspace=0.18,
        hspace=0.22
    )
    
    # ============================================================
    # Save outputs
    # ============================================================
    out_png = DATA_DIR / "O3_label_efficiency_3x3_A4.png"
    out_pdf = DATA_DIR / "O3_label_efficiency_3x3_A4.pdf"
    
    plt.savefig(out_png, dpi=300)
    plt.savefig(out_pdf)
    
    plt.close("all")
    
    print("Saved PNG to:", out_png)
    print("Saved PDF to:", out_pdf)
    
except Exception as exc:
    print('Warning: skipped optional exact-format figure section 2:', exc)
    traceback.print_exc(limit=1)


# Optional exact-format manuscript figure section 3
try:
    
    import pandas as pd
    import numpy as np
    import matplotlib.pyplot as plt
    from pathlib import Path
    
    # ============================================================
    # Data folder
    # Change this to your O2 folder if needed
    # ============================================================
    DATA_DIR = OUT_DIR / "O2_SE_added_value"
    
    # Main input file
    summary_file = DATA_DIR / "O2_SE_added_value_summary.csv"
    
    # If CSV is not found, try Excel file
    excel_file = DATA_DIR / "O2_SE_added_value_results.xlsx"
    
    if summary_file.exists():
        df = pd.read_csv(summary_file)
    elif excel_file.exists():
        df = pd.read_excel(excel_file, sheet_name="Summary")
    else:
        raise FileNotFoundError(
            f"Could not find:\n{summary_file}\n\nor:\n{excel_file}"
        )
    
    # ============================================================
    # Label cleaning
    # ============================================================
    def clean_feature_name(name):
        name = str(name)
        name = name.replace("RGBNIR", "MSI")
        name = name.replace("+", " + ")
        return name
    
    def clean_model_name(model):
        if model == "GradientBoosting":
            return "GB"
        return model
    
    df["With_SE_clean"] = df["With_SE"].apply(clean_feature_name)
    df["Model_clean"] = df["Model"].apply(clean_model_name)
    
    # ============================================================
    # Plot order
    # ============================================================
    row_order = [
        "CHM + SE",
        "MSI + SE",
        "CHM + MSI + SE"
    ]
    
    col_order = [
        "RF",
        "GB",
        "XGBoost"
    ]
    
    # ============================================================
    # Improvement calculations
    # ============================================================
    # R² improvement is calculated as relative improvement:
    # 100 * (R2_with_SE - R2_without_SE) / R2_without_SE
    df["R2_improvement_percent"] = (
        100 * (df["R2_with_SE_mean"] - df["R2_without_SE_mean"]) / df["R2_without_SE_mean"]
    )
    
    # RMSE and MAE improvement columns already exist in the summary file
    df["RMSE_improvement_percent"] = df["RMSE_percent_improvement_mean"]
    df["MAE_improvement_percent"] = df["MAE_percent_improvement_mean"]
    
    # ============================================================
    # Significance labels
    # ============================================================
    def p_to_stars(p):
        if p < 0.001:
            return "***"
        elif p < 0.01:
            return "**"
        elif p < 0.05:
            return "*"
        else:
            return "ns"
    
    # ============================================================
    # Build matrix for each metric
    # ============================================================
    def build_matrix(value_col, p_col):
        values = pd.DataFrame(index=row_order, columns=col_order, dtype=float)
        stars = pd.DataFrame(index=row_order, columns=col_order, dtype=object)
    
        for _, row in df.iterrows():
            r = row["With_SE_clean"]
            c = row["Model_clean"]
    
            if r in row_order and c in col_order:
                values.loc[r, c] = row[value_col]
                stars.loc[r, c] = p_to_stars(row[p_col])
    
        return values, stars
    
    r2_values, r2_stars = build_matrix(
        "R2_improvement_percent",
        "Wilcoxon_p_R2"
    )
    
    rmse_values, rmse_stars = build_matrix(
        "RMSE_improvement_percent",
        "Wilcoxon_p_RMSE"
    )
    
    mae_values, mae_stars = build_matrix(
        "MAE_improvement_percent",
        "Wilcoxon_p_MAE"
    )
    
    # ============================================================
    # Figure style
    # ============================================================
    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 8,
        "axes.titlesize": 10,
        "axes.labelsize": 9,
        "xtick.labelsize": 8,
        "ytick.labelsize": 8,
        "axes.linewidth": 0.8,
    })
    
    fig, axes = plt.subplots(
        1, 3,
        figsize=(11.69, 3.25),
        dpi=300
    )
    
    # Same green/yellow/red style
    cmap = "RdYlGn"
    
    # Color limits matching your example
    r2_vmin, r2_vmax = -5, 85
    err_vmin, err_vmax = -5, 30
    
    panels = [
        {
            "ax": axes[0],
            "values": r2_values,
            "stars": r2_stars,
            "title": "R² improvement (%)",
            "label": "(a)",
            "vmin": r2_vmin,
            "vmax": r2_vmax,
            "cbar_ticks": np.arange(0, 81, 10)
        },
        {
            "ax": axes[1],
            "values": rmse_values,
            "stars": rmse_stars,
            "title": "RMSE improvement (%)",
            "label": "(b)",
            "vmin": err_vmin,
            "vmax": err_vmax,
            "cbar_ticks": np.arange(-5, 31, 5)
        },
        {
            "ax": axes[2],
            "values": mae_values,
            "stars": mae_stars,
            "title": "MAE improvement (%)",
            "label": "(c)",
            "vmin": err_vmin,
            "vmax": err_vmax,
            "cbar_ticks": np.arange(-5, 31, 5)
        }
    ]
    
    # ============================================================
    # Draw heatmaps
    # ============================================================
    for panel in panels:
        ax = panel["ax"]
        values = panel["values"]
        stars = panel["stars"]
    
        im = ax.imshow(
            values.to_numpy(dtype=float),
            cmap=cmap,
            vmin=panel["vmin"],
            vmax=panel["vmax"],
            aspect="auto"
        )
    
        # Axis ticks
        ax.set_xticks(np.arange(len(col_order)))
        ax.set_xticklabels(col_order)
    
        ax.set_yticks(np.arange(len(row_order)))
        ax.set_yticklabels(row_order)
    
        # Only first panel has y-axis label and y tick labels
        if ax == axes[0]:
            ax.set_ylabel("Feature set after adding SE")
        else:
            ax.set_yticklabels([])
    
        # White cell borders
        ax.set_xticks(np.arange(-0.5, len(col_order), 1), minor=True)
        ax.set_yticks(np.arange(-0.5, len(row_order), 1), minor=True)
    
        ax.grid(
            which="minor",
            color="white",
            linestyle="-",
            linewidth=2
        )
    
        ax.tick_params(which="minor", bottom=False, left=False)
    
        # Cell text
        for i in range(len(row_order)):
            for j in range(len(col_order)):
                value = values.iloc[i, j]
                sig = stars.iloc[i, j]
    
                ax.text(
                    j,
                    i,
                    f"{value:.1f}%\n{sig}",
                    ha="center",
                    va="center",
                    color="black",
                    fontsize=8,
                    fontweight="bold" if sig != "ns" else "normal"
                )
    
        # Title and panel label
        ax.set_title(panel["title"], pad=8)
    
        ax.text(
            -0.23,
            1.04,
            panel["label"],
            transform=ax.transAxes,
            fontsize=10,
            fontweight="bold",
            ha="left",
            va="bottom"
        )
    
        # Colorbar
        cbar = fig.colorbar(
            im,
            ax=ax,
            fraction=0.046,
            pad=0.035,
            ticks=panel["cbar_ticks"]
        )
    
        cbar.ax.tick_params(labelsize=7)
    
    # ============================================================
    # Footnote
    # ============================================================
    fig.text(
        0.5,
        0.04,
        "Positive values indicate improved performance after adding SE predictors. "
        "* p < 0.05, ** p < 0.01, *** p < 0.001; ns = not significant.",
        ha="center",
        va="center",
        fontsize=8
    )
    
    # ============================================================
    # Layout
    # ============================================================
    plt.subplots_adjust(
        left=0.08,
        right=0.965,
        top=0.84,
        bottom=0.22,
        wspace=0.14
    )
    
    # ============================================================
    # Save
    # ============================================================
    out_png = DATA_DIR / "O2_SE_added_value_heatmap.png"
    out_pdf = DATA_DIR / "O2_SE_added_value_heatmap.pdf"
    
    plt.savefig(out_png, dpi=300)
    plt.savefig(out_pdf)
    
    plt.close("all")
    
    print("Saved PNG to:", out_png)
    print("Saved PDF to:", out_pdf)
    
except Exception as exc:
    print('Warning: skipped optional exact-format figure section 3:', exc)
    traceback.print_exc(limit=1)


# Optional exact-format manuscript figure section 4
try:
    
    import pandas as pd
    import matplotlib.pyplot as plt
    from pathlib import Path
    
    # ============================================================
    # Data folder
    # ============================================================
    DATA_DIR = OUT_DIR / "O3_label_efficiency"
    
    # Main input file
    summary_file = DATA_DIR / "O3_label_efficiency_SE_added_value_by_training_size.csv"
    
    if not summary_file.exists():
        raise FileNotFoundError(
            f"File not found:\n{summary_file}\n\n"
            "Check that O3_label_efficiency_SE_added_value_by_training_size.csv is inside this folder."
        )
    
    df = pd.read_csv(summary_file)
    
    # ============================================================
    # Clean model names
    # ============================================================
    model_name_map = {
        "RF": "RF",
        "GradientBoosting": "GB",
        "XGBoost": "XGBoost"
    }
    
    df["Model_plot"] = df["Model"].replace(model_name_map)
    
    # ============================================================
    # Plot order and colors
    # ============================================================
    model_order = ["RF", "GB", "XGBoost"]
    
    colors = {
        "RF": "#1f77b4",
        "GB": "#ff7f0e",
        "XGBoost": "#2ca02c"
    }
    
    # ============================================================
    # Figure style
    # ============================================================
    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 12,
        "axes.labelsize": 13,
        "xtick.labelsize": 12,
        "ytick.labelsize": 12,
        "legend.fontsize": 12,
        "legend.title_fontsize": 13,
        "axes.linewidth": 1.0,
    })
    
    fig, ax = plt.subplots(
        figsize=(8.6, 6.6),
        dpi=300
    )
    
    # ============================================================
    # Plot lines with ±1 SD error bars
    # ============================================================
    for model in model_order:
        plot_df = (
            df[df["Model_plot"] == model]
            .sort_values("Train_Proportion")
            .copy()
        )
    
        x = plot_df["Train_Proportion"] * 100
        y = plot_df["RMSE_percent_improvement_mean"]
        yerr = plot_df["RMSE_percent_improvement_sd"]
    
        ax.errorbar(
            x,
            y,
            yerr=yerr,
            marker="o",
            markersize=7,
            linewidth=2.2,
            elinewidth=1.2,
            capsize=4,
            capthick=1.0,
            color=colors[model],
            alpha=0.95,
            label=model
        )
    
    # ============================================================
    # Zero reference line
    # ============================================================
    ax.axhline(
        0,
        color="black",
        linestyle="--",
        linewidth=1.2,
        alpha=0.65
    )
    
    # ============================================================
    # Axes, labels, ticks
    # ============================================================
    ax.set_xlabel("Training data used (%)")
    ax.set_ylabel("RMSE improvement from SE (%)")
    
    ax.set_xlim(16, 104)
    ax.set_xticks([20, 30, 40, 50, 60, 70, 80, 90, 100])
    
    ax.set_ylim(-32, 18)
    ax.set_yticks([-30, -20, -10, 0, 10])
    
    # Grid
    ax.grid(
        True,
        axis="both",
        color="#d9d9d9",
        linewidth=0.7,
        alpha=0.55
    )
    
    ax.set_axisbelow(True)
    
    # Legend
    ax.legend(
        title="Model",
        loc="upper right",
        frameon=True
    )
    
    # ============================================================
    # Layout and save
    # ============================================================
    plt.tight_layout()
    
    out_png = DATA_DIR / "O3_RMSE_improvement_from_SE_by_training_size.png"
    out_pdf = DATA_DIR / "O3_RMSE_improvement_from_SE_by_training_size.pdf"
    
    plt.savefig(out_png, dpi=300)
    plt.savefig(out_pdf)
    
    plt.close("all")
    
    print("Saved PNG to:", out_png)
    print("Saved PDF to:", out_pdf)
except Exception as exc:
    print('Warning: skipped optional exact-format figure section 4:', exc)
    traceback.print_exc(limit=1)




# ============================================================
# FINAL MANUSCRIPT FIGURES WITH FIXED MARGINS
# This block regenerates only the final manuscript figures from
# the CSV outputs created above. It does NOT rerun models.
# Figures are saved before final collection so they are included
# in ALL_FIGURES_TOGETHER and the final zip package.
# ============================================================
from typing import Optional, Iterable
from matplotlib.patches import Patch
from matplotlib.lines import Line2D
from matplotlib.ticker import MultipleLocator, FormatStrFormatter
from matplotlib.gridspec import GridSpec

def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def save_fig(fig: plt.Figure, out_base: Path, dpi: int = 300, tight: bool = False) -> None:
    out_base.parent.mkdir(parents=True, exist_ok=True)
    if tight:
        fig.savefig(out_base.with_suffix(".png"), dpi=dpi, bbox_inches="tight", pad_inches=0.08)
        fig.savefig(out_base.with_suffix(".pdf"), bbox_inches="tight", pad_inches=0.08)
    else:
        fig.savefig(out_base.with_suffix(".png"), dpi=dpi)
        fig.savefig(out_base.with_suffix(".pdf"))
    plt.close(fig)
    print("Saved:", out_base.with_suffix(".png"))
    print("Saved:", out_base.with_suffix(".pdf"))


def first_existing(paths: Iterable[Path]) -> Optional[Path]:
    for p in paths:
        if p.exists():
            return p
    return None


def find_file(root: Path, filename: str) -> Optional[Path]:
    direct = root / filename
    if direct.exists():
        return direct
    matches = list(root.rglob(filename))
    if matches:
        # Prefer shortest path / likely original, not collected copy.
        matches = sorted(matches, key=lambda p: ("FINAL_NECESSARY" in str(p), len(str(p))))
        return matches[0]
    return None


def clean_feature_set_name(name: str) -> str:
    return str(name).replace("RGBNIR", "MSI")


def clean_model_name(name: str) -> str:
    return {"GradientBoosting": "GB", "Gradient_Boosting": "GB"}.get(str(name), str(name))


def safe_label(text: str) -> str:
    return str(text).replace("RGBNIR", "MSI")


# ============================================================
# Figure 1: A4 top predictors
# ============================================================

def generate_top_predictors(o1_dir: Path, out_dir: Path, dpi: int, xlabel: str) -> None:
    data_dir = o1_dir / "O4_SHAP_and_complementarity" / "O4_algorithm_specific_SHAP"

    all_file = data_dir / "All_algorithm_specific_SHAP_predictor_importance.csv"
    overall_file = data_dir / "Overall_mean_SHAP_importance_across_algorithms.csv"

    if not all_file.exists():
        all_file = find_file(o1_dir, "All_algorithm_specific_SHAP_predictor_importance.csv")
    if not overall_file.exists():
        overall_file = find_file(o1_dir, "Overall_mean_SHAP_importance_across_algorithms.csv")

    if all_file is None or not Path(all_file).exists():
        raise FileNotFoundError("Could not find All_algorithm_specific_SHAP_predictor_importance.csv")
    if overall_file is None or not Path(overall_file).exists():
        raise FileNotFoundError("Could not find Overall_mean_SHAP_importance_across_algorithms.csv")

    all_df = pd.read_csv(all_file)
    overall_df = pd.read_csv(overall_file)

    if "Mean_abs_SHAP" not in all_df.columns:
        raise ValueError(f"Expected column 'Mean_abs_SHAP' in {all_file}")
    if "Mean_abs_SHAP_overall" not in overall_df.columns:
        raise ValueError(f"Expected column 'Mean_abs_SHAP_overall' in {overall_file}")

    def clean_group(group: str) -> str:
        group = str(group)
        if group in ["RGB-NIR", "RGBNIR", "MSI", "RGB_NIR"]:
            return "MSI"
        return group

    all_df["Plot_Group"] = all_df["Group"].apply(clean_group)
    overall_df["Plot_Group"] = overall_df["Group"].apply(clean_group)

    if "Model_plot" not in all_df.columns:
        all_df["Model_plot"] = all_df["Model"].apply(clean_model_name)

    group_colors = {
        "CHM": "#6aaed6",
        "MSI": "#e7ad72",
        "SE": "#7bd88f",
    }

    overall_top = (
        overall_df
        .sort_values("Mean_abs_SHAP_overall", ascending=False)
        .head(20)
    )

    per_model = {}
    for m in ["RF", "GB", "XGBoost"]:
        mdf = all_df[all_df["Model_plot"] == m].copy()
        if mdf.empty:
            continue
        per_model[m] = mdf.sort_values("Mean_abs_SHAP", ascending=False).head(10)

    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 8,
        "axes.titlesize": 10,
        "axes.labelsize": 8,
        "xtick.labelsize": 8,
        "ytick.labelsize": 8,
        "axes.linewidth": 0.9,
    })

    fig, axes = plt.subplots(2, 2, figsize=(11.69, 8.27), dpi=dpi)
    axes = axes.flatten()

    def plot_barh(ax, df, title, panel_label, tick_step=0.02):
        value_col = "Mean_abs_SHAP_overall" if "Mean_abs_SHAP_overall" in df.columns else "Mean_abs_SHAP"
        df_plot = df.sort_values(value_col, ascending=True).copy()
        colors = df_plot["Plot_Group"].map(group_colors).fillna("#bdbdbd")
        xmax = float(df[value_col].max()) * 1.10 if len(df) else 0.1
        xmax = max(xmax, 0.01)

        ax.barh(
            df_plot["Predictor"],
            df_plot[value_col],
            color=colors,
            edgecolor=colors,
            height=0.70,
        )
        ax.set_title(title, fontweight="bold", pad=7)
        ax.set_xlabel("Mean |SHAP value|")
        ax.set_xlim(0, xmax)
        ax.xaxis.set_major_locator(MultipleLocator(tick_step))
        ax.xaxis.set_major_formatter(FormatStrFormatter("%.2f"))
        ax.grid(axis="x", linestyle="--", linewidth=0.6, alpha=0.35)
        ax.set_axisbelow(True)
        ax.tick_params(axis="y", labelsize=8, pad=2)
        ax.tick_params(axis="x", labelsize=8)
        ax.text(
            0.0, 1.045, panel_label,
            transform=ax.transAxes,
            fontsize=10,
            fontweight="bold",
            ha="left",
            va="bottom",
            clip_on=False,
        )

    plot_barh(axes[0], overall_top, "Overall Top 20 Predictors", "(a)")
    plot_barh(axes[1], per_model.get("RF", overall_top.head(10)), "RF | Top 10 Predictors", "(b)")
    plot_barh(axes[2], per_model.get("GB", overall_top.head(10)), "GB | Top 10 Predictors", "(c)")
    plot_barh(axes[3], per_model.get("XGBoost", overall_top.head(10)), "XGBoost | Top 10 Predictors", "(d)")

    legend_handles = [
        Patch(facecolor=group_colors["CHM"], edgecolor="black", label="CHM"),
        Patch(facecolor=group_colors["MSI"], edgecolor="black", label="MSI"),
        Patch(facecolor=group_colors["SE"], edgecolor="black", label="SE"),
    ]
    fig.legend(
        handles=legend_handles,
        loc="lower center",
        ncol=3,
        frameon=False,
        bbox_to_anchor=(0.5, 0.035),
        fontsize=8,
    )

    fig.subplots_adjust(
        left=0.185,
        right=0.985,
        top=0.94,
        bottom=0.12,
        wspace=0.32,
        hspace=0.28,
    )

    save_fig(fig, out_dir / "Top_predictors_A4_exact_style_FIXED", dpi=dpi, tight=False)


# ============================================================
# Figure 2: SHAP interaction + complementarity A4 landscape
# ============================================================

def generate_shap_interaction(o1_dir: Path, out_dir: Path, dpi: int) -> None:
    o4_dir = o1_dir / "O4_SHAP_and_complementarity"
    interaction_file = first_existing([
        o4_dir / "O5_SHAP_interaction_values.csv",
        o1_dir / "O5_SHAP_interaction_values.csv",
    ]) or find_file(o1_dir, "O5_SHAP_interaction_values.csv")
    correlation_file = first_existing([
        o4_dir / "O5_SHAP_complementarity_correlation.csv",
        o1_dir / "O5_SHAP_complementarity_correlation.csv",
    ]) or find_file(o1_dir, "O5_SHAP_complementarity_correlation.csv")

    if interaction_file is None or correlation_file is None:
        raise FileNotFoundError(
            "Could not find O5_SHAP_interaction_values.csv and/or "
            "O5_SHAP_complementarity_correlation.csv"
        )

    interaction = pd.read_csv(interaction_file)
    correlation = pd.read_csv(correlation_file)

    interaction["Feature pair"] = interaction["SE_Feature"].astype(str) + " x " + interaction["Other_Feature"].astype(str)
    correlation["Feature pair"] = correlation["SE_Feature"].astype(str) + " x " + correlation["Other_Feature"].astype(str)

    top_n = min(35, len(interaction), len(correlation))
    left_df = interaction.sort_values("Mean_abs_SHAP_Interaction", ascending=False).head(top_n).copy()
    right_df = correlation.sort_values("Abs_SHAP_Correlation", ascending=False).head(top_n).copy()

    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 7,
        "axes.titlesize": 9,
        "axes.labelsize": 8,
        "xtick.labelsize": 7,
        "ytick.labelsize": 5.9,
        "axes.linewidth": 0.8,
    })

    fig, axes = plt.subplots(1, 2, figsize=(11.69, 8.27), dpi=dpi)
    bar_color = "#cd9778"

    def make_panel(ax, df, value_col, title, xlabel, tick_step, tick_format):
        df_plot = df.sort_values(value_col, ascending=True).copy()
        y = np.arange(len(df_plot))
        xmax = float(df[value_col].max()) * 1.10 if len(df) else 1.0
        ax.barh(y, df_plot[value_col], color=bar_color, edgecolor=bar_color, height=0.68)
        ax.set_yticks(y)
        ax.set_yticklabels(df_plot["Feature pair"], fontsize=5.9)
        ax.set_title(title, fontsize=9, pad=8)
        ax.set_xlabel(xlabel, fontsize=8)
        ax.set_ylabel("Feature pair", fontsize=8, labelpad=8)
        ax.set_xlim(0, xmax)
        ax.xaxis.set_major_locator(MultipleLocator(tick_step))
        ax.xaxis.set_major_formatter(FormatStrFormatter(tick_format))
        ax.xaxis.grid(True, color="#e6e6e6", linewidth=0.6)
        ax.set_axisbelow(True)
        ax.tick_params(axis="y", length=3, pad=2, labelsize=5.9)
        ax.tick_params(axis="x", length=3, labelsize=7)

    make_panel(
        axes[0],
        left_df,
        "Mean_abs_SHAP_Interaction",
        "SHAP interaction strength: SE with CHM/MSI",
        "Mean absolute SHAP interaction value",
        tick_step=0.004,
        tick_format="%.3f",
    )
    make_panel(
        axes[1],
        right_df,
        "Abs_SHAP_Correlation",
        "Complementarity between SE and CHM/MSI predictors",
        "Absolute correlation between SHAP values",
        tick_step=0.1,
        tick_format="%.1f",
    )

    axes[0].text(-0.08, 1.025, "(a)", transform=axes[0].transAxes, fontsize=11, fontweight="bold", clip_on=False)
    axes[1].text(-0.08, 1.025, "(b)", transform=axes[1].transAxes, fontsize=11, fontweight="bold", clip_on=False)

    fig.subplots_adjust(left=0.245, right=0.985, top=0.935, bottom=0.085, wspace=0.72)
    save_fig(fig, out_dir / "SHAP_interaction_complementarity_A4_landscape_FIXED", dpi=dpi, tight=False)


# ============================================================
# Figure 3: O3 label efficiency 3 x 3 A4
# ============================================================

def generate_o3_label_efficiency(o1_dir: Path, out_dir: Path, dpi: int) -> None:
    data_dir = o1_dir / "O3_label_efficiency"
    summary_file = data_dir / "O3_label_efficiency_summary.csv"
    if not summary_file.exists():
        summary_file = find_file(o1_dir, "O3_label_efficiency_summary.csv")
    if summary_file is None:
        raise FileNotFoundError("Could not find O3_label_efficiency_summary.csv")

    df = pd.read_csv(summary_file)
    df["Feature_Set_Plot"] = df["Feature_Set"].replace({
        "CHM+RGBNIR": "CHM+MSI",
        "CHM+RGBNIR+SE": "CHM+MSI+SE",
    })

    model_order = ["RF", "GradientBoosting", "XGBoost"]
    model_title_map = {"RF": "RF", "GradientBoosting": "GB", "XGBoost": "XGBoost"}
    feature_order = ["CHM+MSI", "CHM+MSI+SE"]
    colors = {"CHM+MSI": "#1f77b4", "CHM+MSI+SE": "#ff7f0e"}

    metrics = [
        {"mean_col": "R2_mean", "sd_col": "R2_sd", "ylabel": "R²", "ylim": (-0.55, 0.72), "yticks": [-0.4, -0.2, 0.0, 0.2, 0.4, 0.6]},
        {"mean_col": "RMSE_mean", "sd_col": "RMSE_sd", "ylabel": "RMSE", "ylim": (0.105, 0.245), "yticks": [0.12, 0.14, 0.16, 0.18, 0.20, 0.22, 0.24]},
        {"mean_col": "MAE_mean", "sd_col": "MAE_sd", "ylabel": "MAE", "ylim": (0.08, 0.16), "yticks": [0.08, 0.09, 0.10, 0.11, 0.12, 0.13, 0.14, 0.15, 0.16]},
    ]

    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 8,
        "axes.titlesize": 9,
        "axes.labelsize": 8,
        "xtick.labelsize": 8,
        "ytick.labelsize": 8,
        "axes.linewidth": 0.9,
    })

    fig, axes = plt.subplots(3, 3, figsize=(11.69, 8.27), dpi=dpi)
    panel_letters = ["(a)", "(b)", "(c)", "(d)", "(e)", "(f)", "(g)", "(h)", "(i)"]
    panel_i = 0

    for row_i, metric in enumerate(metrics):
        for col_i, model in enumerate(model_order):
            ax = axes[row_i, col_i]
            model_df = df[df["Model"] == model].copy()

            for feature_set in feature_order:
                plot_df = model_df[model_df["Feature_Set_Plot"] == feature_set].sort_values("Train_Proportion").copy()
                if plot_df.empty:
                    continue
                x = plot_df["Train_Proportion"].to_numpy() * 100
                y = plot_df[metric["mean_col"]].to_numpy(dtype=float)
                sd = plot_df[metric["sd_col"]].fillna(0).to_numpy(dtype=float)

                ax.plot(x, y, marker="o", markersize=4.5, linewidth=1.8, color=colors[feature_set], label=feature_set)
                ax.fill_between(x, y - sd, y + sd, color=colors[feature_set], alpha=0.16, linewidth=0)

            if row_i == 0:
                ax.set_title(model_title_map[model], fontweight="bold", pad=8)
            if col_i == 0:
                ax.set_ylabel(metric["ylabel"])
            if row_i == 2:
                ax.set_xlabel("Training data used (%)")
            else:
                ax.set_xticklabels([])

            ax.set_xlim(17, 103)
            ax.set_xticks([20, 40, 60, 80, 100])
            ax.set_ylim(metric["ylim"])
            ax.set_yticks(metric["yticks"])
            ax.grid(True, axis="both", color="#d9d9d9", linewidth=0.6, alpha=0.55)
            ax.set_axisbelow(True)
            ax.text(0.02, 1.04, panel_letters[panel_i], transform=ax.transAxes, fontsize=10, fontweight="bold", ha="left", va="bottom", clip_on=False)
            panel_i += 1

    handles, labels = axes[0, 0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="upper center", bbox_to_anchor=(0.5, 0.985), ncol=2, frameon=False, fontsize=8)
    fig.text(0.5, 0.035, "Lines show mean performance across training proportions; shaded bands indicate ±1 SD.", ha="center", va="center", fontsize=8)
    fig.subplots_adjust(left=0.075, right=0.985, top=0.92, bottom=0.10, wspace=0.18, hspace=0.22)

    save_fig(fig, out_dir / "O3_label_efficiency_3x3_A4_FIXED", dpi=dpi, tight=False)


# ============================================================
# Figure 4: O3 RMSE improvement from SE by training size
# ============================================================

def generate_o3_rmse_improvement(o1_dir: Path, out_dir: Path, dpi: int) -> None:
    data_dir = o1_dir / "O3_label_efficiency"
    summary_file = data_dir / "O3_label_efficiency_SE_added_value_by_training_size.csv"
    if not summary_file.exists():
        summary_file = find_file(o1_dir, "O3_label_efficiency_SE_added_value_by_training_size.csv")
    if summary_file is None:
        raise FileNotFoundError("Could not find O3_label_efficiency_SE_added_value_by_training_size.csv")

    df = pd.read_csv(summary_file)
    df["Model_plot"] = df["Model"].replace({"GradientBoosting": "GB", "Gradient_Boosting": "GB"})

    model_order = ["RF", "GB", "XGBoost"]
    colors = {"RF": "#1f77b4", "GB": "#ff7f0e", "XGBoost": "#2ca02c"}

    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 12,
        "axes.labelsize": 13,
        "xtick.labelsize": 12,
        "ytick.labelsize": 12,
        "legend.fontsize": 12,
        "legend.title_fontsize": 13,
        "axes.linewidth": 1.0,
    })

    fig, ax = plt.subplots(figsize=(8.6, 6.6), dpi=dpi)

    for model in model_order:
        plot_df = df[df["Model_plot"] == model].sort_values("Train_Proportion").copy()
        if plot_df.empty:
            continue
        x = plot_df["Train_Proportion"] * 100
        y = plot_df["RMSE_percent_improvement_mean"]
        yerr = plot_df.get("RMSE_percent_improvement_sd", pd.Series(np.zeros(len(plot_df)), index=plot_df.index)).fillna(0)
        ax.errorbar(x, y, yerr=yerr, marker="o", markersize=7, linewidth=2.2, elinewidth=1.2, capsize=4, capthick=1.0, color=colors[model], alpha=0.95, label=model)

    ax.axhline(0, color="black", linestyle="--", linewidth=1.2, alpha=0.65)
    ax.set_xlabel("Training data used (%)")
    ax.set_ylabel("RMSE improvement from SE (%)")
    ax.set_xlim(16, 104)
    ax.set_xticks([20, 30, 40, 50, 60, 70, 80, 90, 100])

    yvals = df["RMSE_percent_improvement_mean"].replace([np.inf, -np.inf], np.nan).dropna()
    if len(yvals):
        ymin = min(-32, float((df["RMSE_percent_improvement_mean"] - df.get("RMSE_percent_improvement_sd", 0)).min()) - 2)
        ymax = max(18, float((df["RMSE_percent_improvement_mean"] + df.get("RMSE_percent_improvement_sd", 0)).max()) + 2)
        ax.set_ylim(ymin, ymax)
    else:
        ax.set_ylim(-32, 18)

    ax.grid(True, axis="both", color="#d9d9d9", linewidth=0.7, alpha=0.55)
    ax.set_axisbelow(True)
    ax.legend(title="Model", loc="upper right", frameon=True)
    fig.tight_layout(pad=1.2)

    save_fig(fig, out_dir / "O3_RMSE_improvement_from_SE_by_training_size_FIXED", dpi=dpi, tight=True)


# ============================================================
# Figure 5: O7 plot-specific SE benefit split dotplot
# ============================================================

def generate_o7_plot_variable_dotplot(o1_dir: Path, out_dir: Path, dpi: int) -> None:
    data_dir = o1_dir / "O7_plot_specific_SE_benefit"
    corr_file = data_dir / "O7_plot_variable_correlations_with_SE_improvement.csv"

    # Backward-compatible fallback for old run folders.
    if not corr_file.exists():
        corr_file = find_file(o1_dir, "O7_plot_variable_correlations_with_SE_improvement.csv")
    if corr_file is None:
        corr_file = find_file(o1_dir, "O7_condition_correlations_with_SE_improvement.csv")
    if corr_file is None:
        raise FileNotFoundError(
            "Could not find O7_plot_variable_correlations_with_SE_improvement.csv"
        )

    corr_df = pd.read_csv(corr_file)
    corr_df.columns = corr_df.columns.str.strip()
    corr_df["Model"] = corr_df["Model"].replace({
        "GradientBoosting": "GB",
        "Gradient_Boosting": "GB"
    })

    model_order = ["RF", "GB", "XGBoost"]
    model_colors = {"RF": "#1f77b4", "GB": "#ff7f0e", "XGBoost": "#2ca02c"}
    model_markers = {"RF": "o", "GB": "s", "XGBoost": "^"}
    corr_col = "Spearman_Correlation_with_SE_Improvement"

    if corr_col not in corr_df.columns:
        raise ValueError(f"Missing column: {corr_col}")

    if "Plot_Variable" not in corr_df.columns and "Condition_Variable" in corr_df.columns:
        corr_df = corr_df.rename(columns={"Condition_Variable": "Plot_Variable"})

    if "Plot_Variable" not in corr_df.columns:
        raise ValueError("Missing column: Plot_Variable")

    corr_df["Plot_Label"] = (
        corr_df["Plot_Variable"]
        .astype(str)
        .str.replace("RGBNIR", "MSI", regex=False)
        .str.replace("_", " ", regex=False)
    )

    corr_df["Full_Label"] = corr_df["Model"] + " | " + corr_df["Plot_Label"]

    n_show = 10

    neg_df = (
        corr_df[corr_df[corr_col] < 0]
        .sort_values(corr_col, ascending=True)
        .head(n_show)
        .copy()
    )

    pos_df = (
        corr_df[corr_df[corr_col] > 0]
        .sort_values(corr_col, ascending=False)
        .head(n_show)
        .copy()
    )

    neg_df = neg_df.sort_values(corr_col, ascending=False).reset_index(drop=True)
    pos_df = pos_df.sort_values(corr_col, ascending=True).reset_index(drop=True)

    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 9,
        "axes.labelsize": 9,
        "xtick.labelsize": 9,
        "ytick.labelsize": 8.5,
        "axes.linewidth": 0.9,
    })

    fig, axes = plt.subplots(
        nrows=1,
        ncols=2,
        figsize=(13.4, 5.9),
        dpi=dpi,
        sharex=False
    )

    def panel(ax, df, positive: bool):
        y = np.arange(len(df))

        if df.empty:
            ax.axvline(0, linestyle="--", linewidth=1, color="black")
            ax.set_yticks([])
            ax.set_xlabel("Spearman correlation", fontsize=9)
            ax.text(
                0.5,
                0.5,
                "No associations",
                transform=ax.transAxes,
                ha="center",
                va="center",
                fontsize=9
            )
            ax.set_xlim(-0.1, 0.1)
            return

        for i, row in df.iterrows():
            model = row["Model"]

            ax.scatter(
                row[corr_col],
                i,
                color=model_colors.get(model, "gray"),
                marker=model_markers.get(model, "o"),
                s=55,
                zorder=3
            )

            if positive:
                ax.text(
                    row[corr_col] + 0.006,
                    i,
                    f"{row[corr_col]:.2f}",
                    ha="left",
                    va="center",
                    fontsize=8,
                    clip_on=False
                )
            else:
                ax.text(
                    row[corr_col] - 0.006,
                    i,
                    f"{row[corr_col]:.2f}",
                    ha="right",
                    va="center",
                    fontsize=8,
                    clip_on=False
                )

        ax.axvline(0, linestyle="--", linewidth=1, color="black")
        ax.set_yticks(y)
        ax.set_yticklabels(df["Full_Label"], fontsize=8.5)
        ax.set_xlabel("Spearman correlation", fontsize=9)
        ax.grid(axis="x", alpha=0.25)
        ax.set_axisbelow(True)

    panel(axes[0], neg_df, positive=False)

    if not neg_df.empty:
        axes[0].set_xlim(min(neg_df[corr_col].min() - 0.04, -0.22), 0.02)

    axes[0].text(
        0.01,
        1.04,
        "(a)",
        transform=axes[0].transAxes,
        fontsize=11,
        fontweight="bold",
        ha="left",
        va="bottom",
        clip_on=False
    )

    axes[0].text(
        1.12,
        0.5,
        "Plot-level variables associated with weaker SE benefit",
        transform=axes[0].transAxes,
        rotation=-90,
        fontsize=9,
        ha="center",
        va="center",
        clip_on=False
    )

    panel(axes[1], pos_df, positive=True)

    if not pos_df.empty:
        axes[1].set_xlim(-0.02, max(pos_df[corr_col].max() + 0.04, 0.14))

    axes[1].text(
        0.01,
        1.04,
        "(b)",
        transform=axes[1].transAxes,
        fontsize=11,
        fontweight="bold",
        ha="left",
        va="bottom",
        clip_on=False
    )

    axes[1].text(
        1.12,
        0.5,
        "Plot-level variables associated with stronger SE benefit",
        transform=axes[1].transAxes,
        rotation=-90,
        fontsize=9,
        ha="center",
        va="center",
        clip_on=False
    )

    legend_handles = [
        Line2D(
            [0],
            [0],
            marker=model_markers[m],
            color="w",
            markerfacecolor=model_colors[m],
            markeredgecolor=model_colors[m],
            markersize=7,
            label=m
        )
        for m in model_order
    ]

    fig.legend(
        handles=legend_handles,
        loc="upper center",
        ncol=3,
        frameon=False,
        bbox_to_anchor=(0.5, 1.02),
        fontsize=9
    )

    fig.text(
        0.5,
        0.015,
        (
            "Positive correlations indicate greater SE benefit at higher plot-level variable values; "
            "negative correlations indicate weaker SE benefit."
        ),
        ha="center",
        fontsize=8
    )

    fig.subplots_adjust(
        left=0.20,
        right=0.91,
        top=0.88,
        bottom=0.14,
        wspace=0.78
    )

    save_fig(
        fig,
        out_dir / "Figure_O7_split_dotplot_plot_variable_associations_FIXED",
        dpi=dpi,
        tight=True
    )


# Backward-compatible function name for old task lists.
generate_o7_condition_dotplot = generate_o7_plot_variable_dotplot
# ============================================================
# Figure 6: Observed vs predicted overview grid
# ============================================================

def generate_observed_predicted_grid(o1_dir: Path, out_dir: Path, dpi: int) -> None:
    pred_file = o1_dir / "O1_all_models_mean_repeated_outer_test_predictions.csv"
    if not pred_file.exists():
        pred_file = find_file(o1_dir, "O1_all_models_mean_repeated_outer_test_predictions.csv")
    if pred_file is None:
        raise FileNotFoundError("Could not find O1_all_models_mean_repeated_outer_test_predictions.csv")

    df = pd.read_csv(pred_file)
    required = {"Feature_Set", "Model", "Observed", "Predicted_Mean"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Prediction file is missing columns: {missing}")

    feature_sets_left = ["CHM", "RGBNIR", "SE", "CHM+RGBNIR", "CHM+SE", "RGBNIR+SE"]
    full_feature_set = "CHM+RGBNIR+SE"
    model_order = ["RF", "GradientBoosting", "XGBoost"]
    model_title = {"RF": "RF", "GradientBoosting": "GB", "XGBoost": "XGBoost"}

    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 7,
        "axes.titlesize": 7,
        "axes.labelsize": 8,
        "xtick.labelsize": 6,
        "ytick.labelsize": 6,
        "axes.linewidth": 0.8,
    })

    fig = plt.figure(figsize=(8.27, 11.69), dpi=dpi)  # A4 portrait
    gs = GridSpec(6, 5, figure=fig, width_ratios=[1, 1, 1, 0.18, 1.65], wspace=0.42, hspace=0.52)

    all_obs = df["Observed"].to_numpy(dtype=float)
    all_pred = df["Predicted_Mean"].to_numpy(dtype=float)
    lim_min = min(0.0, np.nanmin([all_obs.min(), all_pred.min()]))
    lim_max = max(0.95, np.nanmax([all_obs.max(), all_pred.max()]))
    lims = (lim_min, lim_max)

    def scatter_panel(ax, sub, title, big=False):
        ax.scatter(sub["Observed"], sub["Predicted_Mean"], s=10 if not big else 14, color="#1f77b4", alpha=0.70, edgecolor="none")
        ax.plot(lims, lims, "--", color="red", linewidth=0.9)
        ax.set_xlim(lims)
        ax.set_ylim(lims)
        ax.set_title(title, fontweight="bold", pad=4, fontsize=7 if not big else 9)
        ax.grid(True, color="#d9d9d9", linewidth=0.45, alpha=0.55)
        ax.set_axisbelow(True)
        ax.tick_params(labelsize=6 if not big else 7)

    # Left matrix: six feature sets x three models.
    for row_i, fs in enumerate(feature_sets_left):
        for col_i, model in enumerate(model_order):
            ax = fig.add_subplot(gs[row_i, col_i])
            sub = df[(df["Feature_Set"] == fs) & (df["Model"] == model)]
            if sub.empty:
                ax.axis("off")
                continue
            title = f"{clean_feature_set_name(fs)} | {model_title[model]}"
            scatter_panel(ax, sub, title, big=False)
            if row_i == len(feature_sets_left) - 1:
                ax.set_xlabel("Observed", fontsize=7)
            if col_i == 0:
                ax.set_ylabel("Predicted", fontsize=7)

    # Right larger panels: full predictor set for each model.
    spans = [(0, 2), (2, 4), (4, 6)]
    for model, (r0, r1) in zip(model_order, spans):
        ax = fig.add_subplot(gs[r0:r1, 4])
        sub = df[(df["Feature_Set"] == full_feature_set) & (df["Model"] == model)]
        if sub.empty:
            ax.axis("off")
            continue
        title = f"CHM+MSI+SE | {model_title[model]}"
        scatter_panel(ax, sub, title, big=True)
        ax.set_xlabel("Observed", fontsize=8)
        ax.set_ylabel("Predicted", fontsize=8)

    fig.subplots_adjust(left=0.07, right=0.97, top=0.98, bottom=0.05)
    save_fig(fig, out_dir / "Observed_vs_predicted_grid_A4_FIXED", dpi=dpi, tight=False)




try:
    FINAL_FIXED_FIG_DIR = RUN_ROOT / "FINAL_NECESSARY_FIGURES_FIXED"
    ensure_dir(FINAL_FIXED_FIG_DIR)

    print("\n============================================================")
    print("Generating final manuscript figures with fixed margins")
    print("============================================================")
    print("Output folder:", FINAL_FIXED_FIG_DIR)

    _final_figure_tasks = [
        ("Top predictors A4", lambda: generate_top_predictors(OUT_DIR, FINAL_FIXED_FIG_DIR, 300, "Mean |SHAP value|")),
        ("SHAP interaction/complementarity A4", lambda: generate_shap_interaction(OUT_DIR, FINAL_FIXED_FIG_DIR, 300)),
        ("O3 label-efficiency 3x3 A4", lambda: generate_o3_label_efficiency(OUT_DIR, FINAL_FIXED_FIG_DIR, 300)),
        ("O3 RMSE improvement from SE", lambda: generate_o3_rmse_improvement(OUT_DIR, FINAL_FIXED_FIG_DIR, 300)),
        ("O7 plot-specific SE benefit dotplot", lambda: generate_o7_plot_variable_dotplot(OUT_DIR, FINAL_FIXED_FIG_DIR, 300)),
        ("Observed-vs-predicted grid", lambda: generate_observed_predicted_grid(OUT_DIR, FINAL_FIXED_FIG_DIR, 300)),
    ]

    _generated_final_figures = []
    _skipped_final_figures = []

    for _label, _func in _final_figure_tasks:
        print("\nGenerating:", _label)
        try:
            _func()
            _generated_final_figures.append(_label)
        except Exception as _exc:
            _skipped_final_figures.append((_label, str(_exc)))
            print("WARNING: skipped", _label)
            print(_exc)
            traceback.print_exc(limit=1)

    print("\nFinal fixed figures generated:")
    for _label in _generated_final_figures:
        print(" -", _label)

    if _skipped_final_figures:
        print("\nFinal fixed figures skipped:")
        for _label, _reason in _skipped_final_figures:
            print(f" - {_label}: {_reason}")
        print("Skipped figures usually mean the required CSV output was not produced in this run.")

except Exception as exc:
    print("Warning: skipped final fixed manuscript figure block:", exc)
    traceback.print_exc(limit=1)



# ============================================================
# FINAL MANUSCRIPT TABLES: generate tables as before
# ============================================================

print("\nGenerating final manuscript-style tables...")

FINAL_TABLE_DIR = RUN_ROOT / "FINAL_TABLES_AS_BEFORE"
FINAL_TABLE_DIR.mkdir(parents=True, exist_ok=True)


def _display_feature_set(value):
    """Use manuscript naming while keeping original analysis names in the raw columns."""
    return str(value).replace("RGBNIR", "MSI")


def _display_model(value):
    value = str(value)
    if value == "GradientBoosting":
        return "GB"
    return value


def _display_comparison(value):
    return str(value).replace("RGBNIR", "MSI")


def _fmt_mean_sd(mean_val, sd_val, decimals=3):
    try:
        if pd.isna(mean_val):
            return ""
        if pd.isna(sd_val):
            return f"{mean_val:.{decimals}f}"
        return f"{mean_val:.{decimals}f} ± {sd_val:.{decimals}f}"
    except Exception:
        return ""


def _round_numeric(df, digits=4):
    out = df.copy()
    for col in out.columns:
        if pd.api.types.is_numeric_dtype(out[col]):
            out[col] = out[col].round(digits)
    return out


def _safe_table(df, name):
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        print(f"Skipping empty table: {name}")
        return None
    out = df.copy()
    out.to_csv(FINAL_TABLE_DIR / f"{name}.csv", index=False)
    return out


final_tables = {}

# ------------------------------------------------------------
# Table 1. Main nested-CV model comparison
# ------------------------------------------------------------
try:
    t1 = summary_df.copy()
    t1["Predictor_Set"] = t1["Feature_Set"].apply(_display_feature_set)
    t1["Model_Display"] = t1["Model"].apply(_display_model)
    t1["R2_mean_sd"] = [
        _fmt_mean_sd(m, s, 3)
        for m, s in zip(t1["R2_mean"], t1["R2_sd"])
    ]
    t1["RMSE_mean_sd"] = [
        _fmt_mean_sd(m, s, 3)
        for m, s in zip(t1["RMSE_mean"], t1["RMSE_sd"])
    ]
    t1["MAE_mean_sd"] = [
        _fmt_mean_sd(m, s, 3)
        for m, s in zip(t1["MAE_mean"], t1["MAE_sd"])
    ]
    t1["N_Selected_Features_mean_sd"] = [
        _fmt_mean_sd(m, s, 1)
        for m, s in zip(t1["N_Selected_Features_mean"], t1["N_Selected_Features_sd"])
    ]
    t1 = t1[[
        "Predictor_Set", "Model_Display", "R2_mean_sd", "RMSE_mean_sd", "MAE_mean_sd",
        "N_Selected_Features_mean_sd", "Feature_Set", "Model", "R2_mean", "R2_sd",
        "RMSE_mean", "RMSE_sd", "MAE_mean", "MAE_sd",
        "N_Selected_Features_mean", "N_Selected_Features_sd"
    ]].rename(columns={
        "Predictor_Set": "Predictor set",
        "Model_Display": "Model",
        "R2_mean_sd": "R² mean ± SD",
        "RMSE_mean_sd": "RMSE mean ± SD",
        "MAE_mean_sd": "MAE mean ± SD",
        "N_Selected_Features_mean_sd": "Selected predictors mean ± SD",
        "Feature_Set": "Original feature-set name",
        "Model": "Original model name"
    })
    t1 = _round_numeric(t1, 4)
    final_tables["Table_01_nested_CV_model_comparison"] = _safe_table(t1, "Table_01_nested_CV_model_comparison")
except Exception as exc:
    print("Could not create Table 01:", exc)

# ------------------------------------------------------------
# Table 2. Best tuned model selected by nested-CV R²
# ------------------------------------------------------------
try:
    t2 = best_info.copy()
    t2["Best_Feature_Set_Display"] = t2["Best_Feature_Set"].apply(_display_feature_set)
    t2["Best_Model_Display"] = t2["Best_Model"].apply(_display_model)
    front_cols = ["Best_Feature_Set_Display", "Best_Model_Display"]
    remaining_cols = [c for c in t2.columns if c not in front_cols]
    t2 = t2[front_cols + remaining_cols].rename(columns={
        "Best_Feature_Set_Display": "Best predictor set",
        "Best_Model_Display": "Best model"
    })
    t2 = _round_numeric(t2, 4)
    final_tables["Table_02_best_model_selected_by_R2"] = _safe_table(t2, "Table_02_best_model_selected_by_R2")
except Exception as exc:
    print("Could not create Table 02:", exc)

# ------------------------------------------------------------
# Table 3. Best feature set by algorithm
# ------------------------------------------------------------
try:
    if "all_algorithm_best_info_df" in globals():
        t3 = all_algorithm_best_info_df.copy()
    elif "all_algorithm_best_info" in globals() and len(all_algorithm_best_info) > 0:
        t3 = pd.DataFrame(all_algorithm_best_info)
    else:
        t3 = pd.DataFrame()
    if not t3.empty:
        t3["Model_Display"] = t3["Model"].apply(_display_model)
        t3["Best_Feature_Set_Display"] = t3["Best_Feature_Set"].apply(_display_feature_set)
        t3["R2_mean_sd"] = [_fmt_mean_sd(m, s, 3) for m, s in zip(t3["Nested_CV_R2_mean"], t3["Nested_CV_R2_sd"])]
        t3["RMSE_mean_sd"] = [_fmt_mean_sd(m, s, 3) for m, s in zip(t3["Nested_CV_RMSE_mean"], t3["Nested_CV_RMSE_sd"])]
        t3["MAE_mean_sd"] = [_fmt_mean_sd(m, s, 3) for m, s in zip(t3["Nested_CV_MAE_mean"], t3["Nested_CV_MAE_sd"])]
        t3 = t3[[
            "Model_Display", "Best_Feature_Set_Display", "R2_mean_sd", "RMSE_mean_sd", "MAE_mean_sd",
            "N_Selected_Features", "Model", "Best_Feature_Set", "Nested_CV_R2_mean", "Nested_CV_R2_sd",
            "Nested_CV_RMSE_mean", "Nested_CV_RMSE_sd", "Nested_CV_MAE_mean", "Nested_CV_MAE_sd",
            "Final_Best_Params_JSON", "Selected_Features", "Saved_Pipeline"
        ]].rename(columns={
            "Model_Display": "Model",
            "Best_Feature_Set_Display": "Best predictor set",
            "R2_mean_sd": "R² mean ± SD",
            "RMSE_mean_sd": "RMSE mean ± SD",
            "MAE_mean_sd": "MAE mean ± SD",
            "N_Selected_Features": "Number of selected predictors",
            "Best_Feature_Set": "Original feature-set name",
            "Model": "Original model name"
        })
        t3 = _round_numeric(t3, 4)
    final_tables["Table_03_best_model_by_algorithm"] = _safe_table(t3, "Table_03_best_model_by_algorithm")
except Exception as exc:
    print("Could not create Table 03:", exc)

# ------------------------------------------------------------
# Table 4. SE added-value paired comparison
# ------------------------------------------------------------
try:
    t4 = added_value_df.copy()
    t4["Model_Display"] = t4["Model"].apply(_display_model)
    t4["Comparison_Display"] = t4["Comparison"].apply(_display_comparison)
    t4["Without_SE_Display"] = t4["Without_SE"].apply(_display_feature_set)
    t4["With_SE_Display"] = t4["With_SE"].apply(_display_feature_set)
    keep = [
        "Model_Display", "Comparison_Display", "Without_SE_Display", "With_SE_Display",
        "N_Paired_Outer_Splits", "R2_without_SE_mean", "R2_with_SE_mean", "Delta_R2_mean", "Delta_R2_sd", "Wilcoxon_p_R2",
        "RMSE_without_SE_mean", "RMSE_with_SE_mean", "Delta_RMSE_mean", "Delta_RMSE_sd",
        "RMSE_percent_improvement_mean", "Wilcoxon_p_RMSE",
        "MAE_without_SE_mean", "MAE_with_SE_mean", "Delta_MAE_mean", "Delta_MAE_sd",
        "MAE_percent_improvement_mean", "Wilcoxon_p_MAE"
    ]
    keep = [c for c in keep if c in t4.columns]
    t4 = t4[keep].rename(columns={
        "Model_Display": "Model",
        "Comparison_Display": "Comparison",
        "Without_SE_Display": "Without SE",
        "With_SE_Display": "With SE",
        "N_Paired_Outer_Splits": "Paired outer-CV splits",
        "RMSE_percent_improvement_mean": "RMSE improvement (%) mean",
        "MAE_percent_improvement_mean": "MAE improvement (%) mean"
    })
    t4 = _round_numeric(t4, 4)
    final_tables["Table_04_SE_added_value_paired_tests"] = _safe_table(t4, "Table_04_SE_added_value_paired_tests")
except Exception as exc:
    print("Could not create Table 04:", exc)

# ------------------------------------------------------------
# Table 5. Label-efficiency summary by training size
# ------------------------------------------------------------
try:
    t5 = label_eff_summary_df.copy()
    t5["Training data used (%)"] = (t5["Train_Proportion"] * 100).round(0).astype(int)
    t5["Predictor set"] = t5["Feature_Set"].apply(_display_feature_set)
    t5["Model display"] = t5["Model"].apply(_display_model)
    t5["R2_mean_sd"] = [_fmt_mean_sd(m, s, 3) for m, s in zip(t5["R2_mean"], t5["R2_sd"])]
    t5["RMSE_mean_sd"] = [_fmt_mean_sd(m, s, 3) for m, s in zip(t5["RMSE_mean"], t5["RMSE_sd"])]
    t5["MAE_mean_sd"] = [_fmt_mean_sd(m, s, 3) for m, s in zip(t5["MAE_mean"], t5["MAE_sd"])]
    t5 = t5[[
        "Training data used (%)", "Predictor set", "Model display", "Train_N_mean",
        "R2_mean_sd", "RMSE_mean_sd", "MAE_mean_sd", "Train_Proportion", "Feature_Set", "Model",
        "R2_mean", "R2_sd", "RMSE_mean", "RMSE_sd", "MAE_mean", "MAE_sd"
    ]].rename(columns={
        "Model display": "Model",
        "Train_N_mean": "Mean training N",
        "R2_mean_sd": "R² mean ± SD",
        "RMSE_mean_sd": "RMSE mean ± SD",
        "MAE_mean_sd": "MAE mean ± SD",
        "Feature_Set": "Original feature-set name",
    })
    t5 = _round_numeric(t5, 4)
    final_tables["Table_05_label_efficiency_summary"] = _safe_table(t5, "Table_05_label_efficiency_summary")
except Exception as exc:
    print("Could not create Table 05:", exc)

# ------------------------------------------------------------
# Table 6. SE gain by training size
# ------------------------------------------------------------
try:
    if "label_eff_added_df" in globals() and isinstance(label_eff_added_df, pd.DataFrame) and not label_eff_added_df.empty:
        t6 = label_eff_added_df.copy()
        t6["Training data used (%)"] = (t6["Train_Proportion"] * 100).round(0).astype(int)
        t6["Model display"] = t6["Model"].apply(_display_model)
        front = ["Training data used (%)", "Model display"]
        remaining = [c for c in t6.columns if c not in front]
        t6 = t6[front + remaining].rename(columns={"Model display": "Model"})
    elif "label_eff_gain_df" in globals() and isinstance(label_eff_gain_df, pd.DataFrame):
        t6 = label_eff_gain_df.copy()
        if "Train_Proportion" in t6.columns:
            t6["Training data used (%)"] = (t6["Train_Proportion"] * 100).round(0).astype(int)
        t6["Model display"] = t6["Model"].apply(_display_model)
    else:
        t6 = pd.DataFrame()
    t6 = _round_numeric(t6, 4)
    final_tables["Table_06_SE_gain_by_training_size"] = _safe_table(t6, "Table_06_SE_gain_by_training_size")
except Exception as exc:
    print("Could not create Table 06:", exc)

# ------------------------------------------------------------
# Table 7. SHAP predictor importance
# ------------------------------------------------------------
try:
    algo_shap_dir = OUT_DIR / "O4_SHAP_and_complementarity" / "O4_algorithm_specific_SHAP"

    if "all_algorithm_shap_df" in globals() and isinstance(all_algorithm_shap_df, pd.DataFrame):
        t7 = all_algorithm_shap_df.copy()
    elif (algo_shap_dir / "All_algorithm_specific_SHAP_predictor_importance.csv").exists():
        t7 = pd.read_csv(algo_shap_dir / "All_algorithm_specific_SHAP_predictor_importance.csv")
    else:
        t7 = shap_importance_df.copy()

    t7 = _round_numeric(t7, 6)
    final_tables["Table_07_SHAP_predictor_importance"] = _safe_table(
        t7,
        "Table_07_SHAP_predictor_importance"
    )

    if "overall_shap_df" in globals() and isinstance(overall_shap_df, pd.DataFrame):
        t7a = overall_shap_df.head(20).copy()
    elif (algo_shap_dir / "Table_07a_SHAP_top20_predictor_importance_across_algorithms.csv").exists():
        t7a = pd.read_csv(algo_shap_dir / "Table_07a_SHAP_top20_predictor_importance_across_algorithms.csv")
    elif (algo_shap_dir / "Overall_mean_SHAP_importance_across_algorithms.csv").exists():
        t7a = pd.read_csv(algo_shap_dir / "Overall_mean_SHAP_importance_across_algorithms.csv").head(20)
    else:
        t7a = t7.head(20).copy()

    t7a = _round_numeric(t7a, 6)
    final_tables["Table_07a_SHAP_top20_predictor_importance"] = _safe_table(
        t7a,
        "Table_07a_SHAP_top20_predictor_importance"
    )

except Exception as exc:
    print("Could not create Table 07:", exc)

# ------------------------------------------------------------
# Table 8. SHAP group importance
# ------------------------------------------------------------
try:
    algo_shap_dir = OUT_DIR / "O4_SHAP_and_complementarity" / "O4_algorithm_specific_SHAP"

    if "all_algorithm_group_shap_df" in globals() and isinstance(all_algorithm_group_shap_df, pd.DataFrame):
        t8 = all_algorithm_group_shap_df.copy()
    elif (algo_shap_dir / "All_algorithm_specific_SHAP_group_importance.csv").exists():
        t8 = pd.read_csv(algo_shap_dir / "All_algorithm_specific_SHAP_group_importance.csv")
    else:
        t8 = group_shap_df.copy()

    t8 = _round_numeric(t8, 6)
    final_tables["Table_08_SHAP_group_importance"] = _safe_table(
        t8,
        "Table_08_SHAP_group_importance"
    )

    if "overall_group_shap_df" in globals() and isinstance(overall_group_shap_df, pd.DataFrame):
        t8_overall = overall_group_shap_df.copy()
    elif (algo_shap_dir / "Overall_SHAP_group_importance_across_algorithms.csv").exists():
        t8_overall = pd.read_csv(algo_shap_dir / "Overall_SHAP_group_importance_across_algorithms.csv")
    else:
        t8_overall = pd.DataFrame()

    if not t8_overall.empty:
        t8_overall = _round_numeric(t8_overall, 6)
        final_tables["Table_08a_overall_SHAP_group_importance"] = _safe_table(
            t8_overall,
            "Table_08a_overall_SHAP_group_importance"
        )

except Exception as exc:
    print("Could not create Table 08:", exc)

# ------------------------------------------------------------
# Table 9. SHAP complementarity correlations
# ------------------------------------------------------------
try:
    t9 = complementarity_df.copy()
    t9 = _round_numeric(t9, 6)
    final_tables["Table_09_SHAP_complementarity_correlations"] = _safe_table(t9, "Table_09_SHAP_complementarity_correlations")
except Exception as exc:
    print("Could not create Table 09:", exc)

# ------------------------------------------------------------
# Table 10. SHAP interactions
# ------------------------------------------------------------
try:
    if "interaction_df" in globals() and isinstance(interaction_df, pd.DataFrame) and not interaction_df.empty:
        t10 = _round_numeric(interaction_df.copy(), 6)
    else:
        t10 = pd.DataFrame()
    final_tables["Table_10_SHAP_interactions"] = _safe_table(t10, "Table_10_SHAP_interactions")
except Exception as exc:
    print("Could not create Table 10:", exc)

# ------------------------------------------------------------
# Table 11. Plot-level SE-benefit correlations
# ------------------------------------------------------------
try:
    if "plot_corr_df" in globals() and isinstance(plot_corr_df, pd.DataFrame):
        t11 = plot_corr_df.copy()
    elif "condition_corr_df" in globals() and isinstance(condition_corr_df, pd.DataFrame):
        t11 = condition_corr_df.copy()
        if "Condition_Variable" in t11.columns:
            t11 = t11.rename(columns={"Condition_Variable": "Plot_Variable"})
    else:
        t11 = pd.DataFrame()

    if "Model" in t11.columns and not t11.empty:
        t11["Model display"] = t11["Model"].apply(_display_model)
        cols = ["Model display"] + [c for c in t11.columns if c != "Model display"]
        t11 = t11[cols].rename(columns={"Model display": "Model"})

    t11 = _round_numeric(t11, 6)
    final_tables["Table_11_plot_level_SE_benefit_correlations"] = _safe_table(
        t11,
        "Table_11_plot_level_SE_benefit_correlations"
    )
except Exception as exc:
    print("Could not create Table 11:", exc)

# ------------------------------------------------------------
# Table 12. Plot-variable bin summary
# ------------------------------------------------------------
try:
    if "bin_summary_df" in globals() and isinstance(bin_summary_df, pd.DataFrame) and not bin_summary_df.empty:
        t12 = bin_summary_df.copy()
        if "Condition_Variable" in t12.columns:
            t12 = t12.rename(columns={"Condition_Variable": "Plot_Variable"})
        if "Condition_Bin" in t12.columns:
            t12 = t12.rename(columns={"Condition_Bin": "Plot_Variable_Bin"})
        t12 = _round_numeric(t12, 5)
    else:
        t12 = pd.DataFrame()

    final_tables["Table_12_plot_variable_bin_summary"] = _safe_table(
        t12,
        "Table_12_plot_variable_bin_summary"
    )
except Exception as exc:
    print("Could not create Table 12:", exc)

try:
    if (
        "best_plot_variable_table" in globals()
        and isinstance(best_plot_variable_table, pd.DataFrame)
        and not best_plot_variable_table.empty
    ):
        t13 = best_plot_variable_table.copy()
    elif (
        "best_condition_table" in globals()
        and isinstance(best_condition_table, pd.DataFrame)
        and not best_condition_table.empty
    ):
        t13 = best_condition_table.copy()
        if "Condition_Variable" in t13.columns:
            t13 = t13.rename(columns={"Condition_Variable": "Plot_Variable"})
        if "Condition_Bin" in t13.columns:
            t13 = t13.rename(columns={"Condition_Bin": "Plot_Variable_Bin"})
    else:
        t13 = pd.DataFrame()

    t13 = _round_numeric(t13, 5)
    final_tables["Table_13_best_plot_level_SE_benefit"] = _safe_table(
        t13,
        "Table_13_best_plot_level_SE_benefit"
    )
except Exception as exc:
    print("Could not create Table 13:", exc)

# ------------------------------------------------------------
# Table 14. Mean repeated outer-test prediction metrics
# ------------------------------------------------------------
try:
    t14 = outer_prediction_metrics_df.copy()
    t14["Predictor set"] = t14["Feature_Set"].apply(_display_feature_set)
    t14["Model display"] = t14["Model"].apply(_display_model)
    front = ["Predictor set", "Model display"]
    t14 = t14[front + [c for c in t14.columns if c not in front]].rename(columns={"Model display": "Model"})
    t14 = _round_numeric(t14, 4)
    final_tables["Table_14_mean_repeated_outer_test_prediction_metrics"] = _safe_table(t14, "Table_14_mean_repeated_outer_test_prediction_metrics")
except Exception as exc:
    print("Could not create Table 14:", exc)

# ------------------------------------------------------------
# Combined final manuscript table workbook
# ------------------------------------------------------------
try:
    final_workbook = FINAL_TABLE_DIR / "FINAL_TABLES_AS_BEFORE.xlsx"
    with pd.ExcelWriter(final_workbook, engine="openpyxl") as writer:
        for sheet_name, table in final_tables.items():
            if table is None or not isinstance(table, pd.DataFrame) or table.empty:
                continue
            safe_sheet = _safe_filename(sheet_name, 31)[:31]
            table.to_excel(writer, sheet_name=safe_sheet, index=False)

        # Add a small table index sheet last/first by writing then moving it.
        index_table = pd.DataFrame([
            {"Sheet": _safe_filename(name, 31)[:31], "CSV file": f"{name}.csv", "Rows": 0 if table is None else len(table)}
            for name, table in final_tables.items()
            if table is not None and isinstance(table, pd.DataFrame) and not table.empty
        ])
        index_table.to_excel(writer, sheet_name="Table_index", index=False)

        # Basic formatting for readability.
        wb = writer.book
        if "Table_index" in wb.sheetnames:
            ws_index = wb["Table_index"]
            wb._sheets.remove(ws_index)
            wb._sheets.insert(0, ws_index)

        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)
        thin = Side(style="thin", color="DDDDDD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for col_cells in ws.columns:
                col_letter = col_cells[0].column_letter
                max_len = 0
                for cell in col_cells[:80]:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 45))

    # Also keep a copy at the run root for easy access.
    shutil.copy2(final_workbook, RUN_ROOT / "FINAL_TABLES_AS_BEFORE.xlsx")
    print("Final manuscript-style table workbook saved:", final_workbook)
except Exception as exc:
    print("Could not create FINAL_TABLES_AS_BEFORE.xlsx:", exc)
    traceback.print_exc(limit=1)

# Write a text summary of table outputs.
try:
    (FINAL_TABLE_DIR / "README_FINAL_TABLES.txt").write_text(
        "Final manuscript-style tables generated from the full reanalysis.\n"
        "Display names use MSI instead of RGBNIR for manuscript consistency, while original feature-set names are retained in raw columns where useful.\n\n"
        "Main file: FINAL_TABLES_AS_BEFORE.xlsx\n"
        "Each sheet is also exported as a separate CSV in this folder.\n",
        encoding="utf-8"
    )
except Exception:
    pass


try:
    _write_figure_selection_report()
except Exception as _fig_report_exc:
    print("Warning: could not write figure selection report:", _fig_report_exc)

# ============================================================
# FINAL COLLECTION: all tables, figures, logs, and a zip package
# ============================================================

def _safe_filename(text: str, max_len: int = 150) -> str:
    bad = '<>:"/\\|?*\n\r\t'
    out = ''.join('_' if ch in bad else ch for ch in str(text))
    out = out.replace(' ', '_')
    while '__' in out:
        out = out.replace('__', '_')
    return out[:max_len].strip('._') or 'file'


def _unique_path(folder: Path, name: str) -> Path:
    base = _safe_filename(Path(name).stem)
    suffix = Path(name).suffix
    candidate = folder / f"{base}{suffix}"
    i = 2
    while candidate.exists():
        candidate = folder / f"{base}_{i}{suffix}"
        i += 1
    return candidate


def _collect_outputs(run_root: Path):
    print("\nCollecting all generated tables and figures...")
    figures_dir = run_root / "FIGURES"
    tables_dir = run_root / "TABLES"
    files_dir = run_root / "FILES"
    figures_dir.mkdir(parents=True, exist_ok=True)
    tables_dir.mkdir(parents=True, exist_ok=True)
    files_dir.mkdir(parents=True, exist_ok=True)

    fig_exts = {'.png', '.jpg', '.jpeg', '.tif', '.tiff', '.svg', '.pdf'}
    table_exts = {'.csv', '.xlsx', '.xls'}
    other_exts = {'.json', '.txt', '.joblib', '.pkl'}
    index_rows = []

    skip_dirs = {figures_dir.resolve(), tables_dir.resolve(), files_dir.resolve()}

    for src in sorted(run_root.rglob('*')):
        if not src.is_file():
            continue
        if any(str(src.resolve()).startswith(str(d)) for d in skip_dirs):
            continue
        if src.name.startswith('ALL_RESULTS_PACKAGE'):
            continue
        ext = src.suffix.lower()
        if ext in fig_exts:
            dst = _unique_path(figures_dir, src.name)
            category = 'figure'
        elif ext in table_exts:
            dst = _unique_path(tables_dir, src.name)
            category = 'table'
        elif ext in other_exts:
            dst = _unique_path(files_dir, src.name)
            category = 'other_result_file'
        else:
            continue
        try:
            shutil.copy2(src, dst)
            index_rows.append({
                'category': category,
                'original_path': str(src),
                'collected_path': str(dst),
                'filename': dst.name,
                'extension': ext,
            })
        except Exception as exc:
            index_rows.append({
                'category': f'{category}_COPY_FAILED',
                'original_path': str(src),
                'collected_path': '',
                'filename': src.name,
                'extension': ext,
                'error': str(exc),
            })

    index_df = pd.DataFrame(index_rows)
    index_path = run_root / 'ALL_RESULTS_INDEX.csv'
    index_df.to_csv(index_path, index=False)

    # Combined workbook: all CSV files, plus sheets from XLSX files if readable.
    workbook_path = run_root / 'ALL_RESULTS_TABLES.xlsx'
    try:
        with pd.ExcelWriter(workbook_path, engine='openpyxl') as writer:
            used_sheets = set()
            for table_file in sorted(tables_dir.glob('*')):
                ext = table_file.suffix.lower()
                try:
                    if ext == '.csv':
                        df_tmp = pd.read_csv(table_file)
                        sheet = _safe_filename(table_file.stem, 25)
                        original_sheet = sheet
                        j = 2
                        while sheet in used_sheets:
                            sheet = _safe_filename(f'{original_sheet}_{j}', 31)
                            j += 1
                        used_sheets.add(sheet)
                        df_tmp.to_excel(writer, sheet_name=sheet[:31], index=False)
                    elif ext in {'.xlsx', '.xls'}:
                        xls = pd.ExcelFile(table_file)
                        for sheet_name in xls.sheet_names:
                            df_tmp = pd.read_excel(table_file, sheet_name=sheet_name)
                            sheet = _safe_filename(f'{table_file.stem}_{sheet_name}', 25)
                            original_sheet = sheet
                            j = 2
                            while sheet in used_sheets:
                                sheet = _safe_filename(f'{original_sheet}_{j}', 31)
                                j += 1
                            used_sheets.add(sheet)
                            df_tmp.to_excel(writer, sheet_name=sheet[:31], index=False)
                except Exception as exc:
                    pd.DataFrame([{'file': str(table_file), 'error': str(exc)}]).to_excel(
                        writer,
                        sheet_name=_safe_filename(f'ERROR_{table_file.stem}', 31)[:31],
                        index=False,
                    )
    except Exception as exc:
        (run_root / 'ALL_RESULTS_TABLES_WORKBOOK_ERROR.txt').write_text(
            f'Could not create ALL_RESULTS_TABLES.xlsx. Reason: {exc}\n',
            encoding='utf-8',
        )
        workbook_path = None

    # Simple HTML gallery for image formats that browsers can display.
    gallery_path = run_root / 'ALL_FIGURES_GALLERY.html'
    html_parts = [
        '<html><head><meta charset="utf-8"><title>All generated figures</title>',
        '<style>body{font-family:Arial,sans-serif;margin:24px;} .fig{margin:24px 0;padding-bottom:24px;border-bottom:1px solid #ddd;} img{max-width:1100px;width:100%;height:auto;border:1px solid #ddd;} code{background:#f6f6f6;padding:2px 4px;}</style>',
        '</head><body>',
        '<h1>All generated figures</h1>',
        f'<p>Run folder: <code>{_html.escape(str(run_root))}</code></p>',
    ]
    browser_exts = {'.png', '.jpg', '.jpeg', '.svg'}
    for fig in sorted(figures_dir.glob('*')):
        if fig.suffix.lower() in browser_exts:
            rel = fig.relative_to(run_root).as_posix()
            html_parts.append(
                f'<div class="fig"><h3>{_html.escape(fig.name)}</h3><img src="{_html.escape(rel)}"></div>'
            )
        else:
            rel = fig.relative_to(run_root).as_posix()
            html_parts.append(f'<div class="fig"><h3>{_html.escape(fig.name)}</h3><p><a href="{_html.escape(rel)}">Open file</a></p></div>')
    html_parts.append('</body></html>')
    gallery_path.write_text('\n'.join(html_parts), encoding='utf-8')

    # README / manifest
    readme = run_root / 'README_RUN_OUTPUTS.txt'
    readme.write_text(
        'Full analysis rerun without predictor extraction\n'
        f'Input predictor CSV: {CSV_PATH_FROM_ARGS}\n'
        f'Run folder: {run_root}\n\n'
        'Main folders:\n'
        f'- {OUT_DIR}: actual analysis outputs by objective\n'
        f'- {figures_dir}: all copied figure files together\n'
        f'- {tables_dir}: all copied table files together\n'
        f'- {files_dir}: models, JSON, text logs, and other result files\n\n'
        'Main combined files:\n'
        f'- {index_path}\n'
        f'- {workbook_path if workbook_path else "ALL_RESULTS_TABLES.xlsx was not created"}\n'
        f'- {gallery_path}\n',
        encoding='utf-8',
    )

    zip_path = run_root / f'ALL_RESULTS_PACKAGE_{run_root.name}.zip'
    try:
        with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
            for f in run_root.rglob('*'):
                if f == zip_path or not f.is_file():
                    continue
                zf.write(f, f.relative_to(run_root))
    except Exception as exc:
        (run_root / 'ZIP_PACKAGE_ERROR.txt').write_text(str(exc), encoding='utf-8')
        zip_path = None

    print('\nDONE. New full-analysis outputs are here:')
    print(run_root)
    print('\nMain collected folders:')
    print(' -', figures_dir)
    print(' -', tables_dir)
    print(' -', files_dir)
    print('\nMain files:')
    print(' -', index_path)
    if workbook_path:
        print(' -', workbook_path)
    print(' -', gallery_path)
    if zip_path:
        print(' -', zip_path)


try:
    _collect_outputs(RUN_ROOT)
except Exception as _collect_exc:
    print('Warning: final collection step failed:', _collect_exc)
    traceback.print_exc()
